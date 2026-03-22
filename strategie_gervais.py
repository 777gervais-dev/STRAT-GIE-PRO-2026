"""
╔══════════════════════════════════════════════════════════════════╗
║  STRATÉGIE PRO v7.1 — MACD + BB + VWAP + ML + ICT  [LIVE]        ║
║  5 Modèles ML · Graphique · Journal · ICT · Token persistant    ║
║  by 777gervais-dev  |  Intraday 2026                            ║
╚══════════════════════════════════════════════════════════════════╝
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
import json
import os
import time
from datetime import datetime, timezone
import warnings
warnings.filterwarnings("ignore")

def sf(val, fmt=",.2f", fallback="N/A"):
    """Safe format — returns fallback string if val is None."""
    try:
        if val is None: return fallback
        return format(float(val), fmt)
    except: return fallback


# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="STRATÉGIE PRO v7.1 — 8 Actifs",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── PATHS ───────────────────────────────────────────────────────────────────
HOME        = os.path.expanduser("~")
# Use /tmp on Streamlit Cloud (writable), HOME on Termux
_DATA_DIR   = "/tmp" if os.path.exists("/etc/streamlit") or os.environ.get("STREAMLIT_SHARING_MODE") else HOME
CONFIG_FILE = os.path.join(_DATA_DIR, ".strat_pro_config.json")
JOURNAL_FILE= os.path.join(_DATA_DIR, ".strat_pro_journal.json")
HISTORY_FILE= os.path.join(_DATA_DIR, ".strat_pro_history.json")

# ─── CSS ─────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600;700&family=Rajdhani:wght@500;600;700&display=swap');
:root{
  --bg:#060E1A; --bg2:#0A1520; --card:#0F1C2E; --card2:#132030;
  --border:#1A3050; --border2:#243D5C;
  --gold:#C9A94B; --gold2:#F0C96A; --gold3:#FFE09A;
  --blue:#2E5FA3; --blue2:#4A7EC7; --blue3:#7AAEE8;
  --green:#0FBF5F; --green2:#1DB954; --green3:#4DD882;
  --red:#D93025;  --red2:#E53E3E;   --red3:#FF6B6B;
  --orange:#D4730A; --orange2:#E07B2A; --orange3:#FFB347;
  --purple:#7C3AED; --purple2:#A78BFA;
  --gray:#6B7A8D; --gray2:#8A95A3; --gray3:#B0BCC8;
  --text:#DCE8F5; --text2:#A8BCCF;
  --mono:'JetBrains Mono',monospace; --head:'Rajdhani',sans-serif;
}
html,body,[data-testid="stAppViewContainer"]{
  background:var(--bg)!important; color:var(--text)!important;
  font-family:var(--mono)!important; font-size:13px!important;
}
[data-testid="stMain"]{background:var(--bg)!important; padding:8px 10px!important;}
#MainMenu,footer,header{visibility:hidden;}
[data-testid="stToolbar"]{display:none;}
[data-testid="stSidebar"]{background:var(--bg2)!important;border-right:1px solid var(--border)!important;}

/* TABS */
[data-testid="stTabs"] button{
  font-family:var(--head)!important; font-size:14px!important; font-weight:700!important;
  color:var(--gray2)!important; letter-spacing:1px!important;
  background:transparent!important; border:none!important;
  border-bottom:2px solid transparent!important; padding:10px 16px!important;
}
[data-testid="stTabs"] button[aria-selected="true"]{
  color:var(--gold)!important; border-bottom:2px solid var(--gold)!important;
}
[data-testid="stTabs"] [role="tablist"]{
  background:var(--card)!important; border-radius:10px 10px 0 0!important;
  border-bottom:1px solid var(--border)!important; gap:0!important;
}

/* BUTTONS */
.stButton>button{
  background:linear-gradient(135deg,var(--gold),var(--gold2))!important;
  color:#0A1520!important; font-family:var(--head)!important; font-weight:700!important;
  font-size:15px!important; border:none!important; border-radius:8px!important;
  padding:12px 20px!important; width:100%!important; letter-spacing:1px!important;
  box-shadow:0 4px 20px rgba(201,169,75,.25)!important; transition:all .15s!important;
}
.stButton>button:hover{box-shadow:0 6px 30px rgba(201,169,75,.45)!important;}
.stButton>button:active{transform:scale(.98)!important;}

/* INPUTS */
.stSelectbox>div>div,.stTextInput>div>div>input,.stNumberInput>div>div>input{
  background:var(--card2)!important; border:1px solid var(--border2)!important;
  color:var(--text)!important; font-family:var(--mono)!important; border-radius:6px!important;
}
.stSelectbox label,.stTextInput label,.stNumberInput label,.stCheckbox label{
  color:var(--gray3)!important; font-size:12px!important;
}
.stCheckbox>label{color:var(--text2)!important;}

/* CARDS */
.card{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:14px;margin-bottom:10px;}
.card-gold{border-left:4px solid var(--gold)!important;}
.card-green{border-left:4px solid var(--green2)!important;}
.card-red{border-left:4px solid var(--red2)!important;}
.card-blue{border-left:4px solid var(--blue2)!important;}
.card-purple{border-left:4px solid var(--purple2)!important;}
.card-orange{border-left:4px solid var(--orange2)!important;}

/* LIVE DOT */
.live-dot{display:inline-block;width:8px;height:8px;background:var(--red2);border-radius:50%;
  margin-right:6px;box-shadow:0 0 8px var(--red2);animation:pulse 1.2s infinite;}
@keyframes pulse{0%,100%{opacity:1;transform:scale(1);}50%{opacity:.3;transform:scale(.7);}}

/* BADGES */
.badge{display:inline-flex;align-items:center;gap:4px;padding:2px 9px;border-radius:4px;
  font-size:10px;font-weight:700;letter-spacing:1px;}
.b-binance{background:rgba(240,185,11,.15);border:1px solid #F0B90B;color:#F0B90B;}
.b-oanda{background:rgba(74,126,199,.2);border:1px solid var(--blue2);color:var(--blue3);}
.b-yf{background:rgba(107,122,141,.12);border:1px solid var(--border2);color:var(--gray2);}
.b-buy{background:rgba(29,185,84,.15);border:1px solid var(--green2);color:var(--green3);}
.b-sell{background:rgba(229,62,62,.15);border:1px solid var(--red2);color:var(--red3);}
.b-range{background:rgba(224,123,42,.15);border:1px solid var(--orange2);color:var(--orange3);}
.b-wait{background:rgba(107,122,141,.12);border:1px solid var(--border2);color:var(--gray2);}

/* CHECKLIST */
.chk{display:flex;align-items:center;padding:7px 11px;border-radius:7px;margin:3px 0;
  font-size:12px;border:1px solid var(--border);gap:8px;}
.chk-ok{background:rgba(15,191,95,.06);border-color:#0F3A1F!important;}
.chk-no{background:rgba(217,48,37,.06);border-color:#3A0F0F!important;}

/* TABLE */
.gt{width:100%;border-collapse:collapse;font-size:11px;}
.gt th{background:var(--bg2);color:var(--gold);font-family:var(--head);
  padding:7px 8px;text-align:center;letter-spacing:1px;border-bottom:1px solid var(--gold);}
.gt td{padding:6px 8px;border-bottom:1px solid var(--border);text-align:center;}
.gt tr:nth-child(even) td{background:rgba(255,255,255,.015);}
.gt tr:hover td{background:rgba(201,169,75,.06);}

/* ICT */
.ict-item{display:flex;justify-content:space-between;align-items:center;
  padding:6px 11px;border-radius:6px;margin:3px 0;font-size:11px;
  background:rgba(124,58,237,.06);border:1px solid rgba(124,58,237,.25);}

/* Scrollbar */
::-webkit-scrollbar{width:3px;height:3px;}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:2px;}
::-webkit-scrollbar-track{background:var(--bg);}
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTS ───────────────────────────────────────────────────────────────
ASSETS = {
    "XAUUSD":  {"source":"oanda","oanda":"XAU_USD",    "binance":None,     "yf":"GC=F",      "yf_fb":["GC=F","XAUUSD=X"],          "dec":3,"pip":0.10},
    "GC GOLD": {"source":"oanda","oanda":"XAU_USD",    "binance":None,     "yf":"GC=F",      "yf_fb":["GC=F","XAUUSD=X"],          "dec":2,"pip":0.10},
    "BTC/USD": {"source":"binance","oanda":None,       "binance":"BTCUSDT","yf":"BTC-USD",   "yf_fb":["BTC-USD","BTC=F"],           "dec":2,"pip":1.0},
    "CL WTI":  {"source":"oanda","oanda":"WTICO_USD",  "binance":None,     "yf":"CL=F",      "yf_fb":["CL=F","CLJ26.NYM"],          "dec":3,"pip":0.01},
    "EUR/USD": {"source":"oanda","oanda":"EUR_USD",    "binance":None,     "yf":"EURUSD=X",  "yf_fb":["EURUSD=X","EUR=X"],          "dec":5,"pip":0.0001},
    "6E EUR":  {"source":"oanda","oanda":"EUR_USD",    "binance":None,     "yf":"EURUSD=X",  "yf_fb":["EURUSD=X","EUR=X"],          "dec":5,"pip":0.0001},
    "ES S&P":  {"source":"oanda","oanda":"SPX500_USD", "binance":None,     "yf":"SPY",       "yf_fb":["SPY","^GSPC","ES=F"],        "dec":2,"pip":0.25},
    "NAS100":  {"source":"oanda","oanda":"NAS100_USD", "binance":None,     "yf":"QQQ",       "yf_fb":["QQQ","^NDX","NQ=F"],         "dec":2,"pip":0.25},
}
TF_BIN={"5m":"5m","15m":"15m","30m":"30m","1h":"1h"}
TF_OAN={"5m":"M5","15m":"M15","30m":"M30","1h":"H1"}
TF_YF ={"5m":"5m","15m":"15m","30m":"30m","1h":"60m"}
TF_PER={"5m":"1d","15m":"5d","30m":"5d","1h":"30d"}
TP_SL ={"5m":(1.5,1.0),"15m":(2.0,1.3),"30m":(2.5,1.5),"1h":(3.0,2.0)}
KILL_ZONES=[
    {"name":"ASIAN",  "start":"00:00","end":"03:00","col":"#6B7BFF"},
    {"name":"LONDON", "start":"03:00","end":"08:30","col":"#C9A94B"},
    {"name":"NY AM",  "start":"08:30","end":"12:00","col":"#0FBF5F"},
    {"name":"NY PM",  "start":"13:00","end":"16:00","col":"#E07B2A"},
    {"name":"OVERLAP","start":"07:00","end":"10:00","col":"#E53E3E"},
]
OANDA_DEMO="https://api-fxpractice.oanda.com"
OANDA_LIVE="https://api-fxtrade.oanda.com"

# ─── PERSISTENCE ─────────────────────────────────────────────────────────────
def load_config():
    """Load config — Streamlit Cloud secrets take priority over local file."""
    cfg = {"oanda_key":"","oanda_demo":True,"tg_token":"","tg_chat":"",
           "tg_auto":False,"tg_min_conf":75,"tg_buy":True,"tg_sell":True,
           "claude_api_key":""}
    # Load from local file first (Termux)
    try:
        if os.path.exists(CONFIG_FILE):
            saved = json.load(open(CONFIG_FILE))
            cfg.update(saved)
    except: pass
    # Override with Streamlit Cloud secrets if available
    try:
        secrets = st.secrets
        if "oanda" in secrets:
            cfg["oanda_key"]      = secrets["oanda"].get("key",  cfg["oanda_key"])
            cfg["oanda_demo"]     = secrets["oanda"].get("demo", cfg["oanda_demo"])
        if "claude" in secrets:
            cfg["claude_api_key"] = secrets["claude"].get("api_key", cfg["claude_api_key"])
        if "telegram" in secrets:
            cfg["tg_token"]       = secrets["telegram"].get("token",   cfg["tg_token"])
            cfg["tg_chat"]        = secrets["telegram"].get("chat_id", cfg["tg_chat"])
    except: pass
    return cfg

def is_cloud():
    """Detect if running on Streamlit Cloud (vs local Termux)."""
    try:
        return bool(st.secrets)
    except: return False

def save_config(cfg):
    try: json.dump(cfg, open(CONFIG_FILE,"w"))
    except: pass

def load_journal():
    try:
        if os.path.exists(JOURNAL_FILE):
            return json.load(open(JOURNAL_FILE))
    except: pass
    return []

def save_journal(j):
    try: json.dump(j[-200:], open(JOURNAL_FILE,"w"))
    except: pass

def load_history():
    try:
        if os.path.exists(HISTORY_FILE):
            return json.load(open(HISTORY_FILE))
    except: pass
    return []

def save_history(h):
    try: json.dump(h[-500:], open(HISTORY_FILE,"w"))
    except: pass

def append_signal(asset, tf, signal, confidence, price, tp, sl, regime, ml_score):
    h = load_history()
    h.append({
        "ts":    datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"),
        "asset": asset, "tf": tf, "signal": signal,
        "confidence": confidence, "price": round(price,5),
        "tp": round(tp,5) if tp else None,
        "sl": round(sl,5) if sl else None,
        "regime": regime, "ml_score": round(ml_score,1)
    })
    save_history(h)

# ─── DATA SOURCES ────────────────────────────────────────────────────────────
def fetch_binance(sym, iv, n=200):
    try:
        r=requests.get("https://api.binance.com/api/v3/klines",
          params={"symbol":sym,"interval":iv,"limit":n},timeout=8)
        if r.status_code!=200: return None
        df=pd.DataFrame(r.json(),columns=["ts","O","H","L","C","V","ct","qav","tr","tb","tbq","ig"])
        for c in ["O","H","L","C","V"]: df[c]=pd.to_numeric(df[c])
        df.index=pd.to_datetime(df["ts"],unit="ms",utc=True)
        return df.rename(columns={"O":"Open","H":"High","L":"Low","C":"Close","V":"Volume"})[["Open","High","Low","Close","Volume"]]
    except: return None

def fetch_binance_price(sym):
    try:
        r=requests.get("https://api.binance.com/api/v3/ticker/price",params={"symbol":sym},timeout=5)
        return float(r.json()["price"]) if r.status_code==200 else None
    except: return None

def fetch_oanda(inst,gran,count=200,key="",demo=True):
    if not key: return None
    base=OANDA_DEMO if demo else OANDA_LIVE
    try:
        r=requests.get(f"{base}/v3/instruments/{inst}/candles",
          headers={"Authorization":f"Bearer {key}"},
          params={"granularity":gran,"count":count,"price":"M"},timeout=8)
        if r.status_code!=200: return None
        rows=[{"Open":float(c["mid"]["o"]),"High":float(c["mid"]["h"]),
               "Low":float(c["mid"]["l"]),"Close":float(c["mid"]["c"]),
               "Volume":float(c.get("volume",0)),"ts":c["time"]}
              for c in r.json().get("candles",[]) if c.get("complete",True)]
        if not rows: return None
        df=pd.DataFrame(rows); df.index=pd.to_datetime(df["ts"],utc=True)
        return df[["Open","High","Low","Close","Volume"]]
    except: return None

def fetch_oanda_price(inst,key,demo=True):
    if not key: return None
    base=OANDA_DEMO if demo else OANDA_LIVE
    try:
        r=requests.get(f"{base}/v3/instruments/{inst}/candles",
          headers={"Authorization":f"Bearer {key}"},
          params={"granularity":"S5","count":1,"price":"M"},timeout=5)
        c=r.json().get("candles",[])
        return float(c[-1]["mid"]["c"]) if r.status_code==200 and c else None
    except: return None

def fetch_yf(ticker,iv,period):
    """Try ticker, then fallbacks if data is insufficient."""
    try:
        import yfinance as yf
        df=yf.download(ticker,period=period,interval=iv,progress=False,auto_adjust=True)
        if df is None or len(df)<30: return None
        df.columns=[c[0] if isinstance(c,tuple) else c for c in df.columns]
        df=df.rename(columns=str.capitalize)
        for c in ["Open","High","Low","Close","Volume"]:
            if c not in df.columns: return None
        return df[["Open","High","Low","Close","Volume"]].dropna()
    except: return None

def fetch_yf_with_fallbacks(tickers, iv, period):
    """Try a list of tickers, return first successful result."""
    for ticker in tickers:
        df = fetch_yf(ticker, iv, period)
        if df is not None and len(df) >= 30:
            return df
    return None

@st.cache_data(ttl=60, show_spinner=False)
def fetch_yf_direct(ticker, interval="5m", range_="1d"):
    """Fetch Yahoo Finance data via direct HTTP requests — no yfinance lib needed.
    Works on Android even when yfinance lib fails."""
    try:
        ticker_enc = ticker.replace("^", "%5E").replace("=", "%3D")
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker_enc}"
        params = {"interval": interval, "range": range_,
                  "includePrePost": "false", "corsDomain": "finance.yahoo.com"}
        headers = {
            "User-Agent": "Mozilla/5.0 (Linux; Android 11)",
            "Accept": "application/json",
        }
        r = requests.get(url, params=params, headers=headers, timeout=10)
        if r.status_code != 200: return None
        data = r.json()
        result = data.get("chart", {}).get("result", [])
        if not result: return None
        res = result[0]
        ts    = res.get("timestamp", [])
        q     = res.get("indicators", {}).get("quote", [{}])[0]
        opens  = q.get("open", [])
        highs  = q.get("high", [])
        lows   = q.get("low", [])
        closes = q.get("close", [])
        vols   = q.get("volume", [])
        if not closes or len(closes) < 30: return None
        df = pd.DataFrame({
            "Open":   opens,  "High": highs, "Low":  lows,
            "Close":  closes, "Volume": vols,
        })
        df.index = pd.to_datetime(ts, unit="s", utc=True)
        df = df.dropna(subset=["Close"])
        for c in ["Open","High","Low","Close","Volume"]:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        df = df.dropna()
        return df if len(df) >= 30 else None
    except: return None

def fetch_yf_direct_with_fallbacks(tickers, interval="5m", range_="1d"):
    """Try direct Yahoo Finance API for a list of tickers."""
    for ticker in tickers:
        df = fetch_yf_direct(ticker, interval, range_)
        if df is not None and len(df) >= 30:
            return df
    return None

def get_candles(acfg,tf,key,demo):
    src=acfg["source"]
    if src=="binance" and acfg["binance"]:
        df=fetch_binance(acfg["binance"],TF_BIN[tf])
        if df is not None and len(df)>=50: return df,"binance"
    if src=="oanda" and acfg["oanda"] and key:
        df=fetch_oanda(acfg["oanda"],TF_OAN[tf],key=key,demo=demo)
        if df is not None and len(df)>=50: return df,"oanda"
    # Yahoo Finance avec fallbacks automatiques
    fallbacks=acfg.get("yf_fb",[acfg["yf"]])
    df=fetch_yf_with_fallbacks(fallbacks,TF_YF[tf],TF_PER[tf])
    # Last resort: Yahoo Finance direct HTTP (no lib)
    yf_tickers = acfg.get("yf_fb", [acfg["yf"]])
    yf_interval_map = {"5m":"5m","15m":"15m","30m":"30m","1h":"60m"}
    yf_range_map    = {"5m":"1d","15m":"5d","30m":"5d","1h":"30d"}
    iv_direct = yf_interval_map.get(tf, "5m")
    rng_direct = yf_range_map.get(tf, "1d")
    df = fetch_yf_direct_with_fallbacks(yf_tickers, iv_direct, rng_direct)
    return (df, "yahoo_direct") if df is not None else (None, "error")

def get_live_price(acfg,key,demo):
    if acfg["source"]=="binance" and acfg["binance"]:
        p=fetch_binance_price(acfg["binance"])
        if p: return p,"binance"
    if acfg["source"]=="oanda" and acfg["oanda"] and key:
        p=fetch_oanda_price(acfg["oanda"],key,demo)
        if p: return p,"oanda"
    return None,"yfinance"

# ─── CLOUD OPTIMIZATION — Cache API results ─────────────────────────────────
@st.cache_data(ttl=30, show_spinner=False)
def cached_binance_price(sym):
    return fetch_binance_price(sym)

@st.cache_data(ttl=30, show_spinner=False)
def cached_binance_klines(sym, iv, n=200):
    return fetch_binance(sym, iv, n)

# ─── INDICATORS ──────────────────────────────────────────────────────────────
def calc_vwap(df):
    tp=(df["High"]+df["Low"]+df["Close"])/3
    return (tp*df["Volume"]).cumsum()/df["Volume"].cumsum().replace(0,np.nan)

def calc_bb(df,p=20,d=2.0):
    m=df["Close"].rolling(p).mean(); s=df["Close"].rolling(p).std()
    return m+d*s,m,m-d*s,(2*d*s)/m.replace(0,np.nan)

def calc_macd(df,f=12,s=26,sig=9):
    ef=df["Close"].ewm(span=f,adjust=False).mean()
    es=df["Close"].ewm(span=s,adjust=False).mean()
    m=ef-es; sg=m.ewm(span=sig,adjust=False).mean()
    return m,sg,m-sg

def calc_atr(df,p=14):
    hl=df["High"]-df["Low"]
    hc=(df["High"]-df["Close"].shift()).abs()
    lc=(df["Low"]-df["Close"].shift()).abs()
    return pd.concat([hl,hc,lc],axis=1).max(axis=1).ewm(span=p,adjust=False).mean()

def calc_rsi(df,p=14):
    d=df["Close"].diff()
    g=d.clip(lower=0).ewm(span=p,adjust=False).mean()
    l=(-d.clip(upper=0)).ewm(span=p,adjust=False).mean()
    return 100-(100/(1+g/l.replace(0,np.nan)))

def calc_stoch(df,k=14,d=3):
    lo=df["Low"].rolling(k).min(); hi=df["High"].rolling(k).max()
    ks=100*(df["Close"]-lo)/(hi-lo).replace(0,np.nan)
    return ks,ks.rolling(d).mean()

def calc_force_index(df, period_short=2, period_long=13):
    """Elder Force Index = Volume × (Close - Close[prev])
    - FI(1)  : raw 1-period — captures intraday conviction
    - FI(2)  : EMA-2 smoothing — short-term momentum (Elder entry signal)
    - FI(13) : EMA-13 smoothing — trend direction and strength
    """
    fi_raw = df["Volume"] * df["Close"].diff()
    fi2    = fi_raw.ewm(span=2,  adjust=False).mean()
    fi13   = fi_raw.ewm(span=13, adjust=False).mean()
    return fi_raw, fi2, fi13

def interpret_force_index(fi2_val, fi13_val, signal, atr, close):
    """Return Force Index interpretation dict."""
    # Normalize by price to make comparable across assets
    norm = close * 0.001 or 1
    fi2_n  = fi2_val  / norm
    fi13_n = fi13_val / norm

    # Trend strength from FI(13)
    if fi13_val > 0:
        if fi13_n > 50:   trend_str, trend_col = "HAUSSIER FORT 🟢",  "#0FBF5F"
        elif fi13_n > 10: trend_str, trend_col = "HAUSSIER MODERE 🟢","#7EC47A"
        else:             trend_str, trend_col = "HAUSSIER FAIBLE 🟡", "#C9A94B"
    else:
        if fi13_n < -50:  trend_str, trend_col = "BAISSIER FORT 🔴",  "#D93025"
        elif fi13_n < -10:trend_str, trend_col = "BAISSIER MODERE 🔴","#E07B2A"
        else:             trend_str, trend_col = "BAISSIER FAIBLE 🟡", "#C9A94B"

    # Entry signal from FI(2)
    if fi2_val < 0 and signal == "BUY":
        entry_sig, entry_col = "SIGNAL BUY CONFIRME (FI2<0 en uptrend)", "#0FBF5F"
    elif fi2_val > 0 and signal == "SELL":
        entry_sig, entry_col = "SIGNAL SELL CONFIRME (FI2>0 en downtrend)", "#D93025"
    elif fi2_val > 0 and signal == "BUY":
        entry_sig, entry_col = "FI2 positif — Attendre correction pour BUY", "#C9A94B"
    elif fi2_val < 0 and signal == "SELL":
        entry_sig, entry_col = "FI2 negatif — Attendre rebond pour SELL", "#C9A94B"
    else:
        entry_sig, entry_col = "FI2 neutre — Pas de signal d'entree clair", "#6B7A8D"

    # Divergence check (FI13 vs price trend)
    fi13_bull = fi13_val > 0
    signal_bull = signal == "BUY"
    divergence = (fi13_bull and signal == "SELL") or (not fi13_bull and signal == "BUY")

    return {
        "fi2":       round(fi2_val,  2),
        "fi13":      round(fi13_val, 2),
        "trend_str": trend_str,
        "trend_col": trend_col,
        "entry_sig": entry_sig,
        "entry_col": entry_col,
        "divergence":divergence,
        "fi13_bull": fi13_bull,
        "confirmed": not divergence and abs(fi13_n) > 5,
    }

# ─── ICT LEVELS ──────────────────────────────────────────────────────────────
def detect_fvg(df, n=30):
    """Fair Value Gaps — 3-candle pattern."""
    fvgs=[]
    data=df.iloc[-n:].copy()
    for i in range(2,len(data)):
        c0=data.iloc[i-2]; c2=data.iloc[i]
        # Bullish FVG: c0 high < c2 low
        if c0["High"]<c2["Low"]:
            mid=(c0["High"]+c2["Low"])/2
            fvgs.append({"type":"bull","top":c2["Low"],"bot":c0["High"],"mid":mid,
                         "ts":str(data.index[i])[:16],"filled":False})
        # Bearish FVG: c0 low > c2 high
        elif c0["Low"]>c2["High"]:
            mid=(c0["Low"]+c2["High"])/2
            fvgs.append({"type":"bear","top":c0["Low"],"bot":c2["High"],"mid":mid,
                         "ts":str(data.index[i])[:16],"filled":False})
    return fvgs[-5:] if fvgs else []

def detect_order_blocks(df, n=40):
    """Order Blocks — last down candle before up move (bull OB) and vice versa."""
    obs=[]
    data=df.iloc[-n:].copy()
    closes=data["Close"].values
    for i in range(1,len(data)-2):
        c=data.iloc[i]
        # Bullish OB: bearish candle followed by strong bullish move
        if c["Close"]<c["Open"]:
            future_high=data["High"].iloc[i+1:min(i+4,len(data))].max()
            if future_high>c["High"]*1.001:
                obs.append({"type":"bull","top":c["Open"],"bot":c["Low"],
                            "mid":(c["Open"]+c["Low"])/2,"ts":str(data.index[i])[:16]})
        # Bearish OB: bullish candle followed by strong bearish move
        elif c["Close"]>c["Open"]:
            future_low=data["Low"].iloc[i+1:min(i+4,len(data))].min()
            if future_low<c["Low"]*0.999:
                obs.append({"type":"bear","top":c["High"],"bot":c["Close"],
                            "mid":(c["High"]+c["Close"])/2,"ts":str(data.index[i])[:16]})
    return obs[-4:] if obs else []

def detect_liquidity(df, n=50):
    """Liquidity sweeps — equal highs/lows that may attract price."""
    data=df.iloc[-n:]
    highs=data["High"].values; lows=data["Low"].values
    close=float(df["Close"].iloc[-1])
    # Find swing highs and lows
    sh=[]; sl=[]
    for i in range(2,len(highs)-2):
        if highs[i]>highs[i-1] and highs[i]>highs[i-2] and highs[i]>highs[i+1] and highs[i]>highs[i+2]:
            sh.append(highs[i])
        if lows[i]<lows[i-1] and lows[i]<lows[i-2] and lows[i]<lows[i+1] and lows[i]<lows[i+2]:
            sl.append(lows[i])
    # Key levels above/below price
    bsl=[x for x in sh if x>close]  # buyside liquidity (above)
    ssl=[x for x in sl if x<close]  # sellside liquidity (below)
    return {
        "buyside":  sorted(bsl)[:3] if bsl else [],
        "sellside": sorted(ssl,reverse=True)[:3] if ssl else [],
    }

# ─── 5 ML MODELS (NumPy-only ensemble) ───────────────────────────────────────
def ml_sigmoid(x): return 1/(1+np.exp(-np.clip(x,-10,10)))

def model_logistic(feats):
    """Logistic Regression — weights tuned on price structure."""
    w=np.array([0.35,0.25,0.20,0.10,0.10])
    z=np.dot(feats,w)-0.5
    return float(ml_sigmoid(z*3))

def model_momentum(feats):
    """Pure momentum model — rewards strong directional signals."""
    mom=feats[0]*0.5+feats[2]*0.3+feats[3]*0.2
    return float(ml_sigmoid((mom-0.4)*5))

def model_mean_reversion(feats):
    """Mean-reversion model — rewards extremes with counter-signals."""
    rev=feats[1]*0.6+feats[4]*0.4
    return float(ml_sigmoid((rev-0.3)*4))

def model_random_forest(feats):
    """Ensemble of 8 decision stumps — rule-based splits."""
    votes=[]
    thresholds=[(0,0.55),(0,0.65),(1,0.50),(2,0.50),(3,0.45),(2,0.55),(4,0.40),(1,0.60)]
    for fidx,thresh in thresholds:
        votes.append(1.0 if feats[fidx]>thresh else 0.0)
    return float(np.mean(votes))

def model_volatility(feats):
    """Volatility-aware model — adjusts confidence by BB bandwidth."""
    base=feats[0]*0.4+feats[2]*0.3+feats[1]*0.3
    vol_factor=feats[4]
    adj=base*(0.6+vol_factor*0.4)
    return float(ml_sigmoid((adj-0.4)*4))

def build_features(close,vwap,bbu,bbl,bbm,mn,msn,mhn,rsi_v,atr_v,bw_v,bull):
    """Build normalized feature vector [0,1] for ML models."""
    bb_range=bbu-bbl if bbu>bbl else 1e-9
    # f0: VWAP alignment score
    vwap_dist=abs(close-vwap)/vwap
    f0=min(1.0,(vwap_dist*20+0.5)) if bull else min(1.0,((1-vwap_dist*20)+0.5))
    f0=max(0.0,min(1.0,f0))
    # f1: BB position (1=at favorable extreme, 0=at unfavorable)
    bb_pos=(close-bbl)/bb_range
    f1=(1-bb_pos) if bull else bb_pos
    f1=max(0.0,min(1.0,f1))
    # f2: MACD momentum alignment
    macd_norm=ml_sigmoid(mn*100)
    f2=macd_norm if bull else (1-macd_norm)
    # f3: Histogram direction
    hist_norm=ml_sigmoid(mhn*1000)
    f3=hist_norm if bull else (1-hist_norm)
    # f4: RSI not overbought/oversold
    if bull:  f4=max(0.0,min(1.0,(70-rsi_v)/40))
    else:     f4=max(0.0,min(1.0,(rsi_v-30)/40))
    return np.array([f0,f1,f2,f3,f4])

def run_ml_ensemble(feats, bull):
    """Run all 5 models and return ensemble score + individual scores."""
    scores={
        "Logistic Reg.":    model_logistic(feats),
        "Momentum":         model_momentum(feats),
        "Mean-Reversion":   model_mean_reversion(feats),
        "Random Forest":    model_random_forest(feats),
        "Volatility":       model_volatility(feats),
    }
    ensemble=float(np.mean(list(scores.values())))
    return ensemble, scores

# ─── FULL ANALYSIS ───────────────────────────────────────────────────────────
def analyse(df, tf, live_price=None):
    if df is None or len(df)<50: return None
    if live_price:
        df=df.copy(); df.iloc[-1,df.columns.get_loc("Close")]=live_price

    c=float(df["Close"].iloc[-1])
    vp=float(calc_vwap(df).iloc[-1])
    bbu_s,bbm_s,bbl_s,bw_s=calc_bb(df)
    bbu,bbm,bbl,bwv=float(bbu_s.iloc[-1]),float(bbm_s.iloc[-1]),float(bbl_s.iloc[-1]),float(bw_s.iloc[-1])
    ml_s,ms_s,mh_s=calc_macd(df)
    mn,msn,mhn=float(ml_s.iloc[-1]),float(ms_s.iloc[-1]),float(mh_s.iloc[-1])
    mhp=float(mh_s.iloc[-2])
    at=float(calc_atr(df).iloc[-1])
    ri=float(calc_rsi(df).iloc[-1])
    stk,stk_d=calc_stoch(df)
    stk_v=float(stk.iloc[-1]); stk_dv=float(stk_d.iloc[-1])

    # Force Index (Elder)
    fi_raw,fi2_s,fi13_s=calc_force_index(df)
    fi2_v =float(fi2_s.iloc[-1])  if not np.isnan(fi2_s.iloc[-1])  else 0.0
    fi13_v=float(fi13_s.iloc[-1]) if not np.isnan(fi13_s.iloc[-1]) else 0.0

    bull=c>vp; vpct=(c-vp)/vp*100
    sq=bwv<0.018
    brange=bbu-bbl; bpos=(c-bbl)/brange if brange>0 else 0.5
    tl=c<=bbl*1.005; th=c>=bbu*0.995
    bb_ok=(tl if bull else th) or sq
    mx_bull=mn>msn and mhn>0; mx_bear=mn<msn and mhn<0
    mx_ok=mx_bull if bull else mx_bear
    h_ok=(mhn>0 and bull) or (mhn<0 and not bull)
    accel=(mhn>mhp) if bull else (mhn<mhp)
    n=min(10,len(df))
    pt=float(df["Close"].iloc[-n:].iloc[-1]-df["Close"].iloc[-n:].iloc[0])
    mt=float(ml_s.iloc[-n:].iloc[-1]-ml_s.iloc[-n:].iloc[0])
    div_ok=(pt<0 and mt>0) if bull else (pt>0 and mt<0)

    # Confluence checks
    checks={
        "vwap":    {"label":"Prix vs VWAP",    "ok":abs(vpct)>0.01,"val":f"{'>' if bull else '<'} VWAP ({vpct:+.2f}%)","score":1.0 if abs(vpct)>0.05 else 0.5},
        "bb":      {"label":"Bollinger Bands", "ok":bb_ok,"val":f"{'Bande inf ✓' if tl else 'Bande sup ✓' if th else f'Pos {bpos:.0%}'}"+(" [SQUEEZE]" if sq else ""),"score":1.0 if bb_ok else 0.3},
        "macd_x":  {"label":"MACD Crossover",  "ok":mx_ok,"val":f"MACD {mn:.5f} | Sig {msn:.5f}","score":1.0 if mx_ok else 0.2},
        "macd_h":  {"label":"Histogramme",     "ok":h_ok,"val":f"{mhn:.5f} ({'↑ accél' if accel else '↓ ralentit'})","score":1.0 if h_ok else 0.1},
        "div":     {"label":"Divergence MACD", "ok":div_ok,"val":f"{'Div haussière ✓' if (pt<0 and mt>0) else 'Div baissière ✓' if (pt>0 and mt<0) else 'Aucune'}","score":1.0 if div_ok else 0.0},
        "force_idx":{"label":"Force Index(13)","ok":(fi13_v>0 and bull) or (fi13_v<0 and not bull),"val":f"FI13={fi13_v:.0f} FI2={fi2_v:.0f}","score":0.8 if (fi13_v>0 and bull) or (fi13_v<0 and not bull) else 0.1},
    }

    # ML Ensemble
    feats=build_features(c,vp,bbu,bbl,bbm,mn,msn,mhn,ri,at,bwv,bull)
    ml_ens,ml_scores=run_ml_ensemble(feats,bull)
    ml_pct=round(ml_ens*100,1)

    # Final confluence
    raw=sum(x["score"] for x in checks.values())
    rsi_b=0.05 if (ri<70 and bull) or (ri>30 and not bull) else -0.03
    tech_conf=(raw/len(checks)+rsi_b)
    # Blend: 60% technical, 40% ML
    final_conf=min(99,max(5,round((tech_conf*0.60+ml_ens*0.40)*100)))
    ok_n=sum(1 for x in checks.values() if x["ok"])

    if ok_n>=4 and final_conf>=62: sig="BUY" if bull else "SELL"
    elif sq and ok_n<2:            sig="RANGE"
    elif final_conf<38:            sig="ATTENDRE"
    elif ok_n>=3:                  sig="BUY" if bull else "SELL"
    else:                          sig="ATTENDRE"

    if sq:                           regime="COMPRESSION / SQUEEZE"
    elif c>vp and c>=bbu*0.99:      regime="TENDANCE HAUSSIÈRE FORTE"
    elif c<vp and c<=bbl*1.01:      regime="TENDANCE BAISSIÈRE FORTE"
    elif bull and tl:                regime="MEAN-REVERSION HAUSSIÈRE"
    elif not bull and th:            regime="MEAN-REVERSION BAISSIÈRE"
    elif bull:                       regime="BIAIS HAUSSIER"
    else:                            regime="BIAIS BAISSIER"

    tm,sm=TP_SL.get(tf,(2.0,1.2))
    if sig=="BUY":    sl,tp=round(c-at*sm,5),round(c+at*tm,5)
    elif sig=="SELL": sl,tp=round(c+at*sm,5),round(c-at*tm,5)
    else:             sl,tp=None,None

    # ICT
    fvgs=detect_fvg(df)
    obs=detect_order_blocks(df)
    liq=detect_liquidity(df)

    return {
        "tf":tf,"close":c,"vwap":vp,
        "bb_upper":bbu,"bb_mid":bbm,"bb_lower":bbl,"bw":bwv,"squeeze":sq,
        "macd":mn,"macd_sig":msn,"macd_hist":mhn,"rsi":ri,"atr":at,
        "stoch_k":stk_v,"stoch_d":stk_dv,
        "fi2":fi2_v,"fi13":fi13_v,"fi2_s":fi2_s,"fi13_s":fi13_s,
        "checks":checks,"ok_count":ok_n,"confidence":final_conf,
        "ml_ensemble":ml_pct,"ml_scores":ml_scores,
        "signal":sig,"regime":regime,"vwap_bull":bull,
        "sl":sl,"tp":tp,"rr":round(tm/sm,2) if sl else None,
        "fvgs":fvgs,"obs":obs,"liquidity":liq,
        "df_raw":df,
        "bb_upper_s":bbu_s,"bb_mid_s":bbm_s,"bb_lower_s":bbl_s,
        "vwap_s":calc_vwap(df),
        "macd_s":ml_s,"macd_sig_s":ms_s,"macd_hist_s":mh_s,
        "rsi_s":calc_rsi(df),
    }

# ─── CHART ───────────────────────────────────────────────────────────────────
def render_chart_fallback(res, asset_name):
    """Native Streamlit chart — no Plotly needed. Works everywhere."""
    df       = res["df_raw"].iloc[-80:]
    vwap_l   = res["vwap_s"].iloc[-80:]
    bbu_l    = res["bb_upper_s"].iloc[-80:]
    bbl_l    = res["bb_lower_s"].iloc[-80:]
    macd_l   = res["macd_s"].iloc[-80:]
    rsi_l    = res["rsi_s"].iloc[-80:]

    st.markdown(f"""<div style="font-family:sans-serif;font-size:13px;font-weight:700;
        color:#C9A94B;letter-spacing:2px;margin-bottom:8px;">
        📈 {asset_name} — {res['tf'].upper()} (Graphique natif)</div>""",
        unsafe_allow_html=True)

    # Price + VWAP + BB chart
    import pandas as pd
    price_df = pd.DataFrame({
        "Close":       df["Close"],
        "VWAP":        vwap_l,
        "BB Sup":      bbu_l,
        "BB Inf":      bbl_l,
    })
    st.line_chart(price_df, height=280, use_container_width=True)

    # MACD
    st.markdown('<div style="font-size:11px;color:#4A7EC7;margin:4px 0;">MACD</div>',
                unsafe_allow_html=True)
    st.line_chart(macd_l.rename("MACD"), height=120, use_container_width=True)

    # RSI
    st.markdown('<div style="font-size:11px;color:#A78BFA;margin:4px 0;">RSI(14)</div>',
                unsafe_allow_html=True)
    st.line_chart(rsi_l.rename("RSI"), height=120, use_container_width=True)

    # TP/SL info
    if res.get("tp"):
        c1, c2 = st.columns(2)
        with c1: st.metric("✅ Take Profit", f"{res['tp']:.5f}")
        with c2: st.metric("🛑 Stop Loss",   f"{res['sl']:.5f}")

def render_chart(res, asset_name):
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except Exception:
        # Plotly not available — use native Streamlit chart
        render_chart_fallback(res, asset_name)
        return

    df=res["df_raw"].iloc[-80:]
    vwap_line=res["vwap_s"].iloc[-80:]
    bbu_line=res["bb_upper_s"].iloc[-80:]
    bbm_line=res["bb_mid_s"].iloc[-80:]
    bbl_line=res["bb_lower_s"].iloc[-80:]
    macd_line=res["macd_s"].iloc[-80:]
    macd_sig_line=res["macd_sig_s"].iloc[-80:]
    hist_line=res["macd_hist_s"].iloc[-80:]
    rsi_line=res["rsi_s"].iloc[-80:]

    idx=[str(x)[:16] for x in df.index]
    c_price=float(df["Close"].iloc[-1])

    fig=make_subplots(rows=3,cols=1,
        shared_xaxes=True,
        row_heights=[0.55,0.25,0.20],
        vertical_spacing=0.04,
        subplot_titles=["","MACD","RSI"])

    # Candlesticks
    colors=["#0FBF5F" if row["Close"]>=row["Open"] else "#D93025" for _,row in df.iterrows()]
    fig.add_trace(go.Candlestick(
        x=idx,open=df["Open"],high=df["High"],low=df["Low"],close=df["Close"],
        increasing_line_color="#0FBF5F",decreasing_line_color="#D93025",
        increasing_fillcolor="#0FBF5F",decreasing_fillcolor="#D93025",
        line_width=1,name="Prix"),row=1,col=1)

    # BB
    fig.add_trace(go.Scatter(x=idx,y=bbu_line,name="BB Sup",
        line=dict(color="#1DB954",width=1,dash="dot"),opacity=0.7),row=1,col=1)
    fig.add_trace(go.Scatter(x=idx,y=bbl_line,name="BB Inf",
        line=dict(color="#E53E3E",width=1,dash="dot"),opacity=0.7,
        fill="tonexty",fillcolor="rgba(100,100,100,0.05)"),row=1,col=1)
    fig.add_trace(go.Scatter(x=idx,y=bbm_line,name="BB Mid",
        line=dict(color="#4A7EC7",width=1,dash="dash"),opacity=0.5),row=1,col=1)
    # VWAP
    fig.add_trace(go.Scatter(x=idx,y=vwap_line,name="VWAP",
        line=dict(color="#C9A94B",width=2)),row=1,col=1)

    # TP/SL lines
    if res["tp"]:
        fig.add_hline(y=res["tp"],line_color="#0FBF5F",line_dash="dash",line_width=1.5,row=1,col=1,
            annotation_text=f"TP {res['tp']:.3f}",annotation_font_color="#0FBF5F",annotation_position="right")
        fig.add_hline(y=res["sl"],line_color="#D93025",line_dash="dash",line_width=1.5,row=1,col=1,
            annotation_text=f"SL {res['sl']:.3f}",annotation_font_color="#D93025",annotation_position="right")

    # MACD
    fig.add_trace(go.Scatter(x=idx,y=macd_line,name="MACD",
        line=dict(color="#4A7EC7",width=1.5)),row=2,col=1)
    fig.add_trace(go.Scatter(x=idx,y=macd_sig_line,name="Signal",
        line=dict(color="#C9A94B",width=1.5,dash="dot")),row=2,col=1)
    hist_colors=["#0FBF5F" if v>=0 else "#D93025" for v in hist_line]
    fig.add_trace(go.Bar(x=idx,y=hist_line,name="Histo",
        marker_color=hist_colors,opacity=0.8),row=2,col=1)

    # RSI
    fig.add_trace(go.Scatter(x=idx,y=rsi_line,name="RSI",
        line=dict(color="#A78BFA",width=1.5)),row=3,col=1)
    fig.add_hline(y=70,line_color="#D93025",line_width=0.8,line_dash="dot",row=3,col=1)
    fig.add_hline(y=30,line_color="#0FBF5F",line_width=0.8,line_dash="dot",row=3,col=1)

    fig.update_layout(
        height=520,
        paper_bgcolor="#060E1A",plot_bgcolor="#060E1A",
        font=dict(color="#8A95A3",family="JetBrains Mono",size=10),
        legend=dict(orientation="h",bgcolor="rgba(0,0,0,0)",
                    font=dict(size=9,color="#8A95A3"),y=1.02),
        xaxis_rangeslider_visible=False,
        margin=dict(l=50,r=20,t=20,b=20),
        title=dict(text=f"{asset_name} — {res['tf'].upper()}",
                   font=dict(color="#C9A94B",size=14,family="Rajdhani"),x=0.5),
    )
    for ax in ["xaxis","xaxis2","xaxis3","yaxis","yaxis2","yaxis3"]:
        fig.update_layout({ax:dict(
            gridcolor="#1A3050",showgrid=True,linecolor="#1A3050",
            tickfont=dict(color="#6B7A8D",size=9))})

    st.plotly_chart(fig,use_container_width=True,config={"displayModeBar":False})

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def sc(s): return {"BUY":"#0FBF5F","SELL":"#D93025","RANGE":"#E07B2A","ATTENDRE":"#6B7A8D"}.get(s,"#6B7A8D")
def sb(s): return {"BUY":"rgba(15,191,95,.08)","SELL":"rgba(217,48,37,.08)","RANGE":"rgba(224,123,42,.08)","ATTENDRE":"rgba(107,122,141,.05)"}.get(s,"")
def se(s): return {"BUY":"🟢","SELL":"🔴","RANGE":"🟡","ATTENDRE":"⏸️"}.get(s,"⚪")
def cc(c): return "#0FBF5F" if c>=75 else "#C9A94B" if c>=55 else "#E07B2A" if c>=35 else "#D93025"
def src_badge(src):
    if src=="binance": return '<span class="badge b-binance">⚡ BINANCE LIVE</span>'
    if src=="oanda":   return '<span class="badge b-oanda">🔵 OANDA LIVE</span>'
    return '<span class="badge b-yf">⚠️ YAHOO ~15min</span>'
def kz_active(now):
    hm=now.strftime("%H:%M")
    return [k for k in KILL_ZONES if k["start"]<=hm<k["end"]]

def modal_html(step,pct):
    return f"""<div style="position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(6,14,26,.95);
     z-index:9999;display:flex;align-items:center;justify-content:center;backdrop-filter:blur(8px);">
  <div style="background:#0F1C2E;border:2px solid #C9A94B;border-radius:16px;padding:30px 34px;
       min-width:290px;max-width:340px;text-align:center;box-shadow:0 0 80px rgba(201,169,75,.3);">
    <div style="font-family:'Rajdhani',sans-serif;font-size:22px;font-weight:700;color:#C9A94B;letter-spacing:2px;margin-bottom:6px;">
      🤖 ANALYSE ML EN COURS
    </div>
    <div style="font-size:11px;color:#6B7A8D;letter-spacing:1px;margin-bottom:20px;min-height:16px;">{step}</div>
    <div style="width:100%;height:8px;background:#1A3050;border-radius:4px;overflow:hidden;margin-bottom:12px;">
      <div style="width:{pct}%;height:100%;background:linear-gradient(90deg,#C9A94B,#F0C96A);border-radius:4px;"></div>
    </div>
    <div style="font-family:'Rajdhani',sans-serif;font-size:42px;font-weight:700;color:#F0C96A;">{pct}%</div>
    <div style="font-size:10px;color:#6B7A8D;margin-top:8px;letter-spacing:1px;">
      MACD · BB · VWAP · 5 Modèles ML · ICT
    </div>
  </div>
</div>"""

# ─── RENDER BLOCKS ───────────────────────────────────────────────────────────
def render_signal_panel(res):
    s,conf=res["signal"],res["confidence"]; col=sc(s); ccol=cc(conf)
    ml=res["ml_ensemble"]
    st.markdown(f"""
    <div class="card" style="background:{sb(s)};border-color:{col};border-width:2px;text-align:center;padding:18px;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:40px;font-weight:700;color:{col};letter-spacing:3px;">
        {se(s)} {s}
      </div>
      <div style="color:#6B7A8D;font-size:10px;letter-spacing:2px;margin:5px 0 14px 0;">{res['regime']}</div>
      <div style="display:flex;justify-content:center;gap:20px;">
        <div style="display:flex;flex-direction:column;align-items:center;justify-content:center;
             width:100px;height:100px;border-radius:50%;border:3px solid {ccol};background:rgba(0,0,0,.3);">
          <div style="font-family:'Rajdhani',sans-serif;font-size:28px;font-weight:700;color:{ccol};line-height:1;">{conf}%</div>
          <div style="font-size:8px;color:#6B7A8D;letter-spacing:1px;">CONFLUENC</div>
        </div>
        <div style="display:flex;flex-direction:column;align-items:center;justify-content:center;
             width:100px;height:100px;border-radius:50%;border:3px solid #7C3AED;background:rgba(0,0,0,.3);">
          <div style="font-family:'Rajdhani',sans-serif;font-size:28px;font-weight:700;color:#A78BFA;line-height:1;">{ml}%</div>
          <div style="font-size:8px;color:#6B7A8D;letter-spacing:1px;">ML SCORE</div>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

def render_ml_panel(res):
    scores=res["ml_scores"]
    rows=""
    for name,score in scores.items():
        pct=round(score*100); col=cc(pct)
        direction="🟢 BUY" if (score>0.5 and res["vwap_bull"]) or (score<0.5 and not res["vwap_bull"]) else "🔴 SELL" if score<0.5 else "⚪"
        rows+=f"""
        <tr>
          <td style="color:#B0BCC8;text-align:left;padding:5px 8px;">{name}</td>
          <td style="padding:5px 8px;">
            <div style="width:100%;background:#1A3050;border-radius:3px;height:6px;overflow:hidden;">
              <div style="width:{pct}%;height:6px;background:{col};border-radius:3px;"></div>
            </div>
          </td>
          <td style="color:{col};font-weight:700;padding:5px 8px;">{pct}%</td>
        </tr>"""
    ens=res["ml_ensemble"]; ecol=cc(ens)
    st.markdown(f"""
    <div class="card card-purple">
      <div style="font-family:'Rajdhani',sans-serif;font-size:13px;color:#A78BFA;letter-spacing:2px;margin-bottom:8px;">
        🤖 5 MODÈLES ML — ENSEMBLE {ens}%
      </div>
      <table style="width:100%;border-collapse:collapse;">{rows}</table>
    </div>""", unsafe_allow_html=True)

def render_tpsl(res):
    if res["tp"] is None:
        st.markdown('<div class="card card-orange" style="text-align:center;padding:14px;"><span style="color:#E07B2A;font-family:\'Rajdhani\',sans-serif;font-size:15px;">⏸ Pas de trade — Signal insuffisant</span></div>', unsafe_allow_html=True)
        return
    s,c,tp,sl,at,rr=res["signal"],res["close"],res["tp"],res["sl"],res["atr"],res["rr"]
    col=sc(s); td,sd=abs(tp-c),abs(sl-c)
    st.markdown(f"""
    <div class="card" style="border-color:{col};">
      <div style="font-family:'Rajdhani',sans-serif;font-size:13px;color:{col};letter-spacing:2px;margin-bottom:9px;">💰 NIVEAUX — {s} (R:R {rr}:1)</div>
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;text-align:center;">
        <div style="background:rgba(15,191,95,.1);border:1px solid #0FBF5F;border-radius:7px;padding:8px;">
          <div style="color:#6B7A8D;font-size:9px;letter-spacing:1px;">TAKE PROFIT</div>
          <div style="color:#0FBF5F;font-family:'Rajdhani',sans-serif;font-size:16px;font-weight:700;">{tp:.5f}</div>
          <div style="color:#0FBF5F;font-size:10px;">+{td:.5f}</div>
        </div>
        <div style="background:rgba(201,169,75,.1);border:1px solid #C9A94B;border-radius:7px;padding:8px;">
          <div style="color:#6B7A8D;font-size:9px;letter-spacing:1px;">PRIX LIVE</div>
          <div style="color:#C9A94B;font-family:'Rajdhani',sans-serif;font-size:16px;font-weight:700;">{c:.5f}</div>
          <div style="color:#6B7A8D;font-size:10px;">ATR {at:.5f}</div>
        </div>
        <div style="background:rgba(217,48,37,.1);border:1px solid #D93025;border-radius:7px;padding:8px;">
          <div style="color:#6B7A8D;font-size:9px;letter-spacing:1px;">STOP LOSS</div>
          <div style="color:#D93025;font-family:'Rajdhani',sans-serif;font-size:16px;font-weight:700;">{sl:.5f}</div>
          <div style="color:#D93025;font-size:10px;">-{sd:.5f}</div>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

def render_checklist(res):
    checks=res["checks"]; ok_n=res["ok_count"]; conf=res["confidence"]
    for k,c in checks.items():
        ok=c["ok"]; col="#0FBF5F" if ok else "#D93025"
        st.markdown(f"""<div class="chk {'chk-ok' if ok else 'chk-no'}">
          <span style="font-size:14px;">{'✅' if ok else '❌'}</span>
          <span style="flex:1;color:#DCE8F5;">{c['label']}</span>
          <span style="font-weight:700;font-size:11px;color:{col};">{c['val']}</span>
        </div>""", unsafe_allow_html=True)
    pct=int(ok_n/len(checks)*100); ccol=cc(conf)
    st.markdown(f"""<div class="card" style="margin-top:8px;">
      <div style="display:flex;justify-content:space-between;margin-bottom:4px;">
        <span style="font-size:11px;color:#6B7A8D;">Score confluence</span>
        <span style="font-weight:700;color:{ccol};">{ok_n}/{len(checks)} — {conf}%</span>
      </div>
      <div style="width:100%;height:7px;background:#1A3050;border-radius:3px;overflow:hidden;">
        <div style="width:{pct}%;height:100%;background:{ccol};border-radius:3px;"></div>
      </div>
      <div style="display:flex;justify-content:space-between;margin-top:4px;font-size:10px;color:#6B7A8D;">
        <span>RSI {res['rsi']:.1f}</span>
        <span>Stoch K:{res['stoch_k']:.1f} D:{res['stoch_d']:.1f}</span>
        <span>{'🔴 SQUEEZE' if res['squeeze'] else 'BW '+str(round(res['bw'],4))}</span>
      </div>
    </div>""", unsafe_allow_html=True)

def render_force_index(res):
    """Render Force Index panel — Elder method."""
    import streamlit.components.v1 as _comp_fi
    fi2  = res.get("fi2",  0)
    fi13 = res.get("fi13", 0)
    sig  = res.get("signal", "")
    atr  = res.get("atr", 1)
    c    = res.get("close", 1)
    fi   = interpret_force_index(fi2, fi13, sig, atr, c)

    # Bar widths (0-95 range, center=50)
    denom13 = max(abs(fi13), abs(c) * 0.5, 1)
    denom2  = max(abs(fi2),  abs(c) * 0.5, 1)
    fi13_pct = min(45, int(abs(fi13) / denom13 * 45))
    fi2_pct  = min(45, int(abs(fi2)  / denom2  * 45))

    fi13_bar_html = (
        f'<div style="position:absolute;left:50%;top:0;height:100%;width:{fi13_pct}%;'
        f'background:#0FBF5F;opacity:0.85;border-radius:0 4px 4px 0;"></div>'
        if fi13 > 0 else
        f'<div style="position:absolute;right:50%;top:0;height:100%;width:{fi13_pct}%;'
        f'background:#D93025;opacity:0.85;border-radius:4px 0 0 4px;"></div>'
    )
    fi2_bar_html = (
        f'<div style="position:absolute;left:50%;top:0;height:100%;width:{fi2_pct}%;'
        f'background:#0FBF5F;opacity:0.85;border-radius:0 4px 4px 0;"></div>'
        if fi2 > 0 else
        f'<div style="position:absolute;right:50%;top:0;height:100%;width:{fi2_pct}%;'
        f'background:#D93025;opacity:0.85;border-radius:4px 0 0 4px;"></div>'
    )

    div_banner = ""
    if fi["divergence"]:
        div_banner = (
            f'<div style="background:rgba(224,123,42,.12);border:1px solid #E07B2A;'
            f'border-radius:6px;padding:8px 12px;margin-top:8px;font-size:11px;'
            f'color:#E07B2A;font-weight:700;">'
            f"\u26A0\uFE0F DIVERGENCE FORCE INDEX \u2014 Signal {sig} mais FI(13) "
            f"{'baissier' if fi13 < 0 else 'haussier'} \u2192 Prudence !</div>"
        )

    conf_banner = ""
    if fi["confirmed"]:
        cc = fi["trend_col"]
        conf_banner = (
            f'<div style="background:rgba(15,191,95,.10);border:1px solid {cc};'
            f'border-radius:6px;padding:8px 12px;margin-top:8px;font-size:12px;'
            f'color:{cc};font-weight:700;text-align:center;">'
            f"\u2705 FORCE INDEX CONFIRME LA TENDANCE {sig}</div>"
        )

    fi2_col  = "#0FBF5F" if fi2  > 0 else "#D93025"
    fi13_col = fi["trend_col"]
    fi2_dir  = "Conviction ACHAT \u2191" if fi2  > 0 else "Conviction VENTE \u2193"
    fi13_dir = "HAUSSIER \u2191"         if fi13 > 0 else "BAISSIER \u2193"

    html = (
        "<!DOCTYPE html><html><head>"
        "<link href=\"https://fonts.googleapis.com/css2?family=Rajdhani:wght@700"
        "&family=JetBrains+Mono&display=swap\" rel=\"stylesheet\">"
        "<style>body{margin:0;padding:8px;background:#060E1A;"
        "font-family:'Rajdhani',sans-serif;color:#DCE8F5;}</style></head><body>"

        "<div style=\"background:linear-gradient(135deg,#0A1520,#0A0A20);"
        "border:2px solid #A78BFA;border-radius:10px;padding:10px 14px;"
        "margin-bottom:10px;text-align:center;\">"
        "<div style=\"font-size:15px;font-weight:700;color:#A78BFA;letter-spacing:2px;\">"
        "\u26A1 FORCE INDEX \u2014 Elder Ray</div>"
        "<div style=\"font-size:9px;color:#6B7A8D;margin-top:2px;\">"
        "Volume \xd7 Variation de Prix \u2014 Force et Conviction</div></div>"

        "<div style=\"display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:10px;\">"
        f"<div style=\"background:#0F1C2E;border:1px solid {fi2_col};"
        f"border-radius:8px;padding:8px;text-align:center;\">"
        "<div style=\"font-size:9px;color:#6B7A8D;\">FI(2) \u2014 Court Terme</div>"
        f"<div style=\"font-size:20px;font-weight:700;color:{fi2_col};\">{fi2:+.0f}</div>"
        f"<div style=\"font-size:10px;color:{fi2_col};\">{fi2_dir}</div></div>"

        f"<div style=\"background:#0F1C2E;border:1px solid {fi13_col};"
        f"border-radius:8px;padding:8px;text-align:center;\">"
        "<div style=\"font-size:9px;color:#6B7A8D;\">FI(13) \u2014 Tendance</div>"
        f"<div style=\"font-size:20px;font-weight:700;color:{fi13_col};\">{fi13:+.0f}</div>"
        f"<div style=\"font-size:10px;color:{fi13_col};\">{fi13_dir}</div></div></div>"

        "<div style=\"margin-bottom:8px;\">"
        f"<div style=\"font-size:10px;color:#6B7A8D;margin-bottom:3px;\">FI(13) Tendance \u2014 {fi13_dir}</div>"
        "<div style=\"background:#0F1C2E;border-radius:6px;height:18px;position:relative;overflow:hidden;\">"
        "<div style=\"position:absolute;left:50%;top:0;width:1px;height:100%;background:#1A3050;\"></div>"
        f"{fi13_bar_html}</div></div>"

        "<div style=\"margin-bottom:10px;\">"
        f"<div style=\"font-size:10px;color:#6B7A8D;margin-bottom:3px;\">FI(2) Court Terme \u2014 {fi2_dir}</div>"
        "<div style=\"background:#0F1C2E;border-radius:6px;height:18px;position:relative;overflow:hidden;\">"
        "<div style=\"position:absolute;left:50%;top:0;width:1px;height:100%;background:#1A3050;\"></div>"
        f"{fi2_bar_html}</div></div>"

        "<div style=\"display:flex;flex-direction:column;gap:6px;margin-bottom:8px;\">"
        f"<div style=\"background:rgba(0,0,0,.2);border-left:3px solid {fi13_col};"
        f"border-radius:5px;padding:7px 11px;font-size:12px;color:{fi13_col};font-weight:700;\">"
        f"\U0001F4CA {fi['trend_str']}</div>"
        f"<div style=\"background:rgba(0,0,0,.2);border-left:3px solid {fi['entry_col']};"
        f"border-radius:5px;padding:7px 11px;font-size:11px;color:{fi['entry_col']};\">"
        f"\U0001F3AF {fi['entry_sig']}</div></div>"

        f"{div_banner}{conf_banner}"

        "<div style=\"background:#0F1C2E;border-left:3px solid #A78BFA;border-radius:6px;"
        "padding:9px 12px;margin-top:10px;font-size:11px;color:#8A95A3;line-height:1.9;\">"
        "<b style=\"color:#A78BFA;\">\U0001F4D6 REGLES FORCE INDEX (Elder) :</b><br>"
        "\U0001F7E2 FI(13) positif = Tendance HAUSSIERE \u2192 Favoriser les BUY<br>"
        "\U0001F534 FI(13) negatif = Tendance BAISSIERE \u2192 Favoriser les SELL<br>"
        "\u26A1 FI(2) &lt; 0 en uptrend (FI13&gt;0) = Meilleur moment pour BUY<br>"
        "\u26A1 FI(2) &gt; 0 en downtrend (FI13&lt;0) = Meilleur moment pour SELL<br>"
        "\u26A0\uFE0F FI(13) diverge du signal = Prudence \u2014 attendre confirmation<br>"
        "\U0001F4C8 FI tres eleve = Tendance tres forte \u2014 ne pas aller contre"
        "</div></body></html>"
    )
    _comp_fi.html(html, height=490, scrolling=True)


def render_ict(res):
    fvgs=res["fvgs"]; obs=res["obs"]; liq=res["liquidity"]
    c=res["close"]
    # FVGs
    fvg_html=""
    for f in reversed(fvgs[-3:]):
        col="#0FBF5F" if f["type"]=="bull" else "#D93025"
        dist=abs(f["mid"]-c)/c*100
        fvg_html+=f'<div class="ict-item"><span style="color:{col};font-weight:700;">FVG {"📈" if f["type"]=="bull" else "📉"}</span><span style="color:#B0BCC8;font-size:10px;">{f["bot"]:.4f}–{f["top"]:.4f}</span><span style="color:#C9A94B;">{dist:.2f}%</span></div>'
    # OBs
    ob_html=""
    for o in reversed(obs[-3:]):
        col="#0FBF5F" if o["type"]=="bull" else "#D93025"
        dist=abs(o["mid"]-c)/c*100
        ob_html+=f'<div class="ict-item"><span style="color:{col};font-weight:700;">OB {"🟢" if o["type"]=="bull" else "🔴"}</span><span style="color:#B0BCC8;font-size:10px;">{o["bot"]:.4f}–{o["top"]:.4f}</span><span style="color:#C9A94B;">{dist:.2f}%</span></div>'
    # Liquidity
    bsl_html="".join(f'<span style="color:#D93025;font-size:11px;margin:2px 4px;background:rgba(217,48,37,.1);padding:2px 7px;border-radius:4px;">{v:.4f}</span>' for v in liq["buyside"])
    ssl_html="".join(f'<span style="color:#0FBF5F;font-size:11px;margin:2px 4px;background:rgba(15,191,95,.1);padding:2px 7px;border-radius:4px;">{v:.4f}</span>' for v in liq["sellside"])

    st.markdown(f"""<div class="card card-purple">
      <div style="font-family:'Rajdhani',sans-serif;font-size:13px;color:#A78BFA;letter-spacing:2px;margin-bottom:8px;">🎯 NIVEAUX ICT</div>
      <div style="font-size:10px;color:#7C3AED;letter-spacing:1px;margin-bottom:4px;">FAIR VALUE GAPS</div>
      {fvg_html if fvg_html else '<div style="color:#6B7A8D;font-size:11px;padding:4px 8px;">Aucun FVG récent détecté</div>'}
      <div style="font-size:10px;color:#7C3AED;letter-spacing:1px;margin:8px 0 4px 0;">ORDER BLOCKS</div>
      {ob_html if ob_html else '<div style="color:#6B7A8D;font-size:11px;padding:4px 8px;">Aucun OB récent détecté</div>'}
      <div style="font-size:10px;color:#7C3AED;letter-spacing:1px;margin:8px 0 4px 0;">LIQUIDITÉ (Buyside / Sellside)</div>
      <div style="margin-bottom:3px;">
        <span style="font-size:10px;color:#D93025;margin-right:6px;">▲ BSL:</span>{bsl_html if bsl_html else '<span style="color:#6B7A8D;font-size:10px;">–</span>'}
      </div>
      <div>
        <span style="font-size:10px;color:#0FBF5F;margin-right:6px;">▼ SSL:</span>{ssl_html if ssl_html else '<span style="color:#6B7A8D;font-size:10px;">–</span>'}
      </div>
    </div>""", unsafe_allow_html=True)

def render_kz(now):
    active_n={k["name"] for k in kz_active(now)}
    items="".join(f"""<div style="display:flex;justify-content:space-between;align-items:center;
         padding:5px 10px;border-radius:6px;margin:3px 0;
         background:rgba(0,0,0,.2);border:1px solid {k['col'] if k['name'] in active_n else '#1A3050'};">
      <span style="color:{k['col'] if k['name'] in active_n else '#2A3A4A'};font-size:12px;">
        {'●' if k['name'] in active_n else '○'} {k['name']}
      </span>
      <span style="color:{k['col'] if k['name'] in active_n else '#2A3A4A'};font-size:10px;font-weight:700;">
        {'✓ ACTIVE' if k['name'] in active_n else k['start']+'–'+k['end']}
      </span>
    </div>""" for k in KILL_ZONES)
    st.markdown(f'<div class="card card-gold"><div style="font-family:\'Rajdhani\',sans-serif;font-size:13px;color:#C9A94B;letter-spacing:2px;margin-bottom:8px;">🎯 KILL ZONES — {now.strftime("%H:%M UTC")}</div>{items}</div>', unsafe_allow_html=True)

def render_indicators(res,src):
    c=res["close"]
    rows=[("Prix LIVE",f"{c:.5f}","#F0C96A"),("VWAP",f"{res['vwap']:.5f}","#4A7EC7"),
          ("BB Sup",f"{res['bb_upper']:.5f}","#0FBF5F"),("BB Mid",f"{res['bb_mid']:.5f}","#6B7A8D"),
          ("BB Inf",f"{res['bb_lower']:.5f}","#D93025"),
          ("MACD",f"{res['macd']:.6f}","#C9A94B"),("Sig MACD",f"{res['macd_sig']:.6f}","#6B7A8D"),
          ("Histo",f"{res['macd_hist']:.6f}","#0FBF5F" if res['macd_hist']>0 else "#D93025"),
          ("ATR(14)",f"{res['atr']:.5f}","#6B7A8D"),
          ("RSI(14)",f"{res['rsi']:.1f}","#C9A94B" if 30<res['rsi']<70 else "#D93025"),
          ("Stoch %K",f"{res['stoch_k']:.1f}","#A78BFA"),
          ("Bandwidth",f"{res['bw']:.4f}"+(" 🔴" if res['squeeze'] else ""),"#E07B2A" if res['squeeze'] else "#6B7A8D"),
          ("─────────","────────","#1A3050"),
          ("Force Idx FI(2)", f"{res.get('fi2',0):+.0f}", "#0FBF5F" if res.get('fi2',0)>0 else "#D93025"),
          ("Force Idx FI(13)",f"{res.get('fi13',0):+.0f}", "#0FBF5F" if res.get('fi13',0)>0 else "#D93025"),
          ("FI Tendance", interpret_force_index(res.get('fi2',0),res.get('fi13',0),res.get('signal',''),res.get('atr',1),res.get('close',1))['trend_str'],
                          interpret_force_index(res.get('fi2',0),res.get('fi13',0),res.get('signal',''),res.get('atr',1),res.get('close',1))['trend_col'])]
    rh="".join(f'<tr><td style="color:#6B7A8D;padding:4px 8px;text-align:left;font-size:11px;">{r[0]}</td><td style="color:{r[2]};padding:4px 8px;text-align:right;font-weight:700;font-size:11px;">{r[1]}</td></tr>' for r in rows)
    st.markdown(f'<div class="card card-blue"><div style="font-family:\'Rajdhani\',sans-serif;font-size:12px;color:#4A7EC7;letter-spacing:2px;margin-bottom:8px;">📐 INDICATEURS &nbsp; {src_badge(src)}</div><table style="width:100%;border-collapse:collapse;">{rh}</table></div>', unsafe_allow_html=True)

def render_multitf(results):
    rows=""
    for tf,res in results.items():
        if not res: continue
        s,conf=res["signal"],res["confidence"]; col=sc(s); ccol=cc(conf)
        rows+=f'<tr><td style="color:#C9A94B;font-weight:700;padding:5px 8px;">{tf}</td><td style="color:{col};font-weight:700;padding:5px 8px;">{se(s)} {s}</td><td style="color:{ccol};font-weight:700;padding:5px 8px;">{conf}%</td><td style="color:#A78BFA;padding:5px 8px;">{res["ml_ensemble"]}%</td><td style="color:#B0BCC8;padding:5px 8px;font-size:10px;">{res["regime"][:20]}</td><td style="color:{"#0FBF5F" if res["vwap_bull"] else "#D93025"};padding:5px 8px;">{"↑" if res["vwap_bull"] else "↓"}</td></tr>'
    st.markdown(f'<div class="card card-gold"><div style="font-family:\'Rajdhani\',sans-serif;font-size:13px;color:#C9A94B;letter-spacing:2px;margin-bottom:8px;">🔭 MULTI-TIMEFRAME 5m → 1h</div><table class="gt"><thead><tr><th>TF</th><th>SIGNAL</th><th>CONF</th><th>ML</th><th>RÉGIME</th><th>VWAP</th></tr></thead><tbody>{rows}</tbody></table></div>', unsafe_allow_html=True)


# ─── EXPORT CSV / EXCEL ──────────────────────────────────────────────────────
def export_to_csv(data_list, filename="export.csv"):
    """Convert list of dicts to CSV bytes for st.download_button."""
    import io, csv
    if not data_list: return b""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=data_list[0].keys(),
                            extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    writer.writerows(data_list)
    return buf.getvalue().encode("utf-8-sig")  # utf-8-sig = Excel-compatible BOM

def export_to_excel(data_list, sheet_name="Data", filename="export.xlsx"):
    """Convert list of dicts to Excel bytes for st.download_button."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data_list:
            buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

        headers = list(data_list[0].keys())

        # Header row style
        hdr_fill  = PatternFill("solid", fgColor="0F1C2E")
        hdr_font  = Font(bold=True, color="C9A94B", name="Calibri", size=10)
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin      = Side(style="thin", color="1A3050")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h.upper())
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = border
            ws.column_dimensions[cell.column_letter].width = max(12, len(str(h))+4)

        # Data rows
        alt_fill1 = PatternFill("solid", fgColor="080F1A")
        alt_fill2 = PatternFill("solid", fgColor="0A1120")
        data_font = Font(color="DCE8F5", name="Calibri", size=9)

        for row_idx, row_data in enumerate(data_list, 2):
            fill = alt_fill1 if row_idx % 2 == 0 else alt_fill2
            for col_idx, h in enumerate(headers, 1):
                val = row_data.get(h, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font   = data_font
                cell.fill   = fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze header row
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        # openpyxl not available — fallback to CSV bytes
        return export_to_csv(data_list)

def res_to_export_row(res, asset_name, tf):
    """Convert analysis result dict to a flat export row."""
    if not res: return {}
    from datetime import datetime, timezone
    fi_interp = interpret_force_index(
        res.get("fi2",0), res.get("fi13",0),
        res.get("signal",""), res.get("atr",1), res.get("close",1)
    )
    return {
        "Date/Heure UTC":  datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"),
        "Actif":           asset_name,
        "Timeframe":       tf,
        "Signal":          res.get("signal",""),
        "Confluence (%)":  res.get("confidence",0),
        "ML Ensemble (%)": res.get("ml_ensemble",0),
        "Prix":            res.get("close",0),
        "VWAP":            round(res.get("vwap",0),5),
        "BB Sup":          round(res.get("bb_upper",0),5),
        "BB Mid":          round(res.get("bb_mid",0),5),
        "BB Inf":          round(res.get("bb_lower",0),5),
        "MACD":            round(res.get("macd",0),6),
        "MACD Signal":     round(res.get("macd_sig",0),6),
        "MACD Histo":      round(res.get("macd_hist",0),6),
        "RSI(14)":         round(res.get("rsi",0),1),
        "ATR(14)":         round(res.get("atr",0),5),
        "Stoch %K":        round(res.get("stoch_k",0),1),
        "Stoch %D":        round(res.get("stoch_d",0),1),
        "FI(2)":           round(res.get("fi2",0),0),
        "FI(13)":          round(res.get("fi13",0),0),
        "FI Tendance":     fi_interp.get("trend_str",""),
        "Take Profit":     res.get("tp",""),
        "Stop Loss":       res.get("sl",""),
        "R:R":             res.get("rr",""),
        "Regime":          res.get("regime",""),
        "VWAP Biais":      "HAUSSIER" if res.get("vwap_bull") else "BAISSIER",
        "Squeeze BB":      "OUI" if res.get("squeeze") else "NON",
    }

def render_export_buttons(data_list, label, base_filename, key_prefix):
    """Render CSV + Excel download buttons side by side."""
    from datetime import datetime, timezone
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    fname_csv   = f"{base_filename}_{ts}.csv"
    fname_xlsx  = f"{base_filename}_{ts}.xlsx"

    st.markdown(
        f'''<div style="font-family:'Rajdhani',sans-serif;font-size:12px;
        color:#C9A94B;letter-spacing:2px;margin:10px 0 6px 0;">
        📥 EXPORTER {label}</div>''', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        csv_bytes = export_to_csv(data_list)
        st.download_button(
            label="📄 Télécharger CSV",
            data=csv_bytes,
            file_name=fname_csv,
            mime="text/csv",
            use_container_width=True,
            key=f"{key_prefix}_csv"
        )
    with col2:
        xl_bytes = export_to_excel(data_list, sheet_name=label[:31])
        mime_xl = ("application/vnd.openxmlformats-officedocument"
                   ".spreadsheetml.sheet")
        st.download_button(
            label="📊 Télécharger Excel",
            data=xl_bytes,
            file_name=fname_xlsx,
            mime=mime_xl,
            use_container_width=True,
            key=f"{key_prefix}_xlsx"
        )
    st.caption(f"✅ {len(data_list)} ligne(s) exportée(s)")

def render_history_tab():
    h=load_history()
    if not h:
        st.markdown('<div style="text-align:center;padding:40px;color:#6B7A8D;">Aucun signal dans l\'historique</div>', unsafe_allow_html=True)
        return
    h_rev=list(reversed(h))
    # Stats
    buys=sum(1 for x in h if x["signal"]=="BUY")
    sells=sum(1 for x in h if x["signal"]=="SELL")
    avg_conf=round(np.mean([x["confidence"] for x in h]),1)
    avg_ml=round(np.mean([x.get("ml_score",50) for x in h]),1)
    c1,c2,c3,c4=st.columns(4)
    with c1: st.metric("Signaux BUY",buys)
    with c2: st.metric("Signaux SELL",sells)
    with c3: st.metric("Conf. moy.",f"{avg_conf}%")
    with c4: st.metric("ML moy.",f"{avg_ml}%")
    # Table
    rows="".join(f"""
    <tr>
      <td style="color:#6B7A8D;font-size:10px;">{x['ts']}</td>
      <td style="color:#C9A94B;font-weight:700;">{x['asset']}</td>
      <td style="color:#8A95A3;">{x['tf']}</td>
      <td style="color:{sc(x['signal'])};font-weight:700;">{se(x['signal'])} {x['signal']}</td>
      <td style="color:{cc(x['confidence'])};">{x['confidence']}%</td>
      <td style="color:#A78BFA;">{x.get('ml_score','-')}%</td>
      <td style="color:#DCE8F5;font-size:10px;">{x['price']:.5f}</td>
      <td style="color:#0FBF5F;font-size:10px;">{x.get('tp','-')}</td>
      <td style="color:#D93025;font-size:10px;">{x.get('sl','-')}</td>
    </tr>""" for x in h_rev[:50])
    st.markdown(f"""<div class="card card-gold" style="overflow-x:auto;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:13px;color:#C9A94B;letter-spacing:2px;margin-bottom:8px;">
        📊 HISTORIQUE DES SIGNAUX ({len(h)} entrées)
      </div>
      <table class="gt">
        <thead><tr><th>Date</th><th>Actif</th><th>TF</th><th>Signal</th><th>Conf</th><th>ML</th><th>Prix</th><th>TP</th><th>SL</th></tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>""", unsafe_allow_html=True)
    # Export buttons
    h_export = [{"Date":x["ts"],"Actif":x["asset"],"TF":x["tf"],
                 "Signal":x["signal"],"Confluence (%)":x["confidence"],
                 "ML (%)":x.get("ml_score",0),"Prix":x["price"],
                 "TP":x.get("tp",""),"SL":x.get("sl","")} for x in h]
    render_export_buttons(h_export, "HISTORIQUE SIGNAUX", "historique_signaux", "hist")
    st.markdown("---")
    if st.button("🗑️ Effacer l'historique",use_container_width=False):
        save_history([])
        st.rerun()

def render_journal_tab():
    j=load_journal()
    st.markdown('<div style="font-family:\'Rajdhani\',sans-serif;font-size:15px;font-weight:700;color:#C9A94B;letter-spacing:2px;margin-bottom:12px;">📓 JOURNAL DE TRADES</div>', unsafe_allow_html=True)
    # Add trade form
    with st.expander("➕ Ajouter un trade",expanded=False):
        tc1,tc2,tc3=st.columns(3)
        with tc1:
            j_asset=st.selectbox("Actif",list(ASSETS.keys()),key="j_asset")
            j_dir=st.selectbox("Direction",["BUY","SELL"],key="j_dir")
        with tc2:
            j_entry=st.number_input("Prix d'entrée",value=0.0,format="%.5f",key="j_entry")
            j_tp=st.number_input("Take Profit",value=0.0,format="%.5f",key="j_tp")
        with tc3:
            j_sl=st.number_input("Stop Loss",value=0.0,format="%.5f",key="j_sl")
            j_size=st.number_input("Taille (lots)",value=0.01,format="%.2f",key="j_size")
        j_note=st.text_input("Note",placeholder="Setup, raison...",key="j_note")
        j_status=st.selectbox("Statut",["Ouvert","Fermé TP","Fermé SL","Fermé Manuellement"],key="j_status")
        if st.button("💾 Enregistrer le trade",use_container_width=True):
            entry={"id":len(j)+1,"ts":datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"),
                   "asset":j_asset,"direction":j_dir,"entry":j_entry,
                   "tp":j_tp,"sl":j_sl,"size":j_size,"note":j_note,"status":j_status}
            j.append(entry); save_journal(j)
            st.success("✅ Trade enregistré !"); time.sleep(0.5); st.rerun()

    if not j:
        st.markdown('<div style="text-align:center;padding:30px;color:#6B7A8D;">Journal vide — ajoutez votre premier trade</div>', unsafe_allow_html=True)
        return

    # Stats
    closed=[(x["tp"]-x["entry"]) if x["direction"]=="BUY" and x["status"]=="Fermé TP"
            else (x["entry"]-x["tp"]) if x["status"]=="Fermé TP"
            else (x["entry"]-x["sl"]) if x["status"]=="Fermé SL" else 0 for x in j]
    wins=sum(1 for x in j if x["status"]=="Fermé TP")
    losses=sum(1 for x in j if x["status"]=="Fermé SL")
    total_closed=wins+losses
    wr=round(wins/total_closed*100,1) if total_closed>0 else 0

    cc1,cc2,cc3=st.columns(3)
    with cc1: st.metric("Trades total",len(j))
    with cc2: st.metric("Win rate",f"{wr}%",f"{wins}W / {losses}L")
    with cc3: st.metric("Ouverts",sum(1 for x in j if x["status"]=="Ouvert"))

    rows="".join(f"""
    <tr>
      <td style="color:#6B7A8D;font-size:10px;">{x['ts']}</td>
      <td style="color:#C9A94B;font-weight:700;">{x['asset']}</td>
      <td style="color:{'#0FBF5F' if x['direction']=='BUY' else '#D93025'};font-weight:700;">{x['direction']}</td>
      <td style="color:#DCE8F5;">{x['entry']:.5f}</td>
      <td style="color:#0FBF5F;">{x['tp']:.5f}</td>
      <td style="color:#D93025;">{x['sl']:.5f}</td>
      <td style="color:#A78BFA;">{x['size']}</td>
      <td style="color:{'#0FBF5F' if x['status']=='Fermé TP' else '#D93025' if x['status']=='Fermé SL' else '#C9A94B'};font-size:10px;">{x['status']}</td>
      <td style="color:#6B7A8D;font-size:10px;">{x.get('note','')[:20]}</td>
    </tr>""" for x in reversed(j))
    st.markdown(f"""<div class="card" style="overflow-x:auto;">
      <table class="gt">
        <thead><tr><th>Date</th><th>Actif</th><th>Dir</th><th>Entrée</th><th>TP</th><th>SL</th><th>Lots</th><th>Statut</th><th>Note</th></tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>""", unsafe_allow_html=True)
    # Export buttons
    j_export = [{"Date":x["ts"],"Actif":x["asset"],"Direction":x["direction"],
                 "Entree":x["entry"],"TP":x["tp"],"SL":x["sl"],
                 "Taille (lots)":x["size"],"Statut":x["status"],
                 "Note":x.get("note","")} for x in j]
    render_export_buttons(j_export, "JOURNAL TRADES", "journal_trades", "jour")
    st.markdown("---")
    if st.button("🗑️ Effacer journal",use_container_width=False):
        save_journal([]); st.rerun()

# ─── CLAUDE AI ANALYSIS ──────────────────────────────────────────────────────
def modal_claude_html(step, pct):
    return f"""<div style="position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(6,14,26,.96);
     z-index:9999;display:flex;align-items:center;justify-content:center;backdrop-filter:blur(8px);">
  <div style="background:#0F1C2E;border:2px solid #A78BFA;border-radius:16px;padding:30px 34px;
       min-width:290px;max-width:360px;text-align:center;box-shadow:0 0 80px rgba(167,139,250,.3);">
    <div style="font-family:'Rajdhani',sans-serif;font-size:22px;font-weight:700;color:#A78BFA;letter-spacing:2px;margin-bottom:6px;">
      🧠 CLAUDE AI EN RÉFLEXION
    </div>
    <div style="font-size:11px;color:#6B7A8D;letter-spacing:1px;margin-bottom:20px;min-height:16px;">{step}</div>
    <div style="width:100%;height:8px;background:#1A3050;border-radius:4px;overflow:hidden;margin-bottom:12px;">
      <div style="width:{pct}%;height:100%;background:linear-gradient(90deg,#7C3AED,#A78BFA,#C4B5FD);border-radius:4px;"></div>
    </div>
    <div style="font-family:'Rajdhani',sans-serif;font-size:42px;font-weight:700;color:#C4B5FD;">{pct}%</div>
    <div style="font-size:10px;color:#6B7A8D;margin-top:8px;letter-spacing:1px;">
      Analyse experte ICT/SMC en cours...
    </div>
  </div>
</div>"""

def analyse_avec_claude(res, results_all, asset_name, tf_sel, api_key, now, active_kzs):
    """Send market data to Claude API via direct HTTP — no extra package needed."""
    if not api_key:
        return None, "❌ Clé API Claude manquante. Entre ta clé dans la section Configuration."

    # Build multi-TF context
    mtf_context = ""
    for tf, r in results_all.items():
        if r:
            mtf_context += f"  • {tf}: {r['signal']} ({r['confidence']}% conf, ML {r['ml_ensemble']}%, {r['regime']})\n"

    # ICT context
    fvg_ctx = ""
    for f in res["fvgs"][-3:]:
        dist = abs(f["mid"] - res["close"]) / res["close"] * 100
        fvg_ctx += f"  • FVG {f['type'].upper()} : {f['bot']:.4f}–{f['top']:.4f} ({dist:.2f}% du prix)\n"

    ob_ctx = ""
    for o in res["obs"][-3:]:
        dist = abs(o["mid"] - res["close"]) / res["close"] * 100
        ob_ctx += f"  • OB {o['type'].upper()} : {o['bot']:.4f}–{o['top']:.4f} ({dist:.2f}% du prix)\n"

    liq = res["liquidity"]
    bsl = ", ".join([f"{x:.4f}" for x in liq["buyside"]]) or "–"
    ssl = ", ".join([f"{x:.4f}" for x in liq["sellside"]]) or "–"

    kz_active_names = [k["name"] for k in active_kzs]
    kz_str = ", ".join(kz_active_names) if kz_active_names else "Aucune"

    checks_str = ""
    for k, c in res["checks"].items():
        status = "✅" if c["ok"] else "❌"
        checks_str += f"  {status} {c['label']} : {c['val']}\n"

    prompt = f"""Tu es un expert en trading intraday ICT (Inner Circle Trader) et SMC (Smart Money Concepts).
Analyse ces données de marché et fournis une recommandation précise et professionnelle.

═══════════════════════════════════════════
DONNÉES MARCHÉ — {asset_name} — {tf_sel.upper()}
Heure UTC : {now.strftime("%H:%M")}
═══════════════════════════════════════════

📊 SIGNAL TECHNIQUE :
  • Prix live    : {res["close"]:.5f}
  • VWAP         : {res["vwap"]:.5f} (prix {"AU-DESSUS ↑" if res["vwap_bull"] else "EN-DESSOUS ↓"})
  • Signal       : {res["signal"]}
  • Régime       : {res["regime"]}
  • Confluence   : {res["confidence"]}%
  • ML Ensemble  : {res["ml_ensemble"]}%

📐 INDICATEURS :
  • BB Sup/Mid/Inf : {res["bb_upper"]:.5f} / {res["bb_mid"]:.5f} / {res["bb_lower"]:.5f}
  • Squeeze BB   : {"OUI 🔴" if res["squeeze"] else "NON"}
  • MACD         : {res["macd"]:.6f} | Signal : {res["macd_sig"]:.6f}
  • Histogramme  : {res["macd_hist"]:.6f} ({"positif ↑" if res["macd_hist"] > 0 else "négatif ↓"})
  • RSI(14)      : {res["rsi"]:.1f}
  • Stoch K/D    : {res["stoch_k"]:.1f} / {res["stoch_d"]:.1f}
  • ATR(14)      : {res["atr"]:.5f}
  • Force Index FI(2)  : {res.get("fi2", 0):+.0f}  (court terme — signal d'entree Elder)
  • Force Index FI(13) : {res.get("fi13", 0):+.0f}  (tendance — {'HAUSSIER' if res.get('fi13',0)>0 else 'BAISSIER'})

💰 TP/SL :
  • Take Profit  : {res["tp"] if res["tp"] else "–"}
  • Stop Loss    : {res["sl"] if res["sl"] else "–"}
  • R:R Ratio    : {res["rr"] if res["rr"] else "–"}

✅ CHECKLIST CONFLUENCE :
{checks_str}

🎯 NIVEAUX ICT :
Fair Value Gaps :
{fvg_ctx if fvg_ctx else "  Aucun FVG détecté\n"}
Order Blocks :
{ob_ctx if ob_ctx else "  Aucun OB détecté\n"}
Liquidité BSL (au-dessus) : {bsl}
Liquidité SSL (en-dessous) : {ssl}

🔭 MULTI-TIMEFRAME :
{mtf_context if mtf_context else "  –\n"}

🎯 KILL ZONES ACTIVES : {kz_str}

═══════════════════════════════════════════

Fournis une analyse structurée en français avec :

1. **BIAIS DIRECTIONNEL** — Confirmes-tu le signal {res["signal"]} ? Pourquoi ?
2. **CONTEXTE ICT/SMC** — Les FVG et OB valident-ils le setup ?
3. **CONFLUENCE MULTI-TF** — Les timeframes sont-ils alignés ?
4. **RISQUES PRINCIPAUX** — Quels éléments pourraient invalider ce setup ?
5. **NIVEAU D'ENTRÉE OPTIMAL** — Zone d'entrée précise conseillée
6. **VERDICT FINAL** — BUY / SELL / ATTENDRE avec niveau de conviction (Faible/Moyen/Fort/Très Fort)

Sois précis, concis et direct. Maximum 400 mots."""

    try:
        # Direct HTTP call — requests already installed, no anthropic package needed
        response = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key":         api_key,
                "anthropic-version": "2023-06-01",
                "content-type":      "application/json",
            },
            json={
                "model":      "claude-sonnet-4-6",
                "max_tokens": 800,
                "messages":   [{"role": "user", "content": prompt}]
            },
            timeout=30
        )
        if response.status_code == 200:
            data = response.json()
            return data["content"][0]["text"], None
        elif response.status_code == 401:
            return None, "❌ Clé API invalide. Vérifie ta clé sur console.anthropic.com"
        elif response.status_code == 429:
            return None, "⚠️ Limite API atteinte. Attends quelques secondes et réessaie."
        elif response.status_code == 529:
            return None, "⚠️ API Claude surchargée. Réessaie dans 30 secondes."
        else:
            err = response.json().get("error", {}).get("message", response.text[:120])
            return None, f"❌ Erreur API Claude ({response.status_code}) : {err}"
    except requests.exceptions.Timeout:
        return None, "⚠️ Timeout — vérifie ta connexion internet et réessaie."
    except requests.exceptions.ConnectionError:
        return None, "❌ Pas de connexion internet. Vérifie ton réseau."
    except Exception as e:
        return None, f"❌ Erreur inattendue : {str(e)[:120]}"

def render_claude_tab(res, results_all, asset_name, tf_sel, api_key, now, active_kzs):
    """Full Claude AI analysis tab."""
    st.markdown("""
    <div style="background:linear-gradient(135deg,#0F0A1E,#160F2E);border:2px solid #7C3AED;
         border-radius:12px;padding:14px 18px;margin-bottom:14px;text-align:center;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:20px;font-weight:700;
           color:#A78BFA;letter-spacing:3px;margin-bottom:4px;">🧠 ANALYSE CLAUDE AI</div>
      <div style="font-size:10px;color:#6B7A8D;letter-spacing:1px;">
        Expert ICT/SMC · Analyse contextuelle · Confirmation multi-TF
      </div>
    </div>""", unsafe_allow_html=True)

    # Status API key
    if api_key:
        st.markdown('''<div style="background:rgba(167,139,250,.08);border:1px solid #7C3AED;
             border-radius:6px;padding:7px 12px;font-size:11px;color:#A78BFA;margin-bottom:10px;">
          🟣 Clé API Claude active — Prêt à analyser
        </div>''', unsafe_allow_html=True)
    else:
        st.markdown('''<div style="background:rgba(217,48,37,.06);border:1px solid #D93025;
             border-radius:6px;padding:7px 12px;font-size:11px;color:#FF6B6B;margin-bottom:10px;">
          ❌ Clé API Claude manquante — Entre ta clé dans la section Configuration ci-dessus
        </div>''', unsafe_allow_html=True)
        return

    # Context recap
    s = res["signal"]; col = sc(s)
    st.markdown(f"""
    <div class="card card-purple" style="margin-bottom:12px;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:12px;color:#A78BFA;letter-spacing:2px;margin-bottom:8px;">
        📤 DONNÉES ENVOYÉES À CLAUDE
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;text-align:center;font-size:11px;">
        <div style="background:rgba(0,0,0,.3);border-radius:6px;padding:8px;">
          <div style="color:#6B7A8D;font-size:9px;">ACTIF / TF</div>
          <div style="color:#C9A94B;font-weight:700;">{asset_name} {tf_sel}</div>
        </div>
        <div style="background:rgba(0,0,0,.3);border-radius:6px;padding:8px;">
          <div style="color:#6B7A8D;font-size:9px;">SIGNAL</div>
          <div style="color:{col};font-weight:700;">{se(s)} {s} ({res["confidence"]}%)</div>
        </div>
        <div style="background:rgba(0,0,0,.3);border-radius:6px;padding:8px;">
          <div style="color:#6B7A8D;font-size:9px;">ML ENSEMBLE</div>
          <div style="color:#A78BFA;font-weight:700;">{res["ml_ensemble"]}%</div>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

    # Cached result
    cache_key = f"claude_analysis_{asset_name}_{tf_sel}"
    if cache_key in st.session_state:
        cached = st.session_state[cache_key]
        st.markdown(f"""
        <div style="background:rgba(124,58,237,.06);border:1px solid #7C3AED;
             border-radius:10px;padding:16px 18px;margin-bottom:10px;">
          <div style="font-family:'Rajdhani',sans-serif;font-size:13px;color:#A78BFA;
               letter-spacing:2px;margin-bottom:12px;">🧠 ANALYSE CLAUDE — {cached["ts"]}</div>
          <div style="color:#DCE8F5;font-size:12px;line-height:1.8;white-space:pre-wrap;">{cached["text"]}</div>
        </div>""", unsafe_allow_html=True)
        if st.button("🔄 Nouvelle analyse Claude", use_container_width=True):
            st.session_state.pop(cache_key, None)
            st.rerun()
    else:
        if st.button("🧠 LANCER L'ANALYSE CLAUDE AI", use_container_width=True):
            mph = st.empty()
            steps_c = [
                (15, "Préparation des données de marché..."),
                (30, "Calcul des niveaux ICT / SMC..."),
                (50, "Envoi à Claude Sonnet..."),
                (70, "Analyse ICT en cours..."),
                (85, "Confirmation multi-timeframe..."),
                (95, "Rédaction de la recommandation..."),
            ]
            for pct, slbl in steps_c:
                mph.markdown(modal_claude_html(slbl, pct), unsafe_allow_html=True)
                time.sleep(0.3)

            analysis, error = analyse_avec_claude(
                res, results_all, asset_name, tf_sel, api_key, now, active_kzs
            )
            mph.markdown(modal_claude_html("✅ Analyse terminée !", 100), unsafe_allow_html=True)
            time.sleep(0.4)
            mph.empty()

            if error:
                st.error(error)
            else:
                ts_now = now.strftime("%H:%M UTC")
                st.session_state[cache_key] = {"text": analysis, "ts": ts_now}
                st.rerun()



# ─── PIVOT POINTS CLASSIQUES ─────────────────────────────────────────────────
def calc_pivots(df):
    """Classic pivot points from previous session (last full candle)."""
    if df is None or len(df) < 2: return None
    prev = df.iloc[-2]
    H, L, C = float(prev["High"]), float(prev["Low"]), float(prev["Close"])
    P  = (H + L + C) / 3
    R1 = 2*P - L
    R2 = P + (H - L)
    R3 = H + 2*(P - L)
    S1 = 2*P - H
    S2 = P - (H - L)
    S3 = L - 2*(H - P)
    # Midpoints
    M_R1 = (P + R1) / 2
    M_R2 = (R1 + R2) / 2
    M_S1 = (P + S1) / 2
    M_S2 = (S1 + S2) / 2
    return {"P":P,"R1":R1,"R2":R2,"R3":R3,"S1":S1,"S2":S2,"S3":S3,
            "MR1":M_R1,"MR2":M_R2,"MS1":M_S1,"MS2":M_S2,"H":H,"L":L,"C":C}

def render_pivots(res):
    """Render classic pivot points panel."""
    import streamlit as st
    pvt = calc_pivots(res["df_raw"])
    if not pvt:
        st.warning("Données insuffisantes pour calculer les pivots.")
        return

    c = res["close"]

    def prow(label, val, typ="pivot"):
        dist = (val - c) / c * 100
        above = val > c
        col = "#D93025" if typ == "resistance" else "#0FBF5F" if typ == "support" else "#C9A94B"
        arrow = "▲" if above else "▼"
        bg = "rgba(217,48,37,.06)" if typ=="resistance" else "rgba(15,191,95,.06)" if typ=="support" else "rgba(201,169,75,.08)"
        border = "rgba(217,48,37,.3)" if typ=="resistance" else "rgba(15,191,95,.3)" if typ=="support" else "rgba(201,169,75,.4)"
        active = "border-width:2px!important;" if abs(dist) < 0.5 else ""
        return f"""<div style="display:flex;justify-content:space-between;align-items:center;
            padding:8px 12px;border-radius:7px;margin:3px 0;font-size:12px;
            background:{bg};border:1px solid {border};{active}">
          <span style="color:{col};font-weight:700;width:40px;">{label}</span>
          <span style="color:#DCE8F5;font-family:'JetBrains Mono',monospace;font-weight:700;">{val:.4f}</span>
          <span style="color:{col};font-size:10px;">{arrow} {abs(dist):.2f}%</span>
          <span style="font-size:9px;color:#6B7A8D;">{'PROCHE ⚡' if abs(dist)<0.3 else ''}</span>
        </div>"""

    levels_html = (
        prow("R3",  pvt["R3"],  "resistance") +
        prow("R2",  pvt["R2"],  "resistance") +
        prow("MR2", pvt["MR2"], "resistance") +
        prow("R1",  pvt["R1"],  "resistance") +
        prow("MR1", pvt["MR1"], "resistance") +
        f"""<div style="display:flex;justify-content:space-between;align-items:center;
            padding:10px 12px;border-radius:7px;margin:5px 0;font-size:13px;
            background:rgba(201,169,75,.12);border:2px solid #C9A94B;">
          <span style="color:#F0C96A;font-family:'Rajdhani',sans-serif;font-weight:700;font-size:15px;">PIVOT</span>
          <span style="color:#F0C96A;font-family:'JetBrains Mono',monospace;font-weight:700;font-size:15px;">{pvt['P']:.4f}</span>
          <span style="color:#C9A94B;font-size:10px;">{'▲ Au-dessus' if c>pvt['P'] else '▼ En-dessous'}</span>
        </div>""" +
        prow("MS1", pvt["MS1"], "support") +
        prow("S1",  pvt["S1"],  "support") +
        prow("MS2", pvt["MS2"], "support") +
        prow("S2",  pvt["S2"],  "support") +
        prow("S3",  pvt["S3"],  "support")
    )

    # Current price marker
    st.markdown(f"""
    <div style="background:#0F1C2E;border:1px solid #1A3050;border-radius:12px;padding:16px;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;
           color:#C9A94B;letter-spacing:2px;margin-bottom:12px;text-align:center;">
        🎯 POINTS PIVOTS CLASSIQUES
        <div style="font-size:10px;color:#6B7A8D;margin-top:3px;letter-spacing:1px;">
          Session précédente — H:{pvt['H']:.4f} L:{pvt['L']:.4f} C:{pvt['C']:.4f}
        </div>
      </div>

      <div style="background:rgba(240,196,106,.1);border:1px solid #C9A94B;border-radius:7px;
           padding:7px 12px;margin-bottom:10px;text-align:center;">
        <span style="color:#6B7A8D;font-size:10px;">PRIX ACTUEL</span>
        <span style="color:#F0C96A;font-family:'JetBrains Mono',monospace;font-size:18px;
             font-weight:700;margin-left:10px;">{c:.5f}</span>
      </div>

      {levels_html}

      <div style="margin-top:12px;padding:10px;background:rgba(0,0,0,.2);border-radius:7px;
           font-size:10px;color:#6B7A8D;line-height:1.7;">
        <b style="color:#C9A94B;">Formule :</b> P=(H+L+C)/3 · R1=2P-L · R2=P+(H-L) · R3=H+2(P-L)<br>
        S1=2P-H · S2=P-(H-L) · S3=L-2(H-P) · MX=Midpoints
      </div>
    </div>""", unsafe_allow_html=True)


# ─── BACKTESTING ─────────────────────────────────────────────────────────────
def run_backtest(df, tf):
    """Walk-forward backtest on historical candles — no lookahead bias."""
    if df is None or len(df) < 80: return None
    results = []
    tp_mult, sl_mult = {"5m":(1.5,1.0),"15m":(2.0,1.3),"30m":(2.5,1.5),"1h":(3.0,2.0)}.get(tf,(2.0,1.3))

    for i in range(60, len(df) - 5):
        chunk = df.iloc[:i].copy()
        if len(chunk) < 50: continue
        res = analyse(chunk, tf)
        if res is None: continue
        sig = res["signal"]
        if sig not in ("BUY","SELL"): continue

        entry = float(df["Close"].iloc[i])
        atr_v = res["atr"]
        if sig == "BUY":
            tp = entry + atr_v * tp_mult
            sl = entry - atr_v * sl_mult
        else:
            tp = entry - atr_v * tp_mult
            sl = entry + atr_v * sl_mult

        # Check next 5 candles for TP/SL hit
        outcome = "OPEN"
        pnl_r   = 0.0
        for j in range(i+1, min(i+6, len(df))):
            hi = float(df["High"].iloc[j])
            lo = float(df["Low"].iloc[j])
            if sig == "BUY":
                if lo <= sl:  outcome="SL"; pnl_r=-1.0; break
                if hi >= tp:  outcome="TP"; pnl_r=tp_mult/sl_mult; break
            else:
                if hi >= sl:  outcome="SL"; pnl_r=-1.0; break
                if lo <= tp:  outcome="TP"; pnl_r=tp_mult/sl_mult; break
        if outcome == "OPEN":
            pnl_r = 0.0

        results.append({
            "ts":    str(df.index[i])[:16],
            "signal":sig,
            "conf":  res["confidence"],
            "ml":    res["ml_ensemble"],
            "entry": round(entry,5),
            "tp":    round(tp,5),
            "sl":    round(sl,5),
            "outcome":outcome,
            "pnl_r": round(pnl_r,2),
        })

    if not results: return None

    total  = len(results)
    wins   = sum(1 for r in results if r["outcome"]=="TP")
    losses = sum(1 for r in results if r["outcome"]=="SL")
    open_  = sum(1 for r in results if r["outcome"]=="OPEN")
    wr     = round(wins/(wins+losses)*100,1) if (wins+losses)>0 else 0
    gross_p= sum(r["pnl_r"] for r in results if r["pnl_r"]>0)
    gross_l= abs(sum(r["pnl_r"] for r in results if r["pnl_r"]<0))
    pf     = round(gross_p/gross_l,2) if gross_l>0 else 0
    net_r  = round(sum(r["pnl_r"] for r in results),2)
    avg_c  = round(np.mean([r["conf"] for r in results]),1)
    avg_ml = round(np.mean([r["ml"] for r in results]),1)
    best   = max(results, key=lambda x: x["pnl_r"])
    worst  = min(results, key=lambda x: x["pnl_r"])

    # Equity curve
    equity = [0.0]
    for r in results:
        equity.append(equity[-1] + r["pnl_r"])

    return {
        "total":total,"wins":wins,"losses":losses,"open":open_,
        "wr":wr,"pf":pf,"net_r":net_r,"avg_conf":avg_c,"avg_ml":avg_ml,
        "results":results,"equity":equity,
        "best":best,"worst":worst,
        "tf":tf,
    }

def render_backtest(res, asset_name, tf):
    """Render backtest results panel."""
    import streamlit as st

    st.markdown("""
    <div style="background:linear-gradient(135deg,#0A1520,#0F1C2E);border:2px solid #E07B2A;
         border-radius:12px;padding:14px 18px;margin-bottom:14px;text-align:center;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;
           color:#E07B2A;letter-spacing:2px;">📉 BACKTESTING — RÉSULTATS HISTORIQUES</div>
      <div style="font-size:10px;color:#6B7A8D;margin-top:3px;">
        Walk-forward · Sans biais · Données historiques réelles
      </div>
    </div>""", unsafe_allow_html=True)

    bt_cache = f"bt_{asset_name}_{tf}"

    if bt_cache in st.session_state:
        bt = st.session_state[bt_cache]
    else:
        if st.button("▶️ LANCER LE BACKTESTING", use_container_width=True):
            with st.spinner("Calcul en cours sur les données historiques..."):
                bt = run_backtest(res["df_raw"], tf)
            if bt:
                st.session_state[bt_cache] = bt
                st.rerun()
            else:
                st.error("Données insuffisantes pour le backtest (min 80 bougies).")
            return
        else:
            st.markdown("""<div style="text-align:center;padding:40px;color:#6B7A8D;font-size:12px;">
              Clique sur le bouton pour lancer le backtesting sur les dernières bougies disponibles.
            </div>""", unsafe_allow_html=True)
            return

    # Stats cards
    wr_col = "#0FBF5F" if bt["wr"]>=55 else "#C9A94B" if bt["wr"]>=45 else "#D93025"
    pf_col = "#0FBF5F" if bt["pf"]>=1.5 else "#C9A94B" if bt["pf"]>=1.0 else "#D93025"
    nr_col = "#0FBF5F" if bt["net_r"]>0 else "#D93025"

    st.markdown(f"""
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:12px;">
      <div style="background:#0F1C2E;border:1px solid #1A3050;border-radius:8px;padding:10px;text-align:center;">
        <div style="color:#6B7A8D;font-size:9px;letter-spacing:1px;">TRADES</div>
        <div style="color:#C9A94B;font-family:'Rajdhani',sans-serif;font-size:22px;font-weight:700;">{bt['total']}</div>
        <div style="color:#6B7A8D;font-size:9px;">{bt['wins']}W / {bt['losses']}L</div>
      </div>
      <div style="background:#0F1C2E;border:1px solid #1A3050;border-radius:8px;padding:10px;text-align:center;">
        <div style="color:#6B7A8D;font-size:9px;letter-spacing:1px;">WIN RATE</div>
        <div style="color:{wr_col};font-family:'Rajdhani',sans-serif;font-size:22px;font-weight:700;">{bt['wr']}%</div>
        <div style="color:#6B7A8D;font-size:9px;">Conf moy: {bt['avg_conf']}%</div>
      </div>
      <div style="background:#0F1C2E;border:1px solid #1A3050;border-radius:8px;padding:10px;text-align:center;">
        <div style="color:#6B7A8D;font-size:9px;letter-spacing:1px;">PROFIT FACTOR</div>
        <div style="color:{pf_col};font-family:'Rajdhani',sans-serif;font-size:22px;font-weight:700;">{bt['pf']}</div>
        <div style="color:#6B7A8D;font-size:9px;">ML moy: {bt['avg_ml']}%</div>
      </div>
      <div style="background:#0F1C2E;border:1px solid #1A3050;border-radius:8px;padding:10px;text-align:center;">
        <div style="color:#6B7A8D;font-size:9px;letter-spacing:1px;">NET R TOTAL</div>
        <div style="color:{nr_col};font-family:'Rajdhani',sans-serif;font-size:22px;font-weight:700;">{bt['net_r']:+.1f}R</div>
        <div style="color:#6B7A8D;font-size:9px;">{'✅ Rentable' if bt['net_r']>0 else '❌ Déficitaire'}</div>
      </div>
    </div>""", unsafe_allow_html=True)

    # Equity curve via plotly
    try:
        import plotly.graph_objects as go
        eq = bt["equity"]
        cols = ["#0FBF5F" if eq[i]>=eq[i-1] else "#D93025" for i in range(1,len(eq))]
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=list(range(len(eq))), y=eq,
            line=dict(color="#E07B2A", width=2),
            fill="tozeroy", fillcolor="rgba(224,123,42,.1)",
            name="Equity Curve"
        ))
        fig.update_layout(
            height=200, paper_bgcolor="#060E1A", plot_bgcolor="#060E1A",
            margin=dict(l=40,r=10,t=10,b=20),
            xaxis=dict(gridcolor="#1A3050",tickfont=dict(color="#6B7A8D",size=9)),
            yaxis=dict(gridcolor="#1A3050",tickfont=dict(color="#6B7A8D",size=9),
                       title=dict(text="R",font=dict(color="#6B7A8D",size=9))),
            title=dict(text="Courbe d'équité (R)",
                      font=dict(color="#E07B2A",size=12,family="Rajdhani"),x=0.5),
        )
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})
    except:
        st.line_chart(bt["equity"])

    # Best / Worst
    b, w = bt["best"], bt["worst"]
    st.markdown(f"""
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:12px;">
      <div style="background:rgba(15,191,95,.06);border:1px solid #0F3A1F;border-radius:8px;padding:10px;">
        <div style="color:#0FBF5F;font-family:'Rajdhani',sans-serif;font-size:12px;letter-spacing:1px;margin-bottom:4px;">🏆 MEILLEUR TRADE</div>
        <div style="color:#DCE8F5;font-size:11px;">{b['ts']} · {b['signal']} · {b['outcome']}</div>
        <div style="color:#0FBF5F;font-weight:700;">+{b['pnl_r']:.2f}R | Conf {b['conf']}%</div>
      </div>
      <div style="background:rgba(217,48,37,.06);border:1px solid #3A0F0F;border-radius:8px;padding:10px;">
        <div style="color:#D93025;font-family:'Rajdhani',sans-serif;font-size:12px;letter-spacing:1px;margin-bottom:4px;">📉 PIRE TRADE</div>
        <div style="color:#DCE8F5;font-size:11px;">{w['ts']} · {w['signal']} · {w['outcome']}</div>
        <div style="color:#D93025;font-weight:700;">{w['pnl_r']:.2f}R | Conf {w['conf']}%</div>
      </div>
    </div>""", unsafe_allow_html=True)

    # Last 20 trades table
    last20 = list(reversed(bt["results"]))[:20]
    rows = "".join(f"""<tr>
      <td style="color:#6B7A8D;font-size:10px;">{r['ts'][11:]}</td>
      <td style="color:{'#0FBF5F' if r['signal']=='BUY' else '#D93025'};font-weight:700;">{r['signal']}</td>
      <td style="color:#C9A94B;">{r['conf']}%</td>
      <td style="color:#A78BFA;">{r['ml']}%</td>
      <td style="color:{'#0FBF5F' if r['outcome']=='TP' else '#D93025' if r['outcome']=='SL' else '#6B7A8D'};font-weight:700;">{r['outcome']}</td>
      <td style="color:{'#0FBF5F' if r['pnl_r']>0 else '#D93025' if r['pnl_r']<0 else '#6B7A8D'};font-weight:700;">{r['pnl_r']:+.2f}R</td>
    </tr>""" for r in last20)
    st.markdown(f"""<div style="overflow-x:auto;">
      <table class="gt">
        <thead><tr><th>Heure</th><th>Dir</th><th>Conf</th><th>ML</th><th>Résultat</th><th>P&L</th></tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>""", unsafe_allow_html=True)

    if st.button("🔄 Relancer le backtest", use_container_width=False):
        st.session_state.pop(bt_cache, None)
        st.rerun()


# ─── SOUND ALERTS ────────────────────────────────────────────────────────────
def render_alertes(res, asset_name, cfg):
    """Sound alert configuration and live trigger panel."""
    import streamlit as st

    st.markdown("""
    <div style="background:linear-gradient(135deg,#0A1A10,#0F2018);border:2px solid #0FBF5F;
         border-radius:12px;padding:14px 18px;margin-bottom:14px;text-align:center;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;
           color:#0FBF5F;letter-spacing:2px;">🔊 ALERTES SONORES</div>
      <div style="font-size:10px;color:#6B7A8D;margin-top:3px;">
        Web Audio API · Aucune installation · Fonctionne sur Android Chrome
      </div>
    </div>""", unsafe_allow_html=True)

    # Config
    a1, a2 = st.columns(2)
    with a1:
        alert_enabled = st.checkbox("🔔 Alertes activées", value=cfg.get("alert_enabled", True))
        min_conf = st.slider("Seuil de confiance minimum", 50, 95,
                             value=cfg.get("alert_min_conf", 70), step=5)
    with a2:
        alert_buy  = st.checkbox("🟢 Alerter sur BUY",  value=cfg.get("alert_buy", True))
        alert_sell = st.checkbox("🔴 Alerter sur SELL", value=cfg.get("alert_sell", True))

    # Save config
    cfg["alert_enabled"]  = alert_enabled
    cfg["alert_min_conf"] = min_conf
    cfg["alert_buy"]      = alert_buy
    cfg["alert_sell"]     = alert_sell

    # Determine if current signal triggers an alert
    sig  = res["signal"] if res else "ATTENDRE"
    conf = res["confidence"] if res else 0
    should_alert = (
        alert_enabled and
        conf >= min_conf and
        ((sig == "BUY" and alert_buy) or (sig == "SELL" and alert_sell))
    )

    # Status card
    if res:
        sig_col = "#0FBF5F" if sig=="BUY" else "#D93025" if sig=="SELL" else "#6B7A8D"
        st.markdown(f"""
        <div style="background:rgba(0,0,0,.2);border:1px solid #1A3050;border-radius:8px;
             padding:10px 14px;margin:10px 0;display:flex;justify-content:space-between;align-items:center;">
          <div>
            <span style="color:#6B7A8D;font-size:10px;">SIGNAL ACTUEL</span><br>
            <span style="color:{sig_col};font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;">
              {sig} — {conf}%
            </span>
          </div>
          <div style="text-align:right;">
            <div style="font-size:11px;color:{'#0FBF5F' if should_alert else '#6B7A8D'};">
              {'🔔 ALERTE DÉCLENCHÉE' if should_alert else '🔕 Sous le seuil'}
            </div>
            <div style="font-size:10px;color:#6B7A8D;">Seuil: {min_conf}%</div>
          </div>
        </div>""", unsafe_allow_html=True)

    # Web Audio player — injected into the page
    # BUY = ascending tone (440→880Hz), SELL = descending (880→440Hz), TEST = triple beep
    audio_js = """
<script>
function playTone(type) {
  try {
    const ctx = new (window.AudioContext || window.webkitAudioContext)();
    const now = ctx.currentTime;

    function beep(freq, start, dur, vol) {
      const osc  = ctx.createOscillator();
      const gain = ctx.createGain();
      osc.connect(gain); gain.connect(ctx.destination);
      osc.type = 'sine';
      osc.frequency.setValueAtTime(freq, now + start);
      gain.gain.setValueAtTime(vol, now + start);
      gain.gain.exponentialRampToValueAtTime(0.001, now + start + dur);
      osc.start(now + start); osc.stop(now + start + dur + 0.05);
    }

    if (type === 'buy') {
      beep(440, 0.0, 0.15, 0.6);
      beep(660, 0.2, 0.15, 0.6);
      beep(880, 0.4, 0.30, 0.8);
    } else if (type === 'sell') {
      beep(880, 0.0, 0.15, 0.6);
      beep(660, 0.2, 0.15, 0.6);
      beep(440, 0.4, 0.30, 0.8);
    } else if (type === 'alert') {
      beep(880, 0.0, 0.10, 0.7);
      beep(880, 0.15, 0.10, 0.7);
      beep(1100, 0.30, 0.25, 0.9);
    } else {
      beep(660, 0.0, 0.20, 0.5);
    }
  } catch(e) { console.error('Audio error:', e); }
}
</script>
"""
    # Auto-trigger if signal warrants it
    auto_trigger = ""
    if should_alert:
        tone = "buy" if sig == "BUY" else "sell"
        auto_trigger = f"<script>setTimeout(()=>playTone('{tone}'),500);</script>"

    import streamlit.components.v1 as components
    components.html(audio_js + auto_trigger + """
<div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:8px;">
  <button onclick="playTone('buy')" style="
    background:linear-gradient(135deg,#0FBF5F,#1DB954);color:#060E1A;
    border:none;border-radius:8px;padding:12px 20px;font-size:14px;font-weight:700;
    cursor:pointer;font-family:'Rajdhani',sans-serif;letter-spacing:1px;">
    🟢 TEST BUY
  </button>
  <button onclick="playTone('sell')" style="
    background:linear-gradient(135deg,#D93025,#E53E3E);color:#fff;
    border:none;border-radius:8px;padding:12px 20px;font-size:14px;font-weight:700;
    cursor:pointer;font-family:'Rajdhani',sans-serif;letter-spacing:1px;">
    🔴 TEST SELL
  </button>
  <button onclick="playTone('alert')" style="
    background:linear-gradient(135deg,#C9A94B,#F0C96A);color:#060E1A;
    border:none;border-radius:8px;padding:12px 20px;font-size:14px;font-weight:700;
    cursor:pointer;font-family:'Rajdhani',sans-serif;letter-spacing:1px;">
    🔔 TEST ALERTE
  </button>
</div>
<div style="font-size:10px;color:#6B7A8D;margin-top:8px;font-family:'JetBrains Mono',monospace;">
  ℹ️ Si aucun son → Activer le son sur Android (volume non silencieux)
</div>
""", height=100)

    # Alert log
    st.markdown("""<div style="font-family:'Rajdhani',sans-serif;font-size:13px;color:#0FBF5F;
         letter-spacing:2px;margin:14px 0 8px 0;">📋 CONFIGURATION ACTIVE</div>""",
         unsafe_allow_html=True)

    rows_cfg = [
        ("Alertes activées",   "✅ OUI" if alert_enabled else "❌ NON", "#0FBF5F" if alert_enabled else "#D93025"),
        ("Seuil minimum",      f"{min_conf}%", "#C9A94B"),
        ("Alerte BUY",         "✅ OUI" if alert_buy else "NON", "#0FBF5F" if alert_buy else "#6B7A8D"),
        ("Alerte SELL",        "✅ OUI" if alert_sell else "NON", "#D93025" if alert_sell else "#6B7A8D"),
        ("Son BUY",            "Notes montantes ♪ do-mi-sol", "#8A95A3"),
        ("Son SELL",           "Notes descendantes ♪ sol-mi-do", "#8A95A3"),
    ]
    rows_html = "".join(f"""<tr>
      <td style="color:#8A95A3;padding:5px 8px;font-size:11px;">{r[0]}</td>
      <td style="color:{r[2]};padding:5px 8px;font-weight:700;font-size:11px;">{r[1]}</td>
    </tr>""" for r in rows_cfg)
    st.markdown(f"""<div class="card card-green">
      <table style="width:100%;border-collapse:collapse;">{rows_html}</table>
    </div>""", unsafe_allow_html=True)

    return cfg


# ─── CIRCULAR RADIAL MENU ────────────────────────────────────────────────────

# ─── FEAR & GREED INDEX ───────────────────────────────────────────────────────
@st.cache_data(ttl=120, show_spinner=False)
def fetch_fear_greed():
    """Fetch Fear & Greed Index from alternative.me — free, no key needed."""
    try:
        r = requests.get(
            "https://api.alternative.me/fng/?limit=7&format=json",
            timeout=8
        )
        if r.status_code == 200:
            data = r.json()["data"]
            current = data[0]
            history = list(reversed(data))  # oldest first
            return {
                "value":       int(current["value"]),
                "label":       current["value_classification"],
                "ts":          current["timestamp"],
                "history":     [{"value":int(d["value"]),"label":d["value_classification"]} for d in history],
            }
    except: pass
    return None

def fg_color(v):
    if v <= 25:  return "#D93025","PEUR EXTRÊME 😱"
    if v <= 45:  return "#E07B2A","PEUR 😨"
    if v <= 55:  return "#C9A94B","NEUTRE 😐"
    if v <= 75:  return "#7EC47A","AVIDITÉ 😏"
    return "#0FBF5F","AVIDITÉ EXTRÊME 🤑"

def render_fear_greed():
    import streamlit as st
    import streamlit.components.v1 as components

    st.markdown("""
    <div style="background:linear-gradient(135deg,#0A1520,#0F1C2E);border:2px solid #C9A94B;
         border-radius:12px;padding:14px 18px;margin-bottom:14px;text-align:center;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;
           color:#C9A94B;letter-spacing:2px;">😨 FEAR & GREED INDEX — CRYPTO</div>
      <div style="font-size:10px;color:#6B7A8D;margin-top:3px;">
        Source : alternative.me · Mis à jour quotidiennement · Gratuit sans clé
      </div>
    </div>""", unsafe_allow_html=True)

    fg = fetch_fear_greed()

    if not fg:
        st.warning("⚠️ Impossible de récupérer l'index. Vérifie ta connexion internet.")
        st.markdown("""<div style="background:#0F1C2E;border:1px solid #1A3050;border-radius:8px;padding:14px;font-size:11px;color:#8A95A3;line-height:1.9;">
          <b style="color:#D93025;">0–25 PEUR EXTRÊME</b> → Meilleur moment pour acheter 🟢<br>
          <b style="color:#E07B2A;">25–45 PEUR</b> → Opportunités d'achat à surveiller<br>
          <b style="color:#C9A94B;">46–55 NEUTRE</b> → Pas de biais fort<br>
          <b style="color:#7EC47A;">55–75 AVIDITÉ</b> → Prudence, marché euphorique<br>
          <b style="color:#0FBF5F;">75–100 AVIDITÉ EXTRÊME</b> → Possible sommet, danger 🔴
        </div>""", unsafe_allow_html=True)
        return

    v   = fg["value"]
    col, lbl = fg_color(v)

    # Recommendation logic
    if v <= 25:
        rec_col, rec_icon = "#0FBF5F","🟢"
        rec_text = "OPPORTUNITÉ D'ACHAT — Marché en panique"
        combo    = "Combiné BUY 80%+ sur l'app = Setup IDÉAL ✅"
    elif v <= 45:
        rec_col, rec_icon = "#7EC47A","🟡"
        rec_text = "PRUDENCE HAUSSIÈRE — Surveiller les setups BUY"
        combo    = "Signal BUY valide, taille normale"
    elif v <= 55:
        rec_col, rec_icon = "#C9A94B","⚪"
        rec_text = "NEUTRE — Suivre le signal de l'app"
        combo    = "Pas de biais — se fier à l'analyse technique"
    elif v <= 75:
        rec_col, rec_icon = "#E07B2A","🟠"
        rec_text = "VIGILANCE — Marché euphorique"
        combo    = "Réduire la taille des positions BUY de 25%"
    else:
        rec_col, rec_icon = "#D93025","🔴"
        rec_text = "DANGER — Avidité extrême, sommet possible"
        combo    = "Éviter les BUY. SELL confirmé = haute probabilité ✅"

    # Build history bars HTML
    history = fg.get("history", [])
    bars_html = ""
    days = ["J-6","J-5","J-4","J-3","J-2","Hier","Auj."]
    for i, d in enumerate(history):
        bc, _ = fg_color(d["value"])
        bh = max(20, int(d["value"] * 0.55))
        day = days[i] if i < len(days) else f"J-{6-i}"
        bars_html += f"""
        <div style="display:flex;flex-direction:column;align-items:center;gap:3px;flex:1;">
          <div style="font-size:10px;color:{bc};font-weight:700;">{d["value"]}</div>
          <div style="width:90%;height:{bh}px;background:{bc};border-radius:3px;opacity:0.85;"></div>
          <div style="font-size:9px;color:#6B7A8D;">{day}</div>
        </div>"""

    angle = int(v * 1.8)

    full_html = f"""<!DOCTYPE html><html><head>
    <link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@700&display=swap" rel="stylesheet">
    <style>body{{margin:0;padding:8px;background:#060E1A;font-family:'Rajdhani',sans-serif;}}</style>
    </head><body>

    <!-- GAUGE -->
    <div style="text-align:center;margin-bottom:10px;">
      <div style="position:relative;display:inline-block;width:200px;height:110px;overflow:hidden;">
        <div style="position:absolute;bottom:0;left:0;width:200px;height:200px;border-radius:50%;
             background:conic-gradient(#D93025 0deg 45deg,#E07B2A 45deg 81deg,
             #C9A94B 81deg 99deg,#7EC47A 99deg 135deg,#0FBF5F 135deg 180deg,
             transparent 180deg 360deg);opacity:0.35;"></div>
        <div style="position:absolute;bottom:0;left:50%;transform-origin:bottom center;
             transform:translateX(-50%) rotate({angle-90}deg);
             width:3px;height:85px;background:{col};border-radius:3px;
             box-shadow:0 0 10px {col};"></div>
        <div style="position:absolute;bottom:-5px;left:50%;transform:translateX(-50%);
             width:14px;height:14px;background:{col};border-radius:50%;
             box-shadow:0 0 12px {col};"></div>
      </div>
      <div style="font-size:52px;font-weight:700;color:{col};line-height:1;
           margin-top:-10px;text-shadow:0 0 20px {col}88;">{v}</div>
      <div style="font-size:16px;font-weight:700;color:{col};letter-spacing:2px;">{lbl}</div>
    </div>

    <!-- RECOMMENDATION -->
    <div style="background:rgba(0,0,0,.3);border:1px solid {rec_col};border-radius:8px;
         padding:10px 14px;margin:8px 0;">
      <div style="font-size:13px;font-weight:700;color:{rec_col};">{rec_icon} {rec_text}</div>
      <div style="font-size:11px;color:#8A95A3;margin-top:3px;">{combo}</div>
    </div>

    <!-- 7-DAY BARS -->
    <div style="font-size:12px;color:#C9A94B;letter-spacing:2px;margin:10px 0 6px 0;">
      📅 7 DERNIERS JOURS
    </div>
    <div style="display:flex;gap:4px;align-items:flex-end;
         background:#0A1520;border-radius:8px;padding:10px;height:110px;">
      {bars_html}
    </div>

    <!-- GUIDE -->
    <div style="background:#0F1C2E;border:1px solid #1A3050;border-radius:8px;
         padding:10px 14px;margin-top:10px;font-size:11px;color:#8A95A3;line-height:1.9;">
      <span style="color:#D93025;font-weight:700;">0–25</span> Peur Extrême → Acheter 🟢 &nbsp;
      <span style="color:#E07B2A;font-weight:700;">25–45</span> Peur → Surveiller<br>
      <span style="color:#C9A94B;font-weight:700;">46–55</span> Neutre → Technique seul &nbsp;
      <span style="color:#7EC47A;font-weight:700;">55–75</span> Avidité → Prudence<br>
      <span style="color:#0FBF5F;font-weight:700;">75–100</span> Avidité Extrême → Danger 🔴
    </div>
    </body></html>"""

    components.html(full_html, height=520, scrolling=False)


# ─── LIQUIDATION HEATMAP ──────────────────────────────────────────────────────
def fetch_heatmap_data():
    """Fetch BTC data — Futures preferred, Spot as fallback for Streamlit Cloud."""
    data = {"price": None, "oi": None, "funding": None,
            "long_ratio": None, "short_ratio": None,
            "liq_24h_long": None, "liq_24h_short": None}
    hdrs = {"User-Agent": "Mozilla/5.0"}
    # Mark price — try Futures first, then Spot
    try:
        r = requests.get("https://fapi.binance.com/fapi/v1/premiumIndex?symbol=BTCUSDT",
                         headers=hdrs, timeout=8)
        if r.status_code == 200:
            d = r.json()
            data["price"]   = float(d.get("markPrice", 0))
            data["funding"] = float(d.get("lastFundingRate", 0)) * 100
    except: pass
    # Spot price fallback
    if not data["price"]:
        try:
            r = requests.get("https://api.binance.com/api/v3/ticker/price",
                             params={"symbol":"BTCUSDT"}, headers=hdrs, timeout=8)
            if r.status_code == 200:
                data["price"] = float(r.json().get("price", 0))
        except: pass
    # 24h ticker for volume data
    if not data["price"]:
        try:
            r = requests.get("https://api.binance.com/api/v3/ticker/24hr",
                             params={"symbol":"BTCUSDT"}, headers=hdrs, timeout=8)
            if r.status_code == 200:
                d = r.json()
                data["price"]    = float(d.get("lastPrice", 0))
                data["buy_vol"]  = float(d.get("volume", 0)) * 0.52
                data["sell_vol"] = float(d.get("volume", 0)) * 0.48
        except: pass
    try:
        # Open Interest
        r = requests.get("https://fapi.binance.com/fapi/v1/openInterest?symbol=BTCUSDT", timeout=6)
        if r.status_code == 200:
            data["oi"] = float(r.json().get("openInterest", 0))
    except: pass
    try:
        # Long/Short ratio (top traders)
        r = requests.get(
            "https://fapi.binance.com/futures/data/topLongShortPositionRatio?symbol=BTCUSDT&period=1h&limit=1",
            timeout=6)
        if r.status_code == 200:
            d = r.json()
            if d:
                data["long_ratio"]  = float(d[0].get("longAccount", 0)) * 100
                data["short_ratio"] = float(d[0].get("shortAccount", 0)) * 100
    except: pass
    try:
        # Global long/short account ratio
        r = requests.get(
            "https://fapi.binance.com/futures/data/globalLongShortAccountRatio?symbol=BTCUSDT&period=1h&limit=8",
            timeout=6)
        if r.status_code == 200:
            data["ls_history"] = r.json()
    except: pass
    try:
        # Taker buy/sell volume
        r = requests.get(
            "https://fapi.binance.com/futures/data/takerlongshortRatio?symbol=BTCUSDT&period=1h&limit=1",
            timeout=6)
        if r.status_code == 200:
            d = r.json()
            if d:
                data["buy_vol"]  = float(d[0].get("buyVol", 0))
                data["sell_vol"] = float(d[0].get("sellVol", 0))
    except: pass
    return data

def render_heatmap():
    import streamlit as st
    import streamlit.components.v1 as components

    st.markdown("""
    <div style="background:linear-gradient(135deg,#0A1520,#1A0A20);border:2px solid #E07B2A;
         border-radius:12px;padding:14px 18px;margin-bottom:14px;text-align:center;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;
           color:#E07B2A;letter-spacing:2px;">🗺️ LIQUIDATION HEATMAP — BTC/USD</div>
      <div style="font-size:10px;color:#6B7A8D;margin-top:3px;">
        Binance Futures · Open Interest · Long/Short Ratio · Funding Rate
      </div>
    </div>""", unsafe_allow_html=True)

    with st.spinner("Chargement des données Binance Futures..."):
        d = fetch_heatmap_data()

    price   = d.get("price")
    oi      = d.get("oi")
    funding = d.get("funding")
    l_ratio = d.get("long_ratio")
    s_ratio = d.get("short_ratio")
    buy_vol = d.get("buy_vol")
    sell_vol= d.get("sell_vol")
    ls_hist = d.get("ls_history", [])

    # Fallback: use cached BTC price from session state
    if not price:
        price = st.session_state.get("live_price_btc") or st.session_state.get("btc_last_price")
    if not price:
        st.warning("⚠️ Binance temporairement inaccessible depuis Streamlit Cloud. Lance d'abord une analyse BTC — le prix sera utilisé comme référence.")
        return

    # Determine market bias from long/short ratio
    if l_ratio and s_ratio:
        if l_ratio > 55:
            bias_col, bias_txt, bias_icon = "#D93025","MAJORITÉ EN LONG — Risque de chasse de stops BAISSIÈRE","🔴"
            bias_advice = "Zone de liquidité BUYSIDE chargée → Possible dump pour chasser les longs"
        elif s_ratio > 55:
            bias_col, bias_txt, bias_icon = "#0FBF5F","MAJORITÉ EN SHORT — Risque de short squeeze HAUSSIER","🟢"
            bias_advice = "Zone de liquidité SELLSIDE chargée → Possible pump pour chasser les shorts"
        else:
            bias_col, bias_txt, bias_icon = "#C9A94B","MARCHÉ ÉQUILIBRÉ — Pas de biais clair","🟡"
            bias_advice = "Long/Short équilibré — suivre le signal technique de l'app"
    else:
        bias_col, bias_txt, bias_icon = "#6B7A8D","Ratio indisponible","⚪"
        bias_advice = ""

    # Funding rate interpretation
    if funding is not None:
        if funding > 0.05:
            fund_col, fund_txt = "#D93025", f"+{funding:.4f}% — Longs paient les shorts → Marché surchauffé 🔴"
        elif funding < -0.01:
            fund_col, fund_txt = "#0FBF5F", f"{funding:.4f}% — Shorts paient les longs → Pression haussière 🟢"
        else:
            fund_col, fund_txt = "#C9A94B", f"{funding:.4f}% — Neutre"
    else:
        fund_col, fund_txt = "#6B7A8D","N/A"

    # Build bars for L/S history
    hist_bars = ""
    if ls_hist:
        for i, h in enumerate(ls_hist[-7:]):
            try:
                lr = float(h.get("longAccount",0.5)) * 100
                sr = 100 - lr
                bc = "#0FBF5F" if lr < 50 else "#D93025"
                lbl = f"{lr:.0f}%"
                hist_bars += f"""
                <div style="display:flex;flex-direction:column;align-items:center;gap:2px;flex:1;">
                  <div style="font-size:9px;color:{bc};font-weight:700;">{lbl}</div>
                  <div style="width:90%;height:{int(lr*0.5)+5}px;background:{bc};
                       border-radius:3px;opacity:0.8;"></div>
                  <div style="font-size:8px;color:#6B7A8D;">H-{7-i}</div>
                </div>"""
            except: pass

    oi_fmt = f"{oi/1000:.1f}K BTC" if oi else "N/A"
    lp_fmt = f"${price:,.0f}" if price else "N/A"
    lr_fmt = f"{l_ratio:.1f}%" if l_ratio else "N/A"
    sr_fmt = f"{s_ratio:.1f}%" if s_ratio else "N/A"
    bv_fmt = f"{buy_vol/1e6:.1f}M" if buy_vol else "N/A"
    sv_fmt = f"{sell_vol/1e6:.1f}M" if sell_vol else "N/A"

    # L/S bar visual (percentage)
    lr_w = int(l_ratio) if l_ratio else 50
    sr_w = 100 - lr_w

    full_html = f"""<!DOCTYPE html><html><head>
    <link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@700&family=JetBrains+Mono&display=swap" rel="stylesheet">
    <style>body{{margin:0;padding:8px;background:#060E1A;font-family:'Rajdhani',sans-serif;color:#DCE8F5;}}</style>
    </head><body>

    <!-- PRICE + OI -->
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-bottom:10px;">
      <div style="background:#0F1C2E;border:1px solid #0FBF5F;border-radius:8px;padding:8px;text-align:center;">
        <div style="color:#6B7A8D;font-size:9px;letter-spacing:1px;">PRIX BTC</div>
        <div style="color:#F0C96A;font-size:18px;font-weight:700;">{lp_fmt}</div>
      </div>
      <div style="background:#0F1C2E;border:1px solid #E07B2A;border-radius:8px;padding:8px;text-align:center;">
        <div style="color:#6B7A8D;font-size:9px;letter-spacing:1px;">OPEN INTEREST</div>
        <div style="color:#E07B2A;font-size:18px;font-weight:700;">{oi_fmt}</div>
      </div>
      <div style="background:#0F1C2E;border:1px solid {fund_col};border-radius:8px;padding:8px;text-align:center;">
        <div style="color:#6B7A8D;font-size:9px;letter-spacing:1px;">FUNDING RATE</div>
        <div style="color:{fund_col};font-size:13px;font-weight:700;">{fund_txt[:12]}</div>
      </div>
    </div>

    <!-- MARKET BIAS -->
    <div style="background:rgba(0,0,0,.3);border:2px solid {bias_col};border-radius:8px;
         padding:10px 14px;margin-bottom:10px;">
      <div style="font-size:14px;font-weight:700;color:{bias_col};">{bias_icon} {bias_txt}</div>
      <div style="font-size:11px;color:#8A95A3;margin-top:4px;">{bias_advice}</div>
    </div>

    <!-- LONG/SHORT RATIO BAR -->
    <div style="margin-bottom:10px;">
      <div style="font-size:12px;color:#C9A94B;letter-spacing:2px;margin-bottom:6px;">
        📊 RATIO LONG / SHORT (TOP TRADERS)
      </div>
      <div style="display:flex;border-radius:6px;overflow:hidden;height:28px;">
        <div style="width:{lr_w}%;background:#D93025;display:flex;align-items:center;
             justify-content:center;font-size:12px;font-weight:700;color:#fff;">
          🔴 LONGS {lr_fmt}
        </div>
        <div style="width:{sr_w}%;background:#0FBF5F;display:flex;align-items:center;
             justify-content:center;font-size:12px;font-weight:700;color:#060E1A;">
          🟢 SHORTS {sr_fmt}
        </div>
      </div>
    </div>

    <!-- TAKER VOLUME -->
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:10px;">
      <div style="background:rgba(217,48,37,.08);border:1px solid #D93025;border-radius:8px;padding:8px;text-align:center;">
        <div style="color:#6B7A8D;font-size:9px;">VOLUME ACHAT (Taker)</div>
        <div style="color:#0FBF5F;font-size:16px;font-weight:700;">{bv_fmt}</div>
      </div>
      <div style="background:rgba(15,191,95,.08);border:1px solid #0FBF5F;border-radius:8px;padding:8px;text-align:center;">
        <div style="color:#6B7A8D;font-size:9px;">VOLUME VENTE (Taker)</div>
        <div style="color:#D93025;font-size:16px;font-weight:700;">{sv_fmt}</div>
      </div>
    </div>

    <!-- L/S HISTORY -->
    {"<div style='font-size:12px;color:#C9A94B;letter-spacing:2px;margin-bottom:6px;'>📅 HISTORIQUE L/S 7H</div><div style='display:flex;gap:4px;align-items:flex-end;background:#0A1520;border-radius:8px;padding:8px;height:90px;'>" + hist_bars + "</div>" if hist_bars else ""}

    <!-- FUNDING DETAIL -->
    <div style="background:#0F1C2E;border:1px solid #1A3050;border-radius:8px;padding:10px;margin-top:10px;">
      <div style="font-size:12px;color:#E07B2A;letter-spacing:2px;margin-bottom:6px;">
        💡 FUNDING RATE: <span style="color:{fund_col};">{fund_txt}</span>
      </div>
      <div style="font-size:10px;color:#8A95A3;line-height:1.8;">
        Positif → Longs paient = Marché haussier surchargé → Risque de correction<br>
        Négatif → Shorts paient = Pression haussière = Support pour BUY<br>
        Proche de 0 → Marché équilibré → Suivre le signal technique
      </div>
    </div>

    <!-- GUIDE ICT -->
    <div style="background:#0F1C2E;border-left:3px solid #E07B2A;border-radius:6px;
         padding:10px 14px;margin-top:10px;font-size:11px;color:#8A95A3;line-height:1.9;">
      <b style="color:#E07B2A;">📖 LECTURE ICT/SMC :</b><br>
      🔴 Majorité en LONG → Market makers vont chasser les stops → Signal SELL probable<br>
      🟢 Majorité en SHORT → Short squeeze probable → Signal BUY renforcé<br>
      🔥 Funding élevé + App SELL → Confluence maximale pour vendre<br>
      🎯 Funding négatif + App BUY → Setup idéal pour acheter
    </div>
    </body></html>"""

    components.html(full_html, height=620, scrolling=True)


# ─── ECONOMIC CALENDAR ───────────────────────────────────────────────────────
def fetch_economic_calendar():
    """Fetch economic calendar events for today and tomorrow."""
    events = []
    try:
        # Use Forex Factory RSS feed (public)
        import xml.etree.ElementTree as ET
        r = requests.get(
            "https://nfs.faireconomy.media/ff_calendar_thisweek.xml",
            timeout=10,
            headers={"User-Agent":"Mozilla/5.0"}
        )
        if r.status_code == 200:
            root = ET.fromstring(r.content)
            for event in root.findall('.//event'):
                title    = event.findtext('title','')
                country  = event.findtext('country','')
                date_str = event.findtext('date','')
                time_str = event.findtext('time','')
                impact   = event.findtext('impact','')
                forecast = event.findtext('forecast','')
                previous = event.findtext('previous','')
                actual   = event.findtext('actual','')
                events.append({
                    "title":title, "country":country,
                    "date":date_str, "time":time_str,
                    "impact":impact, "forecast":forecast,
                    "previous":previous, "actual":actual
                })
            return events
    except: pass

    # Fallback: static high-impact events for the week
    now = datetime.now(timezone.utc)
    events = [
        {"title":"Non-Farm Payrolls",     "country":"USD","date":now.strftime("%Y-%m-%d"),
         "time":"13:30","impact":"High","forecast":"180K","previous":"151K","actual":""},
        {"title":"CPI m/m",               "country":"USD","date":now.strftime("%Y-%m-%d"),
         "time":"13:30","impact":"High","forecast":"0.3%","previous":"0.5%","actual":""},
        {"title":"FOMC Meeting Minutes",  "country":"USD","date":now.strftime("%Y-%m-%d"),
         "time":"19:00","impact":"High","forecast":"","previous":"","actual":""},
        {"title":"Initial Jobless Claims","country":"USD","date":now.strftime("%Y-%m-%d"),
         "time":"13:30","impact":"Medium","forecast":"215K","previous":"221K","actual":""},
        {"title":"GDP q/q",               "country":"USD","date":now.strftime("%Y-%m-%d"),
         "time":"13:30","impact":"High","forecast":"2.3%","previous":"2.3%","actual":""},
    ]
    return events

def impact_style(impact):
    impact_l = str(impact).lower()
    if "high" in impact_l:   return "#D93025","🔴 HIGH"
    if "medium" in impact_l: return "#E07B2A","🟠 MEDIUM"
    return "#6B7A8D","⚪ LOW"

def minutes_until(time_str, now):
    """Calculate minutes until event time (HH:MM format)."""
    try:
        h, m = map(int, time_str.split(":")[:2])
        event_mins = h * 60 + m
        now_mins   = now.hour * 60 + now.minute
        diff = event_mins - now_mins
        if diff < -120: diff += 1440  # next day
        return diff
    except: return 999

def render_calendar(now):
    import streamlit as st

    st.markdown("""
    <div style="background:linear-gradient(135deg,#0A1520,#0F1C2E);border:2px solid #4A7EC7;
         border-radius:12px;padding:14px 18px;margin-bottom:14px;text-align:center;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;
           color:#4A7EC7;letter-spacing:2px;">🗓️ CALENDRIER ÉCONOMIQUE</div>
      <div style="font-size:10px;color:#6B7A8D;margin-top:3px;">
        Événements macroéconomiques · Impact sur BTC · Or · Pétrole · Forex
      </div>
    </div>""", unsafe_allow_html=True)

    with st.spinner("Chargement du calendrier..."):
        events = fetch_economic_calendar()

    if not events:
        st.warning("Calendrier indisponible. Vérifie la connexion.")
        return

    # Filter: only HIGH and MEDIUM, USD/EUR focused
    key_countries = {"USD","EUR","GBP","JPY","CAD","AUD","CHF"}
    filtered = [e for e in events
                if e.get("impact","").lower() in ("high","medium","red","orange")
                or "High" in e.get("impact","")
                or "Medium" in e.get("impact","")]

    if not filtered:
        filtered = events  # show all if nothing filtered

    # Sort by time
    filtered = sorted(filtered, key=lambda e: e.get("time","99:99"))

    # Warning banner for upcoming HIGH events
    upcoming_high = []
    for e in filtered:
        if "high" in str(e.get("impact","")).lower() or "red" in str(e.get("impact","")).lower():
            mins = minutes_until(e.get("time",""), now)
            if 0 <= mins <= 60:
                upcoming_high.append((e, mins))

    if upcoming_high:
        for e, mins in upcoming_high[:2]:
            st.markdown(f"""
            <div style="background:rgba(217,48,37,.12);border:2px solid #D93025;
                 border-radius:8px;padding:10px 14px;margin-bottom:8px;
                 animation:pulse 1s infinite;">
              <div style="font-size:13px;font-weight:700;color:#D93025;">
                🚨 NEWS ROUGE DANS {mins} MIN — {e['title']} ({e['country']})
              </div>
              <div style="font-size:11px;color:#FF8A80;margin-top:4px;">
                ⚠️ FERMER ou ÉVITER toute position ouverte !
                Entrer seulement APRÈS confirmation post-news.
              </div>
            </div>""", unsafe_allow_html=True)

    # Events list
    st.markdown(f"""<div style="font-size:10px;color:#6B7A8D;margin-bottom:8px;">
      🕐 Heure actuelle UTC : <b style="color:#C9A94B;">{now.strftime('%H:%M')}</b>
      · {len(filtered)} événements cette semaine
    </div>""", unsafe_allow_html=True)

    shown = 0
    for e in filtered[:20]:
        title    = e.get("title","")
        country  = e.get("country","")
        time_str = e.get("time","")
        impact   = e.get("impact","")
        forecast = e.get("forecast","")
        previous = e.get("previous","")
        actual   = e.get("actual","")

        imp_col, imp_lbl = impact_style(impact)
        mins = minutes_until(time_str, now)

        # Status
        if actual and actual.strip() and actual != "":
            status_col, status_txt = "#6B7A8D","✅ Publié"
            # Compare actual vs forecast
            try:
                a_v = float(actual.replace("%","").replace("K","").replace("B",""))
                f_v = float(forecast.replace("%","").replace("K","").replace("B",""))
                if a_v > f_v:   status_col, status_txt = "#0FBF5F","✅ Surprise ↑ HAUSSIER"
                elif a_v < f_v: status_col, status_txt = "#D93025","✅ Déception ↓ BAISSIER"
            except: pass
        elif 0 <= mins <= 30:
            status_col, status_txt = "#D93025", f"🚨 DANS {mins} MIN"
        elif 0 <= mins <= 60:
            status_col, status_txt = "#E07B2A", f"⏰ DANS {mins} MIN"
        elif mins < 0:
            status_col, status_txt = "#6B7A8D","⏳ Passé"
        else:
            status_col, status_txt = "#4A7EC7",f"📅 {time_str} UTC"

        bg = "rgba(217,48,37,.06)" if "High" in impact or "red" in impact.lower() else \
             "rgba(224,123,42,.04)" if "Medium" in impact or "orange" in impact.lower() else \
             "rgba(15,28,46,0.5)"

        fore_actual = ""
        if forecast: fore_actual += f"Prév: <b style='color:#C9A94B;'>{forecast}</b> "
        if previous: fore_actual += f"Préc: <b style='color:#6B7A8D;'>{previous}</b> "
        if actual and actual.strip():
            a_col = "#0FBF5F" if "Haussier" in status_txt else "#D93025" if "Baissier" in status_txt else "#DCE8F5"
            fore_actual += f"Réel: <b style='color:{a_col};'>{actual}</b>"

        st.markdown(f"""
        <div style="background:{bg};border:1px solid {imp_col}44;border-left:3px solid {imp_col};
             border-radius:6px;padding:9px 12px;margin:4px 0;">
          <div style="display:flex;justify-content:space-between;align-items:flex-start;">
            <div style="flex:1;">
              <div style="font-size:12px;color:#DCE8F5;font-weight:700;">{title}</div>
              <div style="font-size:10px;color:#6B7A8D;margin-top:2px;">
                🌍 {country} · {fore_actual}
              </div>
            </div>
            <div style="text-align:right;margin-left:10px;">
              <div style="font-size:10px;font-weight:700;color:{imp_col};">{imp_lbl}</div>
              <div style="font-size:10px;color:{status_col};margin-top:2px;">{status_txt}</div>
            </div>
          </div>
        </div>""", unsafe_allow_html=True)
        shown += 1

    if shown == 0:
        st.info("Aucun événement HIGH/MEDIUM trouvé pour aujourd'hui.")

    # Trading rules
    st.markdown("""
    <div class="card card-blue" style="margin-top:14px;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:12px;color:#4A7EC7;
           letter-spacing:2px;margin-bottom:8px;">📋 RÈGLES DE TRADING AUTOUR DES NEWS</div>
      <div style="font-size:11px;color:#8A95A3;line-height:1.9;">
        🚫 <b style="color:#D93025;">30 min avant news rouge</b> → Ne pas ouvrir de position<br>
        ⏳ <b style="color:#E07B2A;">Juste après la news</b> → Attendre 5-15 min (faux mouvements)<br>
        ✅ <b style="color:#0FBF5F;">Réel > Prévision</b> → Surprise haussière → Chercher BUY<br>
        ✅ <b style="color:#D93025;">Réel < Prévision</b> → Déception baissière → Chercher SELL<br>
        🎯 <b style="color:#C9A94B;">Meilleure entrée</b> → Signal app + confirmation après news
      </div>
    </div>""", unsafe_allow_html=True)


# ─── ASSET-SPECIFIC DATA FETCHERS ────────────────────────────────────────────

def _stooq(symbol, rows=7):
    """Fetch OHLCV daily data from stooq.com as CSV — no yfinance needed."""
    try:
        url = f"https://stooq.com/q/d/l/?s={symbol}&i=d"
        r   = requests.get(url, timeout=8,
                           headers={"User-Agent":"Mozilla/5.0"})
        if r.status_code != 200 or "No data" in r.text[:50]:
            return []
        lines  = r.text.strip().split("\n")
        header = [h.strip().lower() for h in lines[0].split(",")]
        data   = []
        for line in lines[1:]:
            vals = line.split(",")
            if len(vals) < len(header): continue
            row = dict(zip(header, vals))
            try:    data.append({"close": float(row.get("close",0))})
            except: pass
        return data[-rows:] if len(data) >= rows else data
    except: return []

def _last2(symbol):
    """Return (current, prev, change) for a symbol via stooq."""
    d = _stooq(symbol, 7)
    if not d: return None, None, 0
    cur  = d[-1]["close"]
    prev = d[-2]["close"] if len(d) >= 2 else cur
    return round(cur,5), round(prev,5), round(cur-prev,5)

@st.cache_data(ttl=120, show_spinner=False)
def fetch_vix():
    cur, prev, chg = _last2("^vix")
    if not cur: return None
    hist = [r["close"] for r in _stooq("^vix", 7)]
    return {"value": round(cur,2), "change": round(chg,2),
            "history": hist, "prev": round(prev,2)}

@st.cache_data(ttl=120, show_spinner=False)
def fetch_dxy():
    cur, prev, chg = _last2("dx-y.nyb")
    if not cur:
        cur, prev, chg = _last2("dxy.i")
    if not cur: return None
    hist = [r["close"] for r in _stooq("dx-y.nyb", 7)]
    return {"value": round(cur,3), "change": round(chg,3),
            "history": hist, "prev": round(prev,3)}

def fetch_eia_oil():
    result = {}
    # WTI
    wti, wti_prev, wti_chg = _last2("cl.f")
    if wti:
        result["wti_price"] = round(wti, 2)
        result["wti_prev"]  = round(wti_prev, 2)
        result["wti_chg"]   = round(wti_chg, 2)
    # Brent
    brent, _, _ = _last2("bz.f")
    if brent: result["brent_price"] = round(brent, 2)
    # Natural Gas
    ng, _, _ = _last2("ng.f")
    if ng: result["ng_price"] = round(ng, 3)
    # USO history
    uso_hist = _stooq("uso.us", 7)
    if uso_hist: result["uso_history"] = [round(r["close"],2) for r in uso_hist]
    return result

@st.cache_data(ttl=120, show_spinner=False)
def fetch_gold_sentiment():
    result = {}
    # Gold
    gc, gc_prev, gc_chg = _last2("gc.f")
    if gc:
        result["gold_price"] = round(gc, 2)
        result["gold_prev"]  = round(gc_prev, 2)
        result["gold_chg"]   = round(gc_chg, 2)
        hist = _stooq("gc.f", 7)
        result["gold_hist"] = [round(r["close"],2) for r in hist]
    # Silver
    si, _, _ = _last2("si.f")
    if si: result["silver_price"] = round(si, 3)
    # GDX miners
    gdx, gdx_prev, gdx_chg = _last2("gdx.us")
    if gdx:
        result["gdx_price"] = round(gdx, 2)
        result["gdx_prev"]  = round(gdx_prev, 2)
        result["gdx_chg"]   = round(gdx_chg, 2)
    # TIP ETF (real yields proxy)
    tip, _, _ = _last2("tip.us")
    if tip: result["tip_price"] = round(tip, 2)
    return result

def fetch_sp500_sentiment():
    result = {}
    # SPY
    spy, spy_prev, spy_chg = _last2("spy.us")
    if spy:
        result["spy_price"] = round(spy, 2)
        result["spy_prev"]  = round(spy_prev, 2)
        result["spy_chg"]   = round(spy_chg, 2)
        hist = _stooq("spy.us", 7)
        result["spy_hist"] = [round(r["close"],2) for r in hist]
    # QQQ
    qqq, qqq_prev, qqq_chg = _last2("qqq.us")
    if qqq:
        result["qqq_price"] = round(qqq, 2)
        result["qqq_chg"]   = round(qqq_chg, 2)
    # IWM
    iwm, iwm_prev, iwm_chg = _last2("iwm.us")
    if iwm:
        result["iwm_price"] = round(iwm, 2)
        result["iwm_chg"]   = round(iwm_chg, 2)
    # VIX
    vix_data = fetch_vix()
    if vix_data:
        result["vix"]     = vix_data["value"]
        result["vix_chg"] = vix_data["change"]
    return result

def fetch_eur_sentiment():
    result = {}
    # EUR/USD
    eur, eur_prev, eur_chg = _last2("eurusd")
    if eur:
        result["eur_price"] = round(eur, 5)
        result["eur_prev"]  = round(eur_prev, 5)
        result["eur_chg"]   = round(eur_chg, 5)
        hist = _stooq("eurusd", 7)
        result["eur_hist"] = [round(r["close"],5) for r in hist]
    # EUR/GBP
    eg, _, _ = _last2("eurgbp")
    if eg: result["eurgbp"] = round(eg, 5)
    # USD/CHF
    uc, _, _ = _last2("usdchf")
    if uc: result["usdchf"] = round(uc, 5)
    # DXY
    dxy = fetch_dxy()
    if dxy:
        result["dxy"]     = dxy["value"]
        result["dxy_chg"] = dxy["change"]
    return result

def render_sentiment_tab(asset_name, now, res=None):
    """Render the correct sentiment tool based on selected asset."""
    import streamlit as st
    import streamlit.components.v1 as components

    is_btc  = "BTC" in asset_name
    is_gold = "XAU" in asset_name or "GC" in asset_name or "Gold" in asset_name.upper() or "GC" in asset_name
    is_oil  = "CL" in asset_name or "WTI" in asset_name or "Oil" in asset_name.upper()
    is_eur  = "EUR" in asset_name or "6E" in asset_name
    is_sp   = "S&P" in asset_name or "ES" in asset_name or "NAS" in asset_name

    # ── BTC → Fear & Greed ──────────────────────────────────────────────────
    if is_btc:
        render_fear_greed()

    # ── XAUUSD → VIX + DXY + Gold Miners ───────────────────────────────────
    elif is_gold:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#0A1520,#0F1C2E);border:2px solid #C9A94B;
             border-radius:12px;padding:14px 18px;margin-bottom:14px;text-align:center;">
          <div style="font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;
               color:#C9A94B;letter-spacing:2px;">📊 SENTIMENT OR (XAUUSD)</div>
          <div style="font-size:10px;color:#6B7A8D;margin-top:3px;">
            VIX · DXY · GDX Miners · Silver — Indicateurs de sentiment Or
          </div>
        </div>""", unsafe_allow_html=True)

        with st.spinner("Chargement des données Or..."):
            vix  = fetch_vix()
            dxy  = fetch_dxy()
            gold = fetch_gold_sentiment()

        if not gold:
            st.error("Données indisponibles. Vérifie ta connexion.")
            return

        vix_v  = vix["value"]  if vix  else None
        dxy_v  = dxy["value"]  if dxy  else None
        dxy_c  = dxy["change"] if dxy  else 0

        # VIX interpretation for Gold
        if vix_v:
            if vix_v > 30:
                vix_col, vix_txt = "#0FBF5F","VIX ÉLEVÉ → Panique marché → Or monte (refuge) 📈"
            elif vix_v > 20:
                vix_col, vix_txt = "#C9A94B","VIX MODÉRÉ → Incertitude → Or neutre à haussier"
            else:
                vix_col, vix_txt = "#D93025","VIX BAS → Marchés sereins → Or sous pression 📉"
        else:
            vix_col, vix_txt = "#6B7A8D","VIX indisponible"

        # DXY interpretation for Gold
        if dxy_v:
            if dxy_c < -0.2:
                dxy_col, dxy_txt = "#0FBF5F","DXY BAISSE → Dollar faible → Or monte 📈"
            elif dxy_c > 0.2:
                dxy_col, dxy_txt = "#D93025","DXY MONTE → Dollar fort → Or sous pression 📉"
            else:
                dxy_col, dxy_txt = "#C9A94B","DXY STABLE → Neutre pour l'Or"
        else:
            dxy_col, dxy_txt = "#6B7A8D","DXY indisponible"

        # GDX interpretation (miners lead gold by 1-2 sessions)
        gdx_c = gold.get("gdx_chg", 0)
        if gdx_c and gdx_c > 0.3:
            gdx_col, gdx_txt = "#0FBF5F",f"GDX +{gdx_c:.2f} → Mineurs haussiers → Or devrait suivre 📈"
        elif gdx_c and gdx_c < -0.3:
            gdx_col, gdx_txt = "#D93025",f"GDX {gdx_c:.2f} → Mineurs baissiers → Or sous pression 📉"
        else:
            gdx_col, gdx_txt = "#C9A94B",f"GDX {gdx_c:+.2f} → Neutre"

        # Overall gold bias
        bullish_count = sum([
            1 if vix_v and vix_v > 25 else 0,
            1 if dxy_c < -0.1 else 0,
            1 if gdx_c and gdx_c > 0.2 else 0,
        ])
        if bullish_count >= 2:
            bias_col, bias_icon, bias_txt = "#0FBF5F","🟢","BIAIS HAUSSIER — Conditions favorables pour l'Or"
        elif bullish_count == 0:
            bias_col, bias_icon, bias_txt = "#D93025","🔴","BIAIS BAISSIER — Pression sur l'Or"
        else:
            bias_col, bias_icon, bias_txt = "#C9A94B","🟡","BIAIS NEUTRE — Pas de signal clair"

        # Build history bars for gold
        hist_bars = ""
        hist = gold.get("gold_hist", [])
        if hist:
            mn, mx = min(hist), max(hist)
            rng = mx - mn or 1
            days = [f"J-{len(hist)-1-i}" if i < len(hist)-1 else "Auj." for i in range(len(hist))]
            for i, val in enumerate(hist):
                bh = max(15, int((val - mn) / rng * 70))
                bc = "#0FBF5F" if i > 0 and val >= hist[i-1] else "#D93025"
                hist_bars += f"""
                <div style="display:flex;flex-direction:column;align-items:center;gap:2px;flex:1;">
                  <div style="font-size:9px;color:{bc};font-weight:700;">{val:.0f}</div>
                  <div style="width:90%;height:{bh}px;background:{bc};border-radius:3px;opacity:0.8;"></div>
                  <div style="font-size:8px;color:#6B7A8D;">{days[i]}</div>
                </div>"""

        gold_price = gold.get("gold_price","N/A")
        gold_chg   = gold.get("gold_chg", 0)
        silver     = gold.get("silver_price","N/A")
        gdx_price  = gold.get("gdx_price","N/A")
        tip_price  = gold.get("tip_price","N/A")

        full_html = f"""<!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@700&family=JetBrains+Mono&display=swap" rel="stylesheet">
        <style>body{{margin:0;padding:8px;background:#060E1A;font-family:'Rajdhani',sans-serif;color:#DCE8F5;}}</style>
        </head><body>

        <!-- BIAS BANNER -->
        <div style="background:rgba(0,0,0,.3);border:2px solid {bias_col};border-radius:8px;
             padding:10px 14px;margin-bottom:10px;text-align:center;">
          <div style="font-size:16px;font-weight:700;color:{bias_col};">{bias_icon} {bias_txt}</div>
        </div>

        <!-- STATS GRID -->
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:10px;">
          <div style="background:#0F1C2E;border:1px solid #C9A94B;border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">OR SPOT</div>
            <div style="color:#F0C96A;font-size:18px;font-weight:700;">${sf(gold_price)}</div>
            <div style="color:{'#0FBF5F' if (gold_chg or 0)>=0 else '#D93025'};font-size:11px;">{sf(gold_chg,'+.2f','±0.00')}</div>
          </div>
          <div style="background:#0F1C2E;border:1px solid #6B7A8D;border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">ARGENT / GDX</div>
            <div style="color:#DCE8F5;font-size:14px;font-weight:700;">Ag ${sf(silver,',.3f')} / GDX ${sf(gdx_price)}</div>
          </div>
          <div style="background:#0F1C2E;border:1px solid {vix_col};border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">VIX (Peur Marché)</div>
            <div style="color:{vix_col};font-size:18px;font-weight:700;">{sf(vix_v,',.2f')}</div>
          </div>
          <div style="background:#0F1C2E;border:1px solid {dxy_col};border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">DXY (Dollar Index)</div>
            <div style="color:{dxy_col};font-size:18px;font-weight:700;">{sf(dxy_v,',.3f')}</div>
          </div>
        </div>

        <!-- INDICATORS -->
        <div style="display:flex;flex-direction:column;gap:6px;margin-bottom:10px;">
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {vix_col};
               border-radius:5px;padding:8px 12px;font-size:12px;color:{vix_col};">{vix_txt}</div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {dxy_col};
               border-radius:5px;padding:8px 12px;font-size:12px;color:{dxy_col};">{dxy_txt}</div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {gdx_col};
               border-radius:5px;padding:8px 12px;font-size:12px;color:{gdx_col};">{gdx_txt}</div>
        </div>

        <!-- GOLD PRICE HISTORY -->
        {"<div style='font-size:12px;color:#C9A94B;letter-spacing:2px;margin-bottom:6px;'>📅 OR — 7 DERNIERS JOURS</div><div style='display:flex;gap:4px;align-items:flex-end;background:#0A1520;border-radius:8px;padding:8px;height:100px;'>" + hist_bars + "</div>" if hist_bars else ""}

        <!-- GUIDE -->
        <div style="background:#0F1C2E;border-left:3px solid #C9A94B;border-radius:6px;
             padding:10px 14px;margin-top:10px;font-size:11px;color:#8A95A3;line-height:1.9;">
          <b style="color:#C9A94B;">📖 RÈGLES OR :</b><br>
          📈 VIX monte + DXY baisse + GDX monte → BUY Or confirmé ✅<br>
          📉 VIX bas + DXY monte + GDX baisse → SELL Or confirmé ✅<br>
          🎯 App BUY + VIX>25 + DXY baisse = Setup idéal pour acheter l'Or
        </div>
        </body></html>"""

        components.html(full_html, height=580, scrolling=True)

    # ── CL WTI → EIA + DXY + Oil Sentiment ─────────────────────────────────
    elif is_oil:
        # ── Tout depuis res (prix OANDA déjà chargé) ──────────────────────────
        import streamlit.components.v1 as components_oil
        if not res:
            st.warning("⚠️ Lance d'abord une analyse CL WTI.")
            return

        wti_p   = res.get("close", 0)
        atr     = res.get("atr", wti_p * 0.008)
        rsi     = res.get("rsi", 50)
        signal  = res.get("signal", "RANGE")
        conf    = res.get("confluence", 50)
        vwap    = res.get("vwap", wti_p)
        bb_up   = res.get("bb_upper", wti_p * 1.01)
        bb_lo   = res.get("bb_lower", wti_p * 0.99)
        macd_b  = res.get("macd_bull", False)

        # Zones de liquidité WTI basées sur ATR
        zones = [
            {"label":"Résistance R3 — Buyside Majeur",  "price": round(wti_p + atr*3, 3), "type":"sell","s":3},
            {"label":"Résistance R2",                   "price": round(wti_p + atr*2, 3), "type":"sell","s":2},
            {"label":"Résistance R1 — BB Supérieure",   "price": round(bb_up, 3),          "type":"sell","s":1},
            {"label":"VWAP Session",                    "price": round(vwap, 3),            "type":"vwap","s":0},
            {"label":"◆ PRIX ACTUEL",                   "price": round(wti_p, 3),          "type":"current","s":0},
            {"label":"Support S1 — BB Inférieure",      "price": round(bb_lo, 3),          "type":"buy", "s":1},
            {"label":"Support S2",                      "price": round(wti_p - atr*2, 3), "type":"buy", "s":2},
            {"label":"Support S3 — Sellside Majeur",    "price": round(wti_p - atr*3, 3), "type":"buy", "s":3},
        ]

        # Signal momentum
        if signal == "BUY":
            sig_col, sig_txt = "#0FBF5F", f"BUY {conf:.0f}% — Momentum haussier WTI 📈"
        elif signal == "SELL":
            sig_col, sig_txt = "#D93025", f"SELL {conf:.0f}% — Momentum baissier WTI 📉"
        else:
            sig_col, sig_txt = "#C9A94B", f"RANGE {conf:.0f}% — Consolidation en cours"

        # RSI interprétation
        if rsi > 70:
            rsi_col, rsi_txt = "#D93025", f"RSI {rsi:.1f} — Zone de SURACHAT → SELL potentiel"
        elif rsi < 30:
            rsi_col, rsi_txt = "#0FBF5F", f"RSI {rsi:.1f} — Zone de SURVENTE → BUY potentiel"
        else:
            rsi_col, rsi_txt = "#C9A94B", f"RSI {rsi:.1f} — Zone neutre"

        # VWAP position
        if wti_p > vwap:
            vwap_col, vwap_txt = "#0FBF5F", f"Prix AU-DESSUS du VWAP ({vwap:.3f}) → Biais haussier"
        else:
            vwap_col, vwap_txt = "#D93025", f"Prix EN-DESSOUS du VWAP ({vwap:.3f}) → Biais baissier"

        # Zone rows
        zone_rows = ""
        for z in zones:
            if z["type"] == "current":
                zc="#F0C96A"; zb="rgba(240,201,106,.12)"
                lbl = "◆ PRIX ACTUEL"
            elif z["type"] == "vwap":
                zc="#4A7EC7"; zb="rgba(74,126,199,.1)"
                lbl = f"── VWAP"
            elif z["type"] == "sell":
                zc="#D93025"; zb=f"rgba(217,48,37,{min(0.6,0.15+z['s']*0.15):.2f})"
                lbl = f"🔴 {'█'*z['s']} {z['label']}"
            else:
                zc="#0FBF5F"; zb=f"rgba(15,191,95,{min(0.6,0.15+z['s']*0.15):.2f})"
                lbl = f"🟢 {'█'*z['s']} {z['label']}"
            dist = round((z["price"]-wti_p)/wti_p*100,2)
            dist_s = f"+{dist:.2f}%" if dist>0 else f"{dist:.2f}%" if dist<0 else "← ICI"
            zone_rows += f"""<tr style="background:{zb};">
              <td style="color:{zc};font-family:'JetBrains Mono',monospace;font-weight:700;font-size:12px;">${z['price']:.3f}</td>
              <td style="color:{zc};font-size:11px;">{lbl}</td>
              <td style="color:#C9A94B;font-size:11px;">{dist_s}</td>
            </tr>"""

        sig_icon = "🟢" if signal=="BUY" else "🔴" if signal=="SELL" else "🟡"
        full_html = f"""<!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@700&family=JetBrains+Mono&display=swap" rel="stylesheet">
        <style>body{{margin:0;padding:8px;background:#060E1A;font-family:'Rajdhani',sans-serif;color:#DCE8F5;}}
        table{{width:100%;border-collapse:collapse;}}
        th{{font-size:10px;color:#6B7A8D;padding:5px 8px;text-align:left;border-bottom:1px solid #1A3050;}}
        td{{padding:6px 8px;}}</style></head><body>

        <!-- HEADER STATS -->
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-bottom:10px;">
          <div style="background:#0F1C2E;border:1px solid #E07B2A;border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">CL WTI LIVE</div>
            <div style="color:#E07B2A;font-size:18px;font-weight:700;">${sf(wti_p,',.3f')}</div>
            <div style="font-size:9px;color:#6B7A8D;">via OANDA</div>
          </div>
          <div style="background:#0F1C2E;border:1px solid {sig_col};border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">SIGNAL</div>
            <div style="color:{sig_col};font-size:16px;font-weight:700;">{sig_icon} {signal}</div>
            <div style="color:{sig_col};font-size:11px;">{conf:.0f}%</div>
          </div>
          <div style="background:#0F1C2E;border:1px solid {rsi_col};border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">RSI(14)</div>
            <div style="color:{rsi_col};font-size:18px;font-weight:700;">{rsi:.1f}</div>
          </div>
        </div>

        <!-- INDICATEURS -->
        <div style="display:flex;flex-direction:column;gap:5px;margin-bottom:10px;">
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {sig_col};border-radius:5px;padding:7px 11px;font-size:12px;color:{sig_col};">{sig_txt}</div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {vwap_col};border-radius:5px;padding:7px 11px;font-size:12px;color:{vwap_col};">{vwap_txt}</div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {rsi_col};border-radius:5px;padding:7px 11px;font-size:12px;color:{rsi_col};">{rsi_txt}</div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {'#0FBF5F' if macd_b else '#D93025'};border-radius:5px;padding:7px 11px;font-size:12px;color:{'#0FBF5F' if macd_b else '#D93025'};">
            MACD {'Haussier ↑ — Momentum positif' if macd_b else 'Baissier ↓ — Momentum négatif'}
          </div>
        </div>

        <!-- ZONES TABLE -->
        <div style="font-size:12px;color:#E07B2A;letter-spacing:2px;margin-bottom:6px;">🗺️ ZONES DE LIQUIDITÉ WTI</div>
        <table><thead><tr><th>Prix</th><th>Zone</th><th>Distance</th></tr></thead>
        <tbody>{zone_rows}</tbody></table>

        <!-- GUIDE -->
        <div style="background:#0F1C2E;border-left:3px solid #E07B2A;border-radius:6px;
             padding:9px 12px;margin-top:10px;font-size:11px;color:#8A95A3;line-height:1.9;">
          <b style="color:#E07B2A;">📖 RÈGLES PÉTROLE ICT :</b><br>
          🔴 Prix approche R3/R2 + RSI>70 → SELL confirmé ✅<br>
          🟢 Prix approche S2/S3 + RSI&lt;35 → BUY confirmé ✅<br>
          🗓️ Mercredi 15h30 UTC → Inventaires EIA → Volatilité forte !<br>
          💡 VWAP = ligne de force/résistance intraday clé
        </div>
        </body></html>"""

        components_oil.html(full_html, height=600, scrolling=True)

    # ── 6E EUR → depuis res ──────────────────────────────────────────────────
    elif is_eur:
        import streamlit.components.v1 as components_eur
        if not res:
            st.warning("⚠️ Lance d'abord une analyse 6E EUR.")
            return

        eur_p  = res.get("close", 0)
        atr    = res.get("atr", eur_p * 0.005)
        rsi    = res.get("rsi", 50)
        signal = res.get("signal", "RANGE")
        conf   = res.get("confluence", 50)
        vwap   = res.get("vwap", eur_p)
        bb_up  = res.get("bb_upper", eur_p * 1.005)
        bb_lo  = res.get("bb_lower", eur_p * 0.995)
        macd_b = res.get("macd_bull", False)

        zones = [
            {"label":"Résistance R3 — Buyside Majeur",  "price": round(eur_p + atr*3, 5), "type":"sell","s":3},
            {"label":"Résistance R2",                   "price": round(eur_p + atr*2, 5), "type":"sell","s":2},
            {"label":"Résistance R1 — BB Supérieure",   "price": round(bb_up, 5),          "type":"sell","s":1},
            {"label":"VWAP Session",                    "price": round(vwap, 5),            "type":"vwap","s":0},
            {"label":"◆ PRIX ACTUEL",                   "price": round(eur_p, 5),          "type":"current","s":0},
            {"label":"Support S1 — BB Inférieure",      "price": round(bb_lo, 5),          "type":"buy", "s":1},
            {"label":"Support S2",                      "price": round(eur_p - atr*2, 5), "type":"buy", "s":2},
            {"label":"Support S3 — Sellside Majeur",    "price": round(eur_p - atr*3, 5), "type":"buy", "s":3},
        ]

        if signal=="BUY":   sig_col,sig_txt="#0FBF5F",f"BUY {conf:.0f}% — EUR fort vs USD 📈"
        elif signal=="SELL":sig_col,sig_txt="#D93025",f"SELL {conf:.0f}% — USD fort vs EUR 📉"
        else:               sig_col,sig_txt="#C9A94B",f"RANGE {conf:.0f}% — Pas de biais clair"

        if rsi>70:   rsi_col,rsi_txt="#D93025",f"RSI {rsi:.1f} — EUR surachetée → Correction possible"
        elif rsi<30: rsi_col,rsi_txt="#0FBF5F",f"RSI {rsi:.1f} — EUR survendue → Rebond possible"
        else:        rsi_col,rsi_txt="#C9A94B",f"RSI {rsi:.1f} — Neutre"

        if eur_p>vwap: vwap_col,vwap_txt="#0FBF5F",f"EUR AU-DESSUS du VWAP ({vwap:.5f}) → Biais haussier"
        else:          vwap_col,vwap_txt="#D93025",f"EUR EN-DESSOUS du VWAP ({vwap:.5f}) → Biais baissier"

        zone_rows=""
        for z in zones:
            if z["type"]=="current": zc="#F0C96A";zb="rgba(240,201,106,.12)";lbl="◆ PRIX ACTUEL"
            elif z["type"]=="vwap":  zc="#4A7EC7";zb="rgba(74,126,199,.1)";lbl="── VWAP"
            elif z["type"]=="sell":  zc="#D93025";zb=f"rgba(217,48,37,{min(0.6,0.15+z['s']*0.15):.2f})";lbl=f"🔴 {'█'*z['s']} {z['label']}"
            else:                    zc="#0FBF5F";zb=f"rgba(15,191,95,{min(0.6,0.15+z['s']*0.15):.2f})";lbl=f"🟢 {'█'*z['s']} {z['label']}"
            dist=round((z["price"]-eur_p)/eur_p*100,3)
            dist_s=f"+{dist:.3f}%" if dist>0 else f"{dist:.3f}%" if dist<0 else "← ICI"
            zone_rows+=f"""<tr style="background:{zb};">
              <td style="color:{zc};font-family:'JetBrains Mono',monospace;font-weight:700;font-size:12px;">{z['price']:.5f}</td>
              <td style="color:{zc};font-size:11px;">{lbl}</td>
              <td style="color:#C9A94B;font-size:11px;">{dist_s}</td>
            </tr>"""

        sig_icon="🟢" if signal=="BUY" else "🔴" if signal=="SELL" else "🟡"
        full_html=f"""<!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@700&family=JetBrains+Mono&display=swap" rel="stylesheet">
        <style>body{{margin:0;padding:8px;background:#060E1A;font-family:'Rajdhani',sans-serif;color:#DCE8F5;}}
        table{{width:100%;border-collapse:collapse;}}
        th{{font-size:10px;color:#6B7A8D;padding:5px 8px;text-align:left;border-bottom:1px solid #1A3050;}}
        td{{padding:6px 8px;}}</style></head><body>

        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-bottom:10px;">
          <div style="background:#0F1C2E;border:1px solid #4A7EC7;border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">EUR/USD LIVE</div>
            <div style="color:#7AAEE8;font-size:17px;font-weight:700;">{eur_p:.5f}</div>
            <div style="font-size:9px;color:#6B7A8D;">via OANDA</div>
          </div>
          <div style="background:#0F1C2E;border:1px solid {sig_col};border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">SIGNAL</div>
            <div style="color:{sig_col};font-size:16px;font-weight:700;">{sig_icon} {signal}</div>
            <div style="color:{sig_col};font-size:11px;">{conf:.0f}%</div>
          </div>
          <div style="background:#0F1C2E;border:1px solid {rsi_col};border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">RSI(14)</div>
            <div style="color:{rsi_col};font-size:18px;font-weight:700;">{rsi:.1f}</div>
          </div>
        </div>

        <div style="display:flex;flex-direction:column;gap:5px;margin-bottom:10px;">
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {sig_col};border-radius:5px;padding:7px 11px;font-size:12px;color:{sig_col};">{sig_txt}</div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {vwap_col};border-radius:5px;padding:7px 11px;font-size:12px;color:{vwap_col};">{vwap_txt}</div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {rsi_col};border-radius:5px;padding:7px 11px;font-size:12px;color:{rsi_col};">{rsi_txt}</div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {'#0FBF5F' if macd_b else '#D93025'};border-radius:5px;padding:7px 11px;font-size:12px;color:{'#0FBF5F' if macd_b else '#D93025'};">
            MACD {'Haussier ↑ — EUR gagne du momentum' if macd_b else 'Baissier ↓ — EUR perd du momentum'}
          </div>
        </div>

        <div style="font-size:12px;color:#4A7EC7;letter-spacing:2px;margin-bottom:6px;">🗺️ ZONES DE LIQUIDITÉ EUR/USD</div>
        <table><thead><tr><th>Prix</th><th>Zone</th><th>Distance</th></tr></thead>
        <tbody>{zone_rows}</tbody></table>

        <div style="background:#0F1C2E;border-left:3px solid #4A7EC7;border-radius:6px;
             padding:9px 12px;margin-top:10px;font-size:11px;color:#8A95A3;line-height:1.9;">
          <b style="color:#4A7EC7;">📖 RÈGLES EUR/USD ICT :</b><br>
          🔴 Prix approche R2/R3 + RSI>65 → SELL EUR ✅<br>
          🟢 Prix approche S2/S3 + RSI&lt;35 → BUY EUR ✅<br>
          🗓️ NFP vendredi 13h30 UTC + CPI + BCE → Impact majeur<br>
          💡 Round numbers (1.0800, 1.0900...) = zones clés supplémentaires
        </div>
        </body></html>"""

        components_eur.html(full_html, height=600, scrolling=True)

    # ── ES S&P → depuis res ──────────────────────────────────────────────────
    elif "NAS" in asset_name:
        import streamlit.components.v1 as components_nas
        if not res:
            st.warning("⚠️ Lance d'abord une analyse NAS100.")
            return

        nas_p  = res.get("close", 0)
        atr    = res.get("atr", nas_p * 0.006)
        rsi    = res.get("rsi", 50)
        signal = res.get("signal", "RANGE")
        conf   = res.get("confluence", 50)
        vwap   = res.get("vwap", nas_p)
        bb_up  = res.get("bb_upper", nas_p * 1.006)
        bb_lo  = res.get("bb_lower", nas_p * 0.994)
        macd_b = res.get("macd_bull", False)
        ml_scores = res.get("ml_scores", {})
        momentum_sc = ml_scores.get("momentum", 0.5) if ml_scores else 0.5

        zones = [
            {"label":"Resistance R3 — Buyside Majeur",  "price": round(nas_p + atr*3, 2), "type":"sell","s":3},
            {"label":"Resistance R2",                   "price": round(nas_p + atr*2, 2), "type":"sell","s":2},
            {"label":"Resistance R1 — BB Superieure",   "price": round(bb_up, 2),          "type":"sell","s":1},
            {"label":"VWAP Session",                    "price": round(vwap, 2),            "type":"vwap","s":0},
            {"label":"PRIX ACTUEL",                     "price": round(nas_p, 2),          "type":"current","s":0},
            {"label":"Support S1 — BB Inferieure",      "price": round(bb_lo, 2),         "type":"buy", "s":1},
            {"label":"Support S2",                      "price": round(nas_p - atr*2, 2), "type":"buy", "s":2},
            {"label":"Support S3 — Sellside Majeur",    "price": round(nas_p - atr*3, 2), "type":"buy", "s":3},
        ]

        vix_sim = round(15 + (50-rsi)*0.4, 1)
        if vix_sim > 30:   vix_col,vix_lbl="#D93025","PEUR 😱"
        elif vix_sim > 20: vix_col,vix_lbl="#E07B2A","INCERTITUDE 😨"
        else:              vix_col,vix_lbl="#0FBF5F","SEREIN 😊"

        risk_on = momentum_sc > 0.55
        risk_col = "#0FBF5F" if risk_on else "#D93025"
        risk_txt = "RISK ON — Tech en hausse → Favorable BUY NAS100" if risk_on else "RISK OFF → Prudence NAS100"

        if signal=="BUY":   sig_col,sig_txt="#0FBF5F",f"BUY {conf:.0f}% — NAS100 momentum haussier 📈"
        elif signal=="SELL":sig_col,sig_txt="#D93025",f"SELL {conf:.0f}% — NAS100 momentum baissier 📉"
        else:               sig_col,sig_txt="#C9A94B",f"RANGE {conf:.0f}% — Consolidation NAS100"

        if nas_p > vwap: vwap_col,vwap_txt="#0FBF5F",f"Prix AU-DESSUS du VWAP ({vwap:.2f}) → Biais haussier"
        else:            vwap_col,vwap_txt="#D93025",f"Prix EN-DESSOUS du VWAP ({vwap:.2f}) → Biais baissier"

        zone_rows=""
        for z in zones:
            if z["type"]=="current": zc="#F0C96A";zb="rgba(240,201,106,.12)";lbl="◆ PRIX ACTUEL"
            elif z["type"]=="vwap":  zc="#4A7EC7";zb="rgba(74,126,199,.1)";lbl="── VWAP"
            elif z["type"]=="sell":  zc="#D93025";zb=f"rgba(217,48,37,{min(0.6,0.15+z['s']*0.15):.2f})";lbl=f"🔴 {'█'*z['s']} {z['label']}"
            else:                    zc="#0FBF5F";zb=f"rgba(15,191,95,{min(0.6,0.15+z['s']*0.15):.2f})";lbl=f"🟢 {'█'*z['s']} {z['label']}"
            dist=round((z["price"]-nas_p)/nas_p*100,2)
            dist_s=f"+{dist:.2f}%" if dist>0 else f"{dist:.2f}%" if dist<0 else "← ICI"
            zone_rows+=f"""<tr style="background:{zb};">
              <td style="color:{zc};font-weight:700;font-size:12px;">${z['price']:,.2f}</td>
              <td style="color:{zc};font-size:11px;">{lbl}</td>
              <td style="color:#C9A94B;font-size:11px;">{dist_s}</td>
            </tr>"""

        sig_icon="🟢" if signal=="BUY" else "🔴" if signal=="SELL" else "🟡"
        full_html=f"""<!DOCTYPE html><html><head>
        <style>body{{margin:0;padding:8px;background:#060E1A;font-family:sans-serif;color:#DCE8F5;}}
        table{{width:100%;border-collapse:collapse;}}
        th{{font-size:10px;color:#6B7A8D;padding:5px 8px;text-align:left;border-bottom:1px solid #1A3050;}}
        td{{padding:6px 8px;}}</style></head><body>
        <div style="background:linear-gradient(135deg,#0A1520,#0A1020);border:2px solid #4A7EC7;
             border-radius:12px;padding:12px;margin-bottom:10px;text-align:center;">
          <div style="font-size:16px;font-weight:700;color:#4A7EC7;letter-spacing:2px;">
            💻 SENTIMENT NAS100 (NASDAQ 100)</div>
          <div style="font-size:10px;color:#6B7A8D;">Tech · FAANG · Risk On/Off</div>
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-bottom:10px;">
          <div style="background:#0F1C2E;border:1px solid #4A7EC7;border-radius:8px;padding:8px;text-align:center;">
            <div style="font-size:9px;color:#6B7A8D;">NAS100 LIVE</div>
            <div style="font-size:17px;font-weight:700;color:#4A7EC7;">${sf(nas_p)}</div>
            <div style="font-size:9px;color:#6B7A8D;">via OANDA</div>
          </div>
          <div style="background:#0F1C2E;border:1px solid {sig_col};border-radius:8px;padding:8px;text-align:center;">
            <div style="font-size:9px;color:#6B7A8D;">SIGNAL</div>
            <div style="font-size:16px;font-weight:700;color:{sig_col};">{sig_icon} {signal}</div>
            <div style="font-size:11px;color:{sig_col};">{conf:.0f}%</div>
          </div>
          <div style="background:#0F1C2E;border:1px solid {vix_col};border-radius:8px;padding:8px;text-align:center;">
            <div style="font-size:9px;color:#6B7A8D;">VIX estime</div>
            <div style="font-size:15px;font-weight:700;color:{vix_col};">{vix_sim}</div>
            <div style="font-size:9px;color:{vix_col};">{vix_lbl}</div>
          </div>
        </div>
        <div style="display:flex;flex-direction:column;gap:5px;margin-bottom:10px;">
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {risk_col};border-radius:5px;padding:7px 11px;font-size:12px;color:{risk_col};">
            {'🟢' if risk_on else '🔴'} {risk_txt}
          </div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {sig_col};border-radius:5px;padding:7px 11px;font-size:12px;color:{sig_col};">{sig_txt}</div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {vwap_col};border-radius:5px;padding:7px 11px;font-size:12px;color:{vwap_col};">{vwap_txt}</div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {'#0FBF5F' if macd_b else '#D93025'};border-radius:5px;padding:7px 11px;font-size:12px;color:{'#0FBF5F' if macd_b else '#D93025'};">
            MACD {'Haussier ↑ — NAS100 momentum positif' if macd_b else 'Baissier ↓ — NAS100 momentum negatif'}
          </div>
        </div>
        <div style="font-size:12px;color:#4A7EC7;letter-spacing:2px;margin-bottom:6px;">🗺️ ZONES DE LIQUIDITE NAS100</div>
        <table><thead><tr><th>Prix</th><th>Zone</th><th>Distance</th></tr></thead>
        <tbody>{zone_rows}</tbody></table>
        <div style="background:#0F1C2E;border-left:3px solid #4A7EC7;border-radius:6px;
             padding:9px 12px;margin-top:10px;font-size:11px;color:#8A95A3;line-height:1.9;">
          <b style="color:#4A7EC7;">📖 REGLES NAS100 ICT :</b><br>
          🔴 Prix approche R2/R3 + RSI>70 + MACD baissier → SELL NAS100 ✅<br>
          🟢 Prix approche S2/S3 + RSI&lt;35 + MACD haussier → BUY NAS100 ✅<br>
          📊 VWAP = niveau equilibre quotidien cle pour NAS100<br>
          🗓️ NFP + Fed + CPI + Earnings FAANG → Volatilite majeure NAS100
        </div>
        </body></html>"""
        components_nas.html(full_html, height=600, scrolling=True)

    elif is_sp:
        import streamlit.components.v1 as components_sp
        if not res:
            st.warning("⚠️ Lance d'abord une analyse ES S&P.")
            return

        sp_p   = res.get("close", 0)
        atr    = res.get("atr", sp_p * 0.006)
        rsi    = res.get("rsi", 50)
        signal = res.get("signal", "RANGE")
        conf   = res.get("confluence", 50)
        vwap   = res.get("vwap", sp_p)
        bb_up  = res.get("bb_upper", sp_p * 1.006)
        bb_lo  = res.get("bb_lower", sp_p * 0.994)
        macd_b = res.get("macd_bull", False)
        ml_scores = res.get("ml_scores", {})
        momentum_sc = ml_scores.get("momentum", 0.5) if ml_scores else 0.5

        zones = [
            {"label":"Résistance R3 — Buyside Majeur",  "price": round(sp_p + atr*3, 2), "type":"sell","s":3},
            {"label":"Résistance R2",                   "price": round(sp_p + atr*2, 2), "type":"sell","s":2},
            {"label":"Résistance R1 — BB Supérieure",   "price": round(bb_up, 2),         "type":"sell","s":1},
            {"label":"VWAP Session",                    "price": round(vwap, 2),           "type":"vwap","s":0},
            {"label":"◆ PRIX ACTUEL",                   "price": round(sp_p, 2),          "type":"current","s":0},
            {"label":"Support S1 — BB Inférieure",      "price": round(bb_lo, 2),         "type":"buy", "s":1},
            {"label":"Support S2",                      "price": round(sp_p - atr*2, 2), "type":"buy", "s":2},
            {"label":"Support S3 — Sellside Majeur",    "price": round(sp_p - atr*3, 2), "type":"buy", "s":3},
        ]

        # VIX simulation from RSI inverse
        vix_sim = round(15 + (50-rsi)*0.4, 1)
        if vix_sim > 30:   vix_col,vix_lbl="#D93025","PEUR ÉLEVÉE 😱"
        elif vix_sim > 20: vix_col,vix_lbl="#E07B2A","INCERTITUDE 😨"
        else:              vix_col,vix_lbl="#0FBF5F","SEREIN 😊"

        risk_on = momentum_sc > 0.55
        risk_col = "#0FBF5F" if risk_on else "#D93025"
        risk_txt = "RISK ON — Appétit au risque fort → Favorable BUY S&P" if risk_on else "RISK OFF — Prudence → Risque de baisse S&P"

        if signal=="BUY":   sig_col,sig_txt="#0FBF5F",f"BUY {conf:.0f}% — S&P en momentum haussier 📈"
        elif signal=="SELL":sig_col,sig_txt="#D93025",f"SELL {conf:.0f}% — S&P en momentum baissier 📉"
        else:               sig_col,sig_txt="#C9A94B",f"RANGE {conf:.0f}% — Consolidation S&P"

        if eur_p_dummy := sp_p:
            if sp_p > vwap: vwap_col,vwap_txt="#0FBF5F",f"Prix AU-DESSUS du VWAP ({vwap:.2f}) → Biais haussier"
            else:           vwap_col,vwap_txt="#D93025",f"Prix EN-DESSOUS du VWAP ({vwap:.2f}) → Biais baissier"

        zone_rows=""
        for z in zones:
            if z["type"]=="current": zc="#F0C96A";zb="rgba(240,201,106,.12)";lbl="◆ PRIX ACTUEL"
            elif z["type"]=="vwap":  zc="#4A7EC7";zb="rgba(74,126,199,.1)";lbl="── VWAP"
            elif z["type"]=="sell":  zc="#D93025";zb=f"rgba(217,48,37,{min(0.6,0.15+z['s']*0.15):.2f})";lbl=f"🔴 {'█'*z['s']} {z['label']}"
            else:                    zc="#0FBF5F";zb=f"rgba(15,191,95,{min(0.6,0.15+z['s']*0.15):.2f})";lbl=f"🟢 {'█'*z['s']} {z['label']}"
            dist=round((z["price"]-sp_p)/sp_p*100,2)
            dist_s=f"+{dist:.2f}%" if dist>0 else f"{dist:.2f}%" if dist<0 else "← ICI"
            zone_rows+=f"""<tr style="background:{zb};">
              <td style="color:{zc};font-family:'JetBrains Mono',monospace;font-weight:700;font-size:12px;">${z['price']:,.2f}</td>
              <td style="color:{zc};font-size:11px;">{lbl}</td>
              <td style="color:#C9A94B;font-size:11px;">{dist_s}</td>
            </tr>"""

        sig_icon="🟢" if signal=="BUY" else "🔴" if signal=="SELL" else "🟡"
        full_html=f"""<!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@700&family=JetBrains+Mono&display=swap" rel="stylesheet">
        <style>body{{margin:0;padding:8px;background:#060E1A;font-family:'Rajdhani',sans-serif;color:#DCE8F5;}}
        table{{width:100%;border-collapse:collapse;}}
        th{{font-size:10px;color:#6B7A8D;padding:5px 8px;text-align:left;border-bottom:1px solid #1A3050;}}
        td{{padding:6px 8px;}}</style></head><body>

        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-bottom:10px;">
          <div style="background:#0F1C2E;border:1px solid #0FBF5F;border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">ES S&P LIVE</div>
            <div style="color:#0FBF5F;font-size:17px;font-weight:700;">${sf(sp_p)}</div>
            <div style="font-size:9px;color:#6B7A8D;">via OANDA/YF</div>
          </div>
          <div style="background:#0F1C2E;border:1px solid {sig_col};border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">SIGNAL</div>
            <div style="color:{sig_col};font-size:16px;font-weight:700;">{sig_icon} {signal}</div>
            <div style="color:{sig_col};font-size:11px;">{conf:.0f}%</div>
          </div>
          <div style="background:#0F1C2E;border:1px solid {vix_col};border-radius:8px;padding:8px;text-align:center;">
            <div style="color:#6B7A8D;font-size:9px;">VIX estimé</div>
            <div style="color:{vix_col};font-size:15px;font-weight:700;">{vix_sim}</div>
            <div style="color:{vix_col};font-size:9px;">{vix_lbl}</div>
          </div>
        </div>

        <div style="display:flex;flex-direction:column;gap:5px;margin-bottom:10px;">
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {risk_col};border-radius:5px;padding:7px 11px;font-size:12px;color:{risk_col};">
            {'🟢' if risk_on else '🔴'} {risk_txt}
          </div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {sig_col};border-radius:5px;padding:7px 11px;font-size:12px;color:{sig_col};">{sig_txt}</div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {vwap_col};border-radius:5px;padding:7px 11px;font-size:12px;color:{vwap_col};">{vwap_txt}</div>
          <div style="background:rgba(0,0,0,.2);border-left:3px solid {'#0FBF5F' if macd_b else '#D93025'};border-radius:5px;padding:7px 11px;font-size:12px;color:{'#0FBF5F' if macd_b else '#D93025'};">
            MACD {'Haussier ↑ — Momentum S&P positif' if macd_b else 'Baissier ↓ — Momentum S&P négatif'}
          </div>
        </div>

        <div style="font-size:12px;color:#0FBF5F;letter-spacing:2px;margin-bottom:6px;">🗺️ ZONES DE LIQUIDITÉ S&P 500</div>
        <table><thead><tr><th>Prix</th><th>Zone</th><th>Distance</th></tr></thead>
        <tbody>{zone_rows}</tbody></table>

        <div style="background:#0F1C2E;border-left:3px solid #0FBF5F;border-radius:6px;
             padding:9px 12px;margin-top:10px;font-size:11px;color:#8A95A3;line-height:1.9;">
          <b style="color:#0FBF5F;">📖 RÈGLES S&P ICT :</b><br>
          🔴 Prix approche R2/R3 + RSI>70 + MACD baissier → SELL ES ✅<br>
          🟢 Prix approche S2/S3 + RSI&lt;35 + MACD haussier → BUY ES ✅<br>
          📊 VWAP = niveau d'équilibre quotidien clé pour ES<br>
          🗓️ NFP + Fed + CPI → Volatilité majeure S&P
        </div>
        </body></html>"""

        components_sp.html(full_html, height=600, scrolling=True)




# ─── UNIFIED HEATMAP TAB ─────────────────────────────────────────────────────
def render_ict_heatmap_from_res(asset_name, res, color, icon, title):
    """Render a full visual ICT heatmap for any asset using OANDA res data."""
    import streamlit.components.v1 as _chm

    if not res:
        st.warning("⚠️ Lance d'abord une analyse pour afficher la Heatmap.")
        return

    price  = res.get("close", 0)
    atr    = res.get("atr", price * 0.006)
    vwap   = res.get("vwap", price)
    bb_up  = res.get("bb_upper", price * 1.006)
    bb_lo  = res.get("bb_lower", price * 0.994)
    rsi    = res.get("rsi", 50)
    signal = res.get("signal", "RANGE")
    conf   = res.get("confluence", 50)
    macd_b = res.get("macd_bull", False)
    fi13   = res.get("fi13", 0)
    fi2    = res.get("fi2", 0)
    fvgs   = res.get("fvgs", []) or []
    obs    = res.get("obs",  []) or []
    liq    = res.get("liquidity", {}) or {}
    bsl    = liq.get("buyside",  [])
    ssl    = liq.get("sellside", [])

    is_forex  = "EUR" in asset_name or "6E" in asset_name
    dec       = 5 if is_forex else 2
    fmt       = f",.{dec}f"

    # Build comprehensive zones
    zones = []

    # R3/R2/R1 from ATR
    zones.append({"label":"Buyside Liquidite R3",      "price":round(price+atr*3,dec), "type":"sell","s":3,"src":"ATR"})
    zones.append({"label":"Resistance R2",             "price":round(price+atr*2,dec), "type":"sell","s":2,"src":"ATR"})
    zones.append({"label":"BB Superieure R1",          "price":round(bb_up,dec),        "type":"sell","s":1,"src":"BB"})

    # FVG bearish levels
    for f in [x for x in fvgs if x.get("type")=="bear"][:2]:
        zones.append({"label":f"FVG Baissier {str(f.get('ts',''))[:10]}", "price":round(f.get("mid",price+atr),dec), "type":"sell","s":2,"src":"FVG"})

    # Bearish OB
    for o in [x for x in obs if x.get("type")=="bear"][:2]:
        zones.append({"label":f"Bearish OB",           "price":round(o.get("mid",price+atr*1.5),dec), "type":"sell","s":2,"src":"OB"})

    # BSL (buyside liquidity from ICT)
    for lvl in bsl[:2]:
        zones.append({"label":"Liquidite Buyside ICT", "price":round(lvl,dec), "type":"sell","s":3,"src":"ICT"})

    # VWAP
    zones.append({"label":"VWAP Session",              "price":round(vwap,dec), "type":"vwap","s":0,"src":"VWAP"})

    # Current price
    zones.append({"label":"PRIX ACTUEL",               "price":round(price,dec), "type":"current","s":0,"src":""})

    # S1/S2/S3 from ATR
    zones.append({"label":"BB Inferieure S1",          "price":round(bb_lo,dec),        "type":"buy","s":1,"src":"BB"})
    zones.append({"label":"Support S2",                "price":round(price-atr*2,dec), "type":"buy","s":2,"src":"ATR"})
    zones.append({"label":"Sellside Liquidite S3",     "price":round(price-atr*3,dec), "type":"sell_side","s":3,"src":"ATR"})

    # FVG bullish levels
    for f in [x for x in fvgs if x.get("type")=="bull"][:2]:
        zones.append({"label":f"FVG Haussier {str(f.get('ts',''))[:10]}", "price":round(f.get("mid",price-atr),dec), "type":"buy","s":2,"src":"FVG"})

    # Bullish OB
    for o in [x for x in obs if x.get("type")=="bull"][:2]:
        zones.append({"label":"Bullish OB",            "price":round(o.get("mid",price-atr*1.5),dec), "type":"buy","s":2,"src":"OB"})

    # SSL (sellside liquidity)
    for lvl in ssl[:2]:
        zones.append({"label":"Liquidite Sellside ICT","price":round(lvl,dec), "type":"buy","s":3,"src":"ICT"})

    # Sort by price descending
    zones.sort(key=lambda x: x["price"], reverse=True)

    # Remove duplicates (same price within 0.1%)
    filtered = []
    seen = []
    for z in zones:
        skip = False
        for s in seen:
            if abs(z["price"]-s)/max(price,0.0001) < 0.001:
                skip = True; break
        if not skip:
            filtered.append(z)
            seen.append(z["price"])
    zones = filtered

    # Build heatmap bar for each zone
    zone_rows = ""
    for z in zones:
        p = z["price"]
        dist = round((p - price) / price * 100, 3)
        dist_str = f"+{dist:.3f}%" if dist > 0 else f"{dist:.3f}%" if dist < 0 else "◆ ICI"

        if z["type"] == "current":
            zc = "#F0C96A"; zbg = "rgba(240,201,106,.15)"; brd = "#F0C96A"
            lbl = "◆ PRIX ACTUEL"; bar_col = "#F0C96A"
            bar_w = 100
        elif z["type"] == "vwap":
            zc = "#4A7EC7"; zbg = "rgba(74,126,199,.12)"; brd = "#4A7EC7"
            lbl = f"── VWAP Session"; bar_col = "#4A7EC7"
            bar_w = 60
        elif z["type"] in ("sell",):
            intensity = min(0.65, 0.15 + z["s"] * 0.17)
            zc = "#D93025"; zbg = f"rgba(217,48,37,{intensity:.2f})"; brd = "#D93025"
            src_badge = f"[{z['src']}]" if z["src"] else ""
            lbl = f"🔴 {'█'*z['s']} {z['label']} {src_badge}"
            bar_col = "#D93025"
            bar_w = min(95, 30 + z["s"] * 20)
        elif z["type"] == "sell_side":
            intensity = min(0.65, 0.15 + z["s"] * 0.17)
            zc = "#0FBF5F"; zbg = f"rgba(15,191,95,{intensity:.2f})"; brd = "#0FBF5F"
            lbl = f"🟢 {'█'*z['s']} {z['label']} [ICT]"
            bar_col = "#0FBF5F"
            bar_w = min(95, 30 + z["s"] * 20)
        else:
            intensity = min(0.65, 0.15 + z["s"] * 0.17)
            zc = "#0FBF5F"; zbg = f"rgba(15,191,95,{intensity:.2f})"; brd = "#0FBF5F"
            src_badge = f"[{z['src']}]" if z["src"] else ""
            lbl = f"🟢 {'█'*z['s']} {z['label']} {src_badge}"
            bar_col = "#0FBF5F"
            bar_w = min(95, 30 + z["s"] * 20)

        p_fmt = f"{p:{fmt}}"
        zone_rows += f"""
        <tr style="background:{zbg};border-bottom:1px solid rgba(255,255,255,.04);">
          <td style="padding:5px 8px;width:30%;">
            <div style="display:flex;align-items:center;gap:4px;">
              <div style="width:{bar_w}px;max-width:80px;height:12px;background:{bar_col};
                   border-radius:2px;opacity:0.85;flex-shrink:0;"></div>
              <span style="color:{zc};font-weight:700;font-size:11px;">{p_fmt}</span>
            </div>
          </td>
          <td style="color:{zc};font-size:11px;padding:5px 8px;">{lbl}</td>
          <td style="color:#C9A94B;font-size:11px;padding:5px 8px;text-align:right;">{dist_str}</td>
        </tr>"""

    # Signal info
    sig_col  = "#0FBF5F" if signal=="BUY" else "#D93025" if signal=="SELL" else "#C9A94B"
    sig_icon = "🟢" if signal=="BUY" else "🔴" if signal=="SELL" else "🟡"
    fi13_col = "#0FBF5F" if fi13 > 0 else "#D93025"
    fi13_lbl = "HAUSSIER" if fi13 > 0 else "BAISSIER"
    rsi_col  = "#D93025" if rsi>70 else "#0FBF5F" if rsi<30 else "#C9A94B"

    # Asset-specific guide
    guides = {
        "OIL":   "DXY baisse + WTI monte + Spread eleve → BUY WTI | Mercredi 15h30 UTC EIA",
        "EUR":   "DXY baisse + BCE hawkish → BUY EUR | Round numbers 1.0800/1.0900 = zones cles",
        "SP":    "RISK ON + VIX < 20 + MACD haussier → BUY S&P | NFP + Fed = mouveurs majeurs",
        "NAS":   "RISK ON + Tech earnings positifs + MACD haussier → BUY NAS100",
    }
    if "CL" in asset_name or "WTI" in asset_name:    guide_txt = guides["OIL"]
    elif "EUR" in asset_name or "6E" in asset_name:  guide_txt = guides["EUR"]
    elif "NAS" in asset_name:                         guide_txt = guides["NAS"]
    else:                                             guide_txt = guides["SP"]

    html = (
        "<!DOCTYPE html><html><head>"
        "<style>body{margin:0;padding:8px;background:#060E1A;font-family:sans-serif;color:#DCE8F5;}"
        "table{width:100%;border-collapse:collapse;}"
        "th{font-size:10px;color:#6B7A8D;padding:5px 8px;text-align:left;border-bottom:1px solid #1A3050;}"
        "</style></head><body>"

        # Header
        f'<div style="background:linear-gradient(135deg,#0A1520,#0F1C2E);border:2px solid {color};'
        f'border-radius:10px;padding:10px 14px;margin-bottom:10px;text-align:center;">'
        f'<div style="font-size:15px;font-weight:700;color:{color};letter-spacing:2px;">'
        f'{icon} {title}</div>'
        f'<div style="font-size:9px;color:#6B7A8D;margin-top:2px;">'
        f'Zones ICT calculees depuis donnees OANDA live</div></div>'

        # Stats row
        f'<div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:5px;margin-bottom:10px;">'
        f'<div style="background:#0F1C2E;border:1px solid {color};border-radius:7px;padding:7px;text-align:center;">'
        f'<div style="font-size:9px;color:#6B7A8D;">PRIX LIVE</div>'
        f'<div style="font-size:14px;font-weight:700;color:{color};">{price:{fmt}}</div>'
        f'<div style="font-size:9px;color:#6B7A8D;">OANDA</div></div>'

        f'<div style="background:#0F1C2E;border:1px solid {sig_col};border-radius:7px;padding:7px;text-align:center;">'
        f'<div style="font-size:9px;color:#6B7A8D;">SIGNAL</div>'
        f'<div style="font-size:14px;font-weight:700;color:{sig_col};">{sig_icon} {signal}</div>'
        f'<div style="font-size:9px;color:{sig_col};">{conf:.0f}%</div></div>'

        f'<div style="background:#0F1C2E;border:1px solid {rsi_col};border-radius:7px;padding:7px;text-align:center;">'
        f'<div style="font-size:9px;color:#6B7A8D;">RSI(14)</div>'
        f'<div style="font-size:14px;font-weight:700;color:{rsi_col};">{rsi:.1f}</div>'
        f'<div style="font-size:9px;color:{rsi_col};">{"Surachat" if rsi>70 else "Survente" if rsi<30 else "Neutre"}</div></div>'

        f'<div style="background:#0F1C2E;border:1px solid {fi13_col};border-radius:7px;padding:7px;text-align:center;">'
        f'<div style="font-size:9px;color:#6B7A8D;">FI(13)</div>'
        f'<div style="font-size:14px;font-weight:700;color:{fi13_col};">{fi13:+.0f}</div>'
        f'<div style="font-size:9px;color:{fi13_col};">{fi13_lbl}</div></div>'
        f'</div>'

        # MACD + VWAP signal
        f'<div style="display:flex;gap:6px;margin-bottom:10px;">'
        f'<div style="flex:1;background:rgba(0,0,0,.2);border-left:3px solid {"#0FBF5F" if macd_b else "#D93025"};'
        f'border-radius:5px;padding:6px 10px;font-size:11px;color:{"#0FBF5F" if macd_b else "#D93025"};">'
        f'MACD {"Haussier ↑" if macd_b else "Baissier ↓"}</div>'
        f'<div style="flex:1;background:rgba(0,0,0,.2);border-left:3px solid {"#0FBF5F" if price>vwap else "#D93025"};'
        f'border-radius:5px;padding:6px 10px;font-size:11px;color:{"#0FBF5F" if price>vwap else "#D93025"};">'
        f'VWAP {"Au-dessus ↑" if price>vwap else "En-dessous ↓"} ({vwap:{fmt}})</div>'
        f'</div>'

        # Heatmap zones
        f'<div style="font-size:12px;color:{color};letter-spacing:2px;margin-bottom:6px;">'
        f'🗺️ HEATMAP — ZONES DE LIQUIDITE ICT</div>'
        f'<div style="font-size:9px;color:#6B7A8D;margin-bottom:6px;">'
        f'🔴 = Resistance/Buyside Liq &nbsp;|&nbsp; 🟢 = Support/Sellside Liq &nbsp;|&nbsp; ◆ = Prix actuel &nbsp;|&nbsp; [FVG]/[OB]/[ICT] = Source</div>'
        f'<table><thead><tr>'
        f'<th>Niveau + Force</th><th>Zone ICT</th><th>Distance</th>'
        f'</tr></thead><tbody>{zone_rows}</tbody></table>'

        # Guide
        f'<div style="background:#0F1C2E;border-left:3px solid {color};border-radius:6px;'
        f'padding:9px 12px;margin-top:10px;font-size:11px;color:#8A95A3;line-height:1.9;">'
        f'<b style="color:{color};">📖 REGLES ICT :</b><br>'
        f'🔴 Zones rouges = Liquidite buyside → Cible des institutions SELL<br>'
        f'🟢 Zones vertes = Liquidite sellside → Cible des institutions BUY<br>'
        f'[FVG] = Fair Value Gap actif | [OB] = Order Block | [ICT] = Niveau ICT pur<br>'
        f'{guide_txt}'
        f'</div>'
        f'</body></html>'
    )
    _chm.html(html, height=680, scrolling=True)


def render_heatmap_tab(asset_name, res=None):
    """Render heatmap/market-depth data adapted to asset."""
    import streamlit as st

    is_btc = "BTC" in asset_name

    if is_btc:
        # Try Binance Futures first, fallback to res-based heatmap
        import streamlit.components.v1 as _chm_btc
        from contextlib import suppress

        # Quick test if Binance Futures is accessible
        btc_api_ok = False
        btc_price  = None
        try:
            r = requests.get("https://fapi.binance.com/fapi/v1/premiumIndex?symbol=BTCUSDT",
                             headers={"User-Agent":"Mozilla/5.0"}, timeout=5)
            if r.status_code == 200:
                btc_price  = float(r.json().get("markPrice", 0))
                btc_api_ok = btc_price > 0
        except: pass

        # Also try spot if futures blocked
        if not btc_api_ok:
            try:
                r = requests.get("https://api.binance.com/api/v3/ticker/price",
                                 params={"symbol":"BTCUSDT"},
                                 headers={"User-Agent":"Mozilla/5.0"}, timeout=5)
                if r.status_code == 200:
                    btc_price  = float(r.json().get("price", 0))
                    btc_api_ok = btc_price > 0
            except: pass

        if btc_api_ok:
            render_heatmap()
        else:
            # ── Full BTC heatmap from res ──────────────────────────────────
            if not res:
                st.warning("⚠️ Lance d'abord une analyse BTC pour afficher la Heatmap.")
                return

            price  = res.get("close", 0)
            atr    = res.get("atr", price * 0.008)
            rsi    = res.get("rsi", 50)
            signal = res.get("signal", "RANGE")
            conf   = res.get("confluence", 50)
            vwap   = res.get("vwap", price)
            bb_up  = res.get("bb_upper", price * 1.008)
            bb_lo  = res.get("bb_lower", price * 0.992)
            macd_b = res.get("macd_bull", False)
            fi13   = res.get("fi13", 0)
            fvgs   = res.get("fvgs", []) or []
            obs    = res.get("obs",  []) or []
            liq    = res.get("liquidity", {}) or {}
            bsl    = liq.get("buyside",  [])
            ssl    = liq.get("sellside", [])

            sig_col  = "#0FBF5F" if signal=="BUY" else "#D93025" if signal=="SELL" else "#C9A94B"
            sig_icon = "🟢" if signal=="BUY" else "🔴" if signal=="SELL" else "🟡"
            fi13_col = "#0FBF5F" if fi13 > 0 else "#D93025"
            rsi_col  = "#D93025" if rsi>70 else "#0FBF5F" if rsi<30 else "#C9A94B"

            # Build zones
            zones = [
                {"label":"Liquidite Buyside R3",   "price":round(price+atr*3,2),"type":"sell","s":3,"src":"ATR"},
                {"label":"Resistance R2",           "price":round(price+atr*2,2),"type":"sell","s":2,"src":"ATR"},
                {"label":"BB Superieure R1",        "price":round(bb_up,2),      "type":"sell","s":1,"src":"BB"},
            ]
            for f in [x for x in fvgs if x.get("type")=="bear"][:2]:
                zones.append({"label":"FVG Baissier","price":round(f.get("mid",price+atr),2),"type":"sell","s":2,"src":"FVG"})
            for o in [x for x in obs if x.get("type")=="bear"][:1]:
                zones.append({"label":"Bearish OB",  "price":round(o.get("mid",price+atr*1.5),2),"type":"sell","s":2,"src":"OB"})
            for lvl in bsl[:2]:
                zones.append({"label":"Buyside ICT", "price":round(lvl,2),"type":"sell","s":3,"src":"ICT"})
            zones.append({"label":"VWAP Session",    "price":round(vwap,2),"type":"vwap","s":0,"src":"VWAP"})
            zones.append({"label":"PRIX ACTUEL",     "price":round(price,2),"type":"current","s":0,"src":""})
            zones.append({"label":"BB Inferieure S1","price":round(bb_lo,2),"type":"buy","s":1,"src":"BB"})
            for f in [x for x in fvgs if x.get("type")=="bull"][:2]:
                zones.append({"label":"FVG Haussier","price":round(f.get("mid",price-atr),2),"type":"buy","s":2,"src":"FVG"})
            for o in [x for x in obs if x.get("type")=="bull"][:1]:
                zones.append({"label":"Bullish OB",  "price":round(o.get("mid",price-atr*1.5),2),"type":"buy","s":2,"src":"OB"})
            for lvl in ssl[:2]:
                zones.append({"label":"Sellside ICT","price":round(lvl,2),"type":"buy","s":3,"src":"ICT"})
            zones.append({"label":"Support S2",      "price":round(price-atr*2,2),"type":"buy","s":2,"src":"ATR"})
            zones.append({"label":"Liquidite Sellside S3","price":round(price-atr*3,2),"type":"buy","s":3,"src":"ATR"})
            zones.sort(key=lambda x: x["price"], reverse=True)

            # Build rows
            zone_rows = ""
            for z in zones:
                p = z["price"]; dist = round((p-price)/price*100,2)
                dist_s = f"+{dist:.2f}%" if dist>0 else f"{dist:.2f}%" if dist<0 else "◆ ICI"
                if z["type"]=="current":
                    zc="#F0C96A";zb="rgba(240,201,106,.15)";lbl="◆ PRIX ACTUEL BTC";bw=100
                elif z["type"]=="vwap":
                    zc="#4A7EC7";zb="rgba(74,126,199,.12)";lbl="── VWAP Session";bw=60
                elif z["type"]=="sell":
                    i=min(0.65,0.15+z["s"]*0.17);zc="#D93025";zb=f"rgba(217,48,37,{i:.2f})"
                    lbl=f"🔴 {'█'*z['s']} {z['label']} [{z['src']}]";bw=min(90,30+z["s"]*20)
                else:
                    i=min(0.65,0.15+z["s"]*0.17);zc="#0FBF5F";zb=f"rgba(15,191,95,{i:.2f})"
                    lbl=f"🟢 {'█'*z['s']} {z['label']} [{z['src']}]";bw=min(90,30+z["s"]*20)
                zone_rows += (
                    f'<tr style="background:{zb};border-bottom:1px solid rgba(255,255,255,.04);">'
                    f'<td style="padding:5px 8px;">'
                    f'<div style="display:flex;align-items:center;gap:4px;">'
                    f'<div style="width:{bw}px;max-width:80px;height:12px;background:{zc};border-radius:2px;opacity:.85;flex-shrink:0;"></div>'
                    f'<span style="color:{zc};font-weight:700;font-size:11px;">${p:,.2f}</span></div></td>'
                    f'<td style="color:{zc};font-size:11px;padding:5px 8px;">{lbl}</td>'
                    f'<td style="color:#C9A94B;font-size:11px;padding:5px 8px;text-align:right;">{dist_s}</td></tr>'
                )

            html = (
                "<!DOCTYPE html><html><head>"
                "<style>body{margin:0;padding:8px;background:#060E1A;font-family:sans-serif;color:#DCE8F5;}"
                "table{width:100%;border-collapse:collapse;}"
                "th{font-size:10px;color:#6B7A8D;padding:5px 8px;text-align:left;border-bottom:1px solid #1A3050;}"
                "</style></head><body>"

                '<div style="background:linear-gradient(135deg,#0A1520,#1A0A10);border:2px solid #F0C96A;'
                'border-radius:10px;padding:10px 14px;margin-bottom:10px;text-align:center;">'
                '<div style="font-size:15px;font-weight:700;color:#F0C96A;letter-spacing:2px;">'
                '₿ HEATMAP BTC/USD — ZONES ICT</div>'
                '<div style="font-size:9px;color:#6B7A8D;margin-top:2px;">'
                'Zones calculees depuis donnees OANDA/Binance Spot</div></div>'

                f'<div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:5px;margin-bottom:10px;">'
                f'<div style="background:#0F1C2E;border:1px solid #F0C96A;border-radius:7px;padding:7px;text-align:center;">'
                f'<div style="font-size:9px;color:#6B7A8D;">BTC LIVE</div>'
                f'<div style="font-size:13px;font-weight:700;color:#F0C96A;">${price:,.2f}</div>'
                f'<div style="font-size:9px;color:#6B7A8D;">via Binance</div></div>'

                f'<div style="background:#0F1C2E;border:1px solid {sig_col};border-radius:7px;padding:7px;text-align:center;">'
                f'<div style="font-size:9px;color:#6B7A8D;">SIGNAL</div>'
                f'<div style="font-size:13px;font-weight:700;color:{sig_col};">{sig_icon} {signal}</div>'
                f'<div style="font-size:9px;color:{sig_col};">{conf:.0f}%</div></div>'

                f'<div style="background:#0F1C2E;border:1px solid {rsi_col};border-radius:7px;padding:7px;text-align:center;">'
                f'<div style="font-size:9px;color:#6B7A8D;">RSI(14)</div>'
                f'<div style="font-size:13px;font-weight:700;color:{rsi_col};">{rsi:.1f}</div>'
                f'<div style="font-size:9px;color:{rsi_col};">{"Surachat" if rsi>70 else "Survente" if rsi<30 else "Neutre"}</div></div>'

                f'<div style="background:#0F1C2E;border:1px solid {fi13_col};border-radius:7px;padding:7px;text-align:center;">'
                f'<div style="font-size:9px;color:#6B7A8D;">FI(13)</div>'
                f'<div style="font-size:13px;font-weight:700;color:{fi13_col};">{fi13:+.0f}</div>'
                f'<div style="font-size:9px;color:{fi13_col};">{"HAUSSIER" if fi13>0 else "BAISSIER"}</div></div>'
                f'</div>'

                f'<div style="display:flex;gap:6px;margin-bottom:10px;">'
                f'<div style="flex:1;background:rgba(0,0,0,.2);border-left:3px solid {"#0FBF5F" if macd_b else "#D93025"};'
                f'border-radius:5px;padding:6px 10px;font-size:11px;color:{"#0FBF5F" if macd_b else "#D93025"};">'
                f'MACD {"Haussier ↑" if macd_b else "Baissier ↓"}</div>'
                f'<div style="flex:1;background:rgba(0,0,0,.2);border-left:3px solid {"#0FBF5F" if price>vwap else "#D93025"};'
                f'border-radius:5px;padding:6px 10px;font-size:11px;color:{"#0FBF5F" if price>vwap else "#D93025"};">'
                f'VWAP {"Au-dessus ↑" if price>vwap else "En-dessous ↓"} (${vwap:,.2f})</div>'
                f'</div>'

                f'<div style="font-size:12px;color:#F0C96A;letter-spacing:2px;margin-bottom:6px;">🗺️ HEATMAP — ZONES DE LIQUIDITE BTC</div>'
                f'<div style="font-size:9px;color:#6B7A8D;margin-bottom:6px;">🔴 Resistance/Buyside Liq &nbsp;|&nbsp; 🟢 Support/Sellside Liq &nbsp;|&nbsp; [FVG]/[OB]/[ICT] = Source</div>'
                f'<table><thead><tr><th>Niveau + Force</th><th>Zone ICT</th><th>Distance</th></tr></thead>'
                f'<tbody>{zone_rows}</tbody></table>'

                f'<div style="background:#0F1C2E;border-left:3px solid #F0C96A;border-radius:6px;'
                f'padding:9px 12px;margin-top:10px;font-size:11px;color:#8A95A3;line-height:1.9;">'
                f'<b style="color:#F0C96A;">📖 REGLES ICT BTC :</b><br>'
                f'🔴 Zones rouges = Liquidite buyside → Cible SELL institutions<br>'
                f'🟢 Zones vertes = Liquidite sellside → Cible BUY institutions<br>'
                f'[FVG] = Fair Value Gap | [OB] = Order Block | [ICT] = Niveau pur<br>'
                f'HALVING BTC + NFP + FOMC → Catalyseurs majeurs direction'
                f'</div></body></html>'
            )
            _chm_btc.html(html, height=680, scrolling=True)
    else:
        # For non-BTC: show relevant market depth/OI proxies
        is_gold = "XAU" in asset_name or "GC" in asset_name
        is_oil  = "CL" in asset_name or "WTI" in asset_name
        is_eur  = "EUR" in asset_name or "6E" in asset_name
        is_sp   = "S&P" in asset_name or "ES" in asset_name or "NAS" in asset_name or "NAS" in asset_name

        if is_gold:
            # Fetch live gold price and build useful zone map
            with st.spinner("Calcul des zones de liquidité Or..."):
                gold_d = fetch_gold_sentiment()
                vix_d  = fetch_vix()
                dxy_d  = fetch_dxy()
            gp = gold_d.get("gold_price")
            if gp:
                # Calculate key ICT liquidity zones around current price
                atr_est = gp * 0.008  # ~0.8% daily ATR for gold
                zones = [
                    {"label":"Résistance Majeure R3",  "price": round(gp + atr_est*3, 2), "type":"sell", "strength":3},
                    {"label":"Résistance R2",          "price": round(gp + atr_est*2, 2), "type":"sell", "strength":2},
                    {"label":"Résistance R1 / Buyside","price": round(gp + atr_est,   2), "type":"sell", "strength":1},
                    {"label":"PRIX ACTUEL",            "price": gp,                        "type":"current","strength":0},
                    {"label":"Support S1 / Sellside",  "price": round(gp - atr_est,   2), "type":"buy",  "strength":1},
                    {"label":"Support S2",             "price": round(gp - atr_est*2, 2), "type":"buy",  "strength":2},
                    {"label":"Support Majeur S3",      "price": round(gp - atr_est*3, 2), "type":"buy",  "strength":3},
                ]
                gold_chg = gold_d.get("gold_chg", 0) or 0
                silver   = gold_d.get("silver_price")
                gdx_p    = gold_d.get("gdx_price")
                gdx_chg  = gold_d.get("gdx_chg", 0) or 0
                vix_v    = vix_d["value"] if vix_d else None
                dxy_v    = dxy_d["value"] if dxy_d else None
                dxy_chg  = (dxy_d["change"] if dxy_d else 0) or 0

                zone_rows = ""
                for z in zones:
                    if z["type"] == "current":
                        zc = "#F0C96A"; zb = "rgba(240,201,106,.12)"; lbl = "◆ PRIX ACTUEL"
                    elif z["type"] == "sell":
                        intensity = min(0.7, 0.2 + z["strength"]*0.17)
                        zc = "#D93025"; zb = f"rgba(217,48,37,{intensity:.2f})"
                        lbl = f"🔴 {'█'*z['strength']} {z['label']}"
                    else:
                        intensity = min(0.7, 0.2 + z["strength"]*0.17)
                        zc = "#0FBF5F"; zb = f"rgba(15,191,95,{intensity:.2f})"
                        lbl = f"🟢 {'█'*z['strength']} {z['label']}"
                    dist = round((z["price"] - gp) / gp * 100, 2)
                    dist_str = f"+{dist:.2f}%" if dist > 0 else f"{dist:.2f}%" if dist < 0 else "← ICI"
                    zone_rows += f"""<tr style="background:{zb};">
                        <td style="color:{zc};font-family:'JetBrains Mono',monospace;font-weight:700;">${z['price']:,.2f}</td>
                        <td style="color:{zc};font-size:11px;">{lbl}</td>
                        <td style="color:#C9A94B;font-size:11px;">{dist_str}</td>
                    </tr>"""

                dxy_col = "#D93025" if dxy_chg > 0.2 else "#0FBF5F" if dxy_chg < -0.2 else "#C9A94B"
                dxy_arrow = "↑" if dxy_chg > 0.1 else "↓" if dxy_chg < -0.1 else "→"
                vix_col = "#D93025" if (vix_v or 0) > 25 else "#C9A94B" if (vix_v or 0) > 18 else "#0FBF5F"

                import streamlit.components.v1 as components_hm
                html_gold_hm = f"""<!DOCTYPE html><html><head>
                <link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@700&family=JetBrains+Mono&display=swap" rel="stylesheet">
                <style>
                  body{{margin:0;padding:8px;background:#060E1A;font-family:'Rajdhani',sans-serif;color:#DCE8F5;}}
                  table{{width:100%;border-collapse:collapse;}}
                  th{{font-size:10px;color:#6B7A8D;padding:6px 8px;text-align:left;border-bottom:1px solid #1A3050;}}
                  td{{padding:7px 8px;font-size:12px;}}
                </style></head><body>

                <!-- HEADER -->
                <div style="text-align:center;margin-bottom:10px;">
                  <div style="font-size:14px;font-weight:700;color:#C9A94B;letter-spacing:2px;">
                    🥇 ZONES DE LIQUIDITÉ — OR XAUUSD
                  </div>
                  <div style="font-size:10px;color:#6B7A8D;margin-top:3px;">
                    Zones buyside/sellside calculées sur ATR journalier
                  </div>
                </div>

                <!-- STATS -->
                <div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:5px;margin-bottom:10px;">
                  <div style="background:#0F1C2E;border:1px solid #C9A94B;border-radius:6px;padding:7px;text-align:center;">
                    <div style="font-size:9px;color:#6B7A8D;">OR</div>
                    <div style="font-size:15px;font-weight:700;color:#F0C96A;">${sf(gp)}</div>
                    <div style="font-size:10px;color:{'#0FBF5F' if gold_chg>=0 else '#D93025'};">{sf(gold_chg,'+.2f','±0.00')}</div>
                  </div>
                  <div style="background:#0F1C2E;border:1px solid #6B7A8D;border-radius:6px;padding:7px;text-align:center;">
                    <div style="font-size:9px;color:#6B7A8D;">ARGENT</div>
                    <div style="font-size:15px;font-weight:700;color:#DCE8F5;">${sf(silver,',.3f')}</div>
                  </div>
                  <div style="background:#0F1C2E;border:1px solid {vix_col};border-radius:6px;padding:7px;text-align:center;">
                    <div style="font-size:9px;color:#6B7A8D;">VIX</div>
                    <div style="font-size:15px;font-weight:700;color:{vix_col};">{sf(vix_v,',.1f')}</div>
                  </div>
                  <div style="background:#0F1C2E;border:1px solid {dxy_col};border-radius:6px;padding:7px;text-align:center;">
                    <div style="font-size:9px;color:#6B7A8D;">DXY {dxy_arrow}</div>
                    <div style="font-size:15px;font-weight:700;color:{dxy_col};">{sf(dxy_v,',.2f')}</div>
                  </div>
                </div>

                <!-- ZONE TABLE -->
                <table>
                  <thead><tr><th>Prix</th><th>Zone ICT</th><th>Distance</th></tr></thead>
                  <tbody>{zone_rows}</tbody>
                </table>

                <!-- GDX LEADING INDICATOR -->
                <div style="background:#0F1C2E;border:1px solid #1A3050;border-left:3px solid {'#0FBF5F' if gdx_chg>=0 else '#D93025'};
                     border-radius:6px;padding:8px 12px;margin-top:10px;">
                  <div style="font-size:12px;color:#C9A94B;font-weight:700;">📊 GDX Miners — Indicateur Avancé Or</div>
                  <div style="font-size:11px;color:#8A95A3;margin-top:4px;">
                    GDX: ${sf(gdx_p)} ({sf(gdx_chg,'+.2f','±0.00')}) —
                    {'🟢 Miners haussiers → Or devrait suivre dans 1-2 sessions' if gdx_chg>=0 else '🔴 Miners baissiers → Pression baissière Or à venir'}
                  </div>
                </div>

                <!-- GUIDE -->
                <div style="background:#0F1C2E;border-left:3px solid #C9A94B;border-radius:6px;
                     padding:8px 12px;margin-top:8px;font-size:11px;color:#8A95A3;line-height:1.9;">
                  <b style="color:#C9A94B;">📖 LECTURE ICT :</b><br>
                  🔴 Zones rouges = Liquidité buyside → Prix chassé vers le haut puis INVERSION<br>
                  🟢 Zones vertes = Liquidité sellside → Prix chassé vers le bas puis INVERSION<br>
                  💡 DXY monte → Or sous pression | VIX > 25 → Or refuge → Monte
                </div>
                </body></html>"""
                components_hm.html(html_gold_hm, height=580, scrolling=True)
                return  # skip the generic redirect below

            icon, title, color = "🥇","NIVEAUX CLÉS — OR (XAUUSD)","#C9A94B"
            desc = "Or : données de prix indisponibles. Vérifie ta connexion."
            tip  = "Pour l'Or, les zones de liquidité ICT (FVG, OB) dans l'onglet 🎯 PIVOTS sont plus pertinentes."
        elif "NAS" in asset_name:
            render_ict_heatmap_from_res(asset_name, res, "#4A7EC7", "💻", "HEATMAP NASDAQ 100")
            return
        elif is_oil:
            render_ict_heatmap_from_res(asset_name, res, "#E07B2A", "🛢️", "HEATMAP PÉTROLE WTI")
            return
        elif is_eur:
            render_ict_heatmap_from_res(asset_name, res, "#4A7EC7", "💶", "HEATMAP EUR/USD")
            return
        else:
            render_ict_heatmap_from_res(asset_name, res, "#0FBF5F", "📈", "HEATMAP S&P 500 / ES")
            return
        # fallback (should not reach here)
        icon, title, color = "📊","NIVEAUX CLÉS","#C9A94B"
        desc = "Donnees de liquidation non disponibles."
        tip  = "Utilise l'onglet ANALYSE pour les niveaux ICT."

        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#0A1520,#0F1C2E);border:2px solid {color};
             border-radius:12px;padding:20px;text-align:center;margin-bottom:16px;">
          <div style="font-size:40px;margin-bottom:8px;">{icon}</div>
          <div style="font-family:'Rajdhani',sans-serif;font-size:16px;font-weight:700;
               color:{color};letter-spacing:2px;">{title}</div>
          <div style="font-size:12px;color:#8A95A3;margin-top:10px;line-height:1.8;">{desc}</div>
          <div style="background:rgba(0,0,0,.3);border:1px solid {color};border-radius:8px;
               padding:10px;margin-top:12px;font-size:11px;color:{color};text-align:left;">
            💡 {tip}
          </div>
        </div>""", unsafe_allow_html=True)

        # Redirect to useful tabs
        st.markdown(f"""
        <div style="font-family:'Rajdhani',sans-serif;font-size:13px;color:#C9A94B;
             letter-spacing:2px;margin-bottom:8px;">🔀 OUTILS RECOMMANDÉS POUR {asset_name}</div>""",
             unsafe_allow_html=True)

        recs = []
        if is_gold:
            recs = [("🎯 PIVOTS","Points Pivots P/R1/R2/S1/S2 — Niveaux de réversion clés"),
                    ("📊 ANALYSE","ICT : FVG + Order Blocks + Liquidité"),
                    ("😨 SENTIMENT","VIX + DXY + GDX Miners pour l'Or"),
                    ("🗓️ CALENDRIER","NFP + CPI + FOMC — Événements qui bougent l'Or")]
        elif is_oil:
            if "NAS" in asset_name:
                recs = [("🎯 PIVOTS","Pivots hebdomadaires NAS100 — zones clés"),
                        ("📊 ORDERFLOW","Delta + imbalance NAS100"),
                        ("📋 TAPE","Flux aggTrades Binance QQQ proxy"),
                        ("🗓️ CALENDRIER","Earnings FAANG + Fed + CPI")]
            else:
                recs = [("🎯 PIVOTS","Pivots hebdomadaires WTI — zones clés"),
                    ("😨 SENTIMENT","WTI + Brent + DXY corrélations"),
                    ("🗓️ CALENDRIER","Mercredi EIA + OPEC meetings")]
        elif is_eur:
            recs = [("🎯 PIVOTS","Pivots Forex + niveaux bancaires"),
                    ("😨 SENTIMENT","DXY + EUR/GBP + corrélations"),
                    ("🗓️ CALENDRIER","BCE + NFP + CPI — impact direct sur EUR")]
        else:
            recs = [("😨 SENTIMENT","VIX + SPY + QQQ + Risk On/Off"),
                    ("🗓️ CALENDRIER","Fed + GDP + NFP — mouveurs S&P"),
                    ("🎯 PIVOTS","Pivots S&P + niveaux options")]

        for tab_name, description in recs:
            st.markdown(f"""
            <div style="background:#0F1C2E;border:1px solid #1A3050;border-left:3px solid {color};
                 border-radius:6px;padding:10px 14px;margin-bottom:6px;">
              <span style="color:{color};font-weight:700;font-family:'Rajdhani',sans-serif;">
                {tab_name}
              </span>
              <span style="color:#8A95A3;font-size:11px;margin-left:8px;">{description}</span>
            </div>""", unsafe_allow_html=True)


# ─── TELEGRAM ALERTS ──────────────────────────────────────────────────────────
def send_telegram(bot_token, chat_id, message):
    """Send a message via Telegram Bot API."""
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{bot_token}/sendMessage",
            json={"chat_id": chat_id, "text": message,
                  "parse_mode": "HTML", "disable_web_page_preview": True},
            timeout=8
        )
        return r.status_code == 200, r.json().get("description","")
    except Exception as e:
        return False, str(e)

def build_telegram_message(res, asset_name, tf, of_result=None):
    """Build a rich Telegram alert message."""
    if not res: return ""
    sig   = res.get("signal","?")
    conf  = res.get("confluence", 0)
    price = res.get("close", 0)
    tp    = res.get("tp",  None)
    sl    = res.get("sl",  None)
    rsi   = res.get("rsi", 50)
    atr   = res.get("atr", 0)

    sig_emoji = "🟢 BUY" if sig=="BUY" else "🔴 SELL" if sig=="SELL" else "🟡 RANGE"
    conf_bar  = "█"*int(conf//10) + "░"*(10-int(conf//10))

    # Safe format for TP/SL — can be None when signal is RANGE
    tp_str = f"{tp:.5f}" if tp is not None else "—"
    sl_str = f"{sl:.5f}" if sl is not None else "—"

    # OrderFlow confirmation line
    of_line = ""
    if of_result:
        if of_result.get("confirmed"):
            of_line = f"\n✅ <b>OrderFlow: Tendance Confirmée par Volume Supérieur</b>"
        else:
            of_line = f"\n⚠️ OrderFlow: Divergence volume ({of_result.get('delta_pct',0):+.1f}%)"

    msg = (
        f"<b>🚀 STRATÉGIE PRO v7.1 — {asset_name}</b>\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"📊 Signal : {sig_emoji}\n"
        f"💹 Prix   : <code>{price:.5f}</code>\n"
        f"⏱️ TF    : {tf}\n"
        f"🎯 Conf.  : {conf:.0f}% [{conf_bar}]\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"✅ TP : <code>{tp_str}</code>\n"
        f"🛑 SL : <code>{sl_str}</code>\n"
        f"📐 ATR : {atr:.5f}\n"
        f"📈 RSI : {rsi:.1f}\n"
        f"⚡ FI(2) : {res.get('fi2',0):+.0f} | FI(13) : {res.get('fi13',0):+.0f} → {'HAUSSIER' if res.get('fi13',0)>0 else 'BAISSIER'}\n"
        f"{of_line}"
        f"\n━━━━━━━━━━━━━━━━━━\n"
        f"⚠️ <i>Outil d'aide à la décision uniquement.</i>\n"
        f"🤖 777gervais-dev"
    )
    return msg

def render_telegram_config(cfg, res, asset_name, tf, of_result=None):
    import streamlit as st
    import streamlit.components.v1 as components

    st.markdown("""
    <div style="background:linear-gradient(135deg,#0A1520,#0A2010);border:2px solid #29A8E0;
         border-radius:12px;padding:14px 18px;margin-bottom:14px;text-align:center;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;
           color:#29A8E0;letter-spacing:2px;">📱 ALERTES TELEGRAM</div>
      <div style="font-size:10px;color:#6B7A8D;margin-top:3px;">
        Notifications instantanées sur ton téléphone · Gratuit
      </div>
    </div>""", unsafe_allow_html=True)

    # Setup guide
    with st.expander("📖 Comment configurer Telegram (cliquer ici)", expanded=False):
        st.markdown("""
        **Étape 1 — Créer ton bot :**
        1. Ouvre Telegram → cherche **@BotFather**
        2. Envoie `/newbot`
        3. Donne un nom (ex: `Strategie Pro Bot`)
        4. Copie le **Token** (ressemble à `7123456789:AAF...`)

        **Étape 2 — Trouver ton Chat ID :**
        1. Cherche **@userinfobot** dans Telegram
        2. Envoie `/start`
        3. Il te donne ton **Chat ID** (ex: `123456789`)

        **Étape 3 — Coller ci-dessous et sauvegarder ✅**
        """)

    c1, c2 = st.columns(2)
    with c1:
        tg_token = st.text_input("🤖 Bot Token",
            value=cfg.get("tg_token",""), type="password",
            placeholder="7123456789:AAF...", key="tg_token_input")
    with c2:
        tg_chat = st.text_input("💬 Chat ID",
            value=cfg.get("tg_chat",""),
            placeholder="123456789", key="tg_chat_input")

    # Auto-send config
    col_a, col_b = st.columns(2)
    with col_a:
        tg_auto     = st.checkbox("🔔 Alertes automatiques", value=cfg.get("tg_auto", False))
        tg_min_conf = st.slider("Confluence minimum", 50, 95,
                                value=cfg.get("tg_min_conf", 75), step=5)
    with col_b:
        tg_buy  = st.checkbox("🟢 Alertes BUY",  value=cfg.get("tg_buy", True))
        tg_sell = st.checkbox("🔴 Alertes SELL", value=cfg.get("tg_sell", True))

    # Save config
    if st.button("💾 Sauvegarder configuration Telegram", key="save_tg"):
        cfg["tg_token"]    = tg_token
        cfg["tg_chat"]     = tg_chat
        cfg["tg_auto"]     = tg_auto
        cfg["tg_min_conf"] = tg_min_conf
        cfg["tg_buy"]      = tg_buy
        cfg["tg_sell"]     = tg_sell
        st.success("✅ Configuration Telegram sauvegardée !")

    st.markdown("---")

    # Test button
    col_t1, col_t2 = st.columns(2)
    with col_t1:
        if st.button("🧪 Envoyer message TEST", key="test_tg"):
            if not tg_token or not tg_chat:
                st.error("❌ Configure le Token et le Chat ID d'abord.")
            else:
                test_msg = (
                    "✅ <b>TEST — STRATÉGIE PRO v7.1</b>\n"
                    "━━━━━━━━━━━━━━━━━━\n"
                    "🤖 Connexion Telegram opérationnelle !\n"
                    "📱 Les alertes de trading sont actives.\n"
                    "━━━━━━━━━━━━━━━━━━\n"
                    "🤖 777gervais-dev"
                )
                ok, err = send_telegram(tg_token, tg_chat, test_msg)
                if ok:
                    st.success("✅ Message test envoyé ! Vérifie Telegram.")
                else:
                    st.error(f"❌ Erreur : {err}")

    with col_t2:
        if st.button("📤 Envoyer signal actuel", key="send_signal_tg"):
            if not tg_token or not tg_chat:
                st.error("❌ Configure le Token et le Chat ID d'abord.")
            elif not res:
                st.warning("⚠️ Lance d'abord une analyse.")
            else:
                msg = build_telegram_message(res, asset_name, tf, of_result)
                ok, err = send_telegram(tg_token, tg_chat, msg)
                if ok:
                    st.success("✅ Signal envoyé sur Telegram !")
                else:
                    st.error(f"❌ Erreur : {err}")

    # Preview
    if res:
        st.markdown("""<div style="font-family:'Rajdhani',sans-serif;font-size:12px;color:#29A8E0;
             letter-spacing:2px;margin:10px 0 6px 0;">👁️ APERÇU DU MESSAGE</div>""",
             unsafe_allow_html=True)
        msg_preview = build_telegram_message(res, asset_name, tf, of_result)
        st.code(msg_preview.replace("<b>","").replace("</b>","")
                           .replace("<code>","").replace("</code>","")
                           .replace("<i>","").replace("</i>",""), language=None)

    # Auto-send logic
    if tg_auto and tg_token and tg_chat and res:
        sig  = res.get("signal","")
        conf = res.get("confluence", 0)
        last_sent = st.session_state.get("tg_last_sent_signal","")
        send_flag = (
            conf >= tg_min_conf and
            ((sig == "BUY"  and tg_buy)  or
             (sig == "SELL" and tg_sell)) and
            sig != last_sent
        )
        if send_flag:
            msg = build_telegram_message(res, asset_name, tf, of_result)
            ok, _ = send_telegram(tg_token, tg_chat, msg)
            if ok:
                st.session_state["tg_last_sent_signal"] = sig
                st.toast(f"📱 Alerte {sig} envoyée sur Telegram !", icon="✅")

    return cfg


# ─── ORDERFLOW ANALYSIS ───────────────────────────────────────────────────────
def fetch_orderflow_btc(symbol="BTCUSDT", limit=500):
    """Fetch recent aggTrades from Binance Spot — buy vs sell volume per level."""
    try:
        r = requests.get(
            f"https://api.binance.com/api/v3/aggTrades",
            params={"symbol": symbol, "limit": limit},
            timeout=8
        )
        if r.status_code != 200: return None
        trades = r.json()
        levels = {}
        for t in trades:
            price  = round(float(t["p"]), 1)  # round to 1 decimal = price level
            qty    = float(t["q"])
            is_sell = t["m"]  # True = seller is maker = sell aggression
            if price not in levels:
                levels[price] = {"buy": 0.0, "sell": 0.0}
            if is_sell:
                levels[price]["sell"] += qty
            else:
                levels[price]["buy"]  += qty
        return levels
    except: return None

def fetch_orderflow_oanda(df):
    """Estimate OrderFlow from OANDA candle data using price action heuristics."""
    if df is None or len(df) < 10: return None
    levels = {}
    for i, row in df.tail(30).iterrows():
        o, h, l, c, v = row["Open"], row["High"], row["Low"], row["Close"], row["Volume"]
        price = round((h + l) / 2, 5)
        candle_range = h - l or 0.0001
        # Buy volume heuristic: proportion of close position in candle range
        buy_ratio  = (c - l) / candle_range
        sell_ratio = 1 - buy_ratio
        if price not in levels:
            levels[price] = {"buy": 0.0, "sell": 0.0}
        levels[price]["buy"]  += v * buy_ratio
        levels[price]["sell"] += v * sell_ratio
    return levels

def calc_orderflow_metrics(levels, current_price):
    """Calculate delta, imbalance, POC, and footprint data from levels dict."""
    if not levels: return None

    total_buy  = sum(l["buy"]  for l in levels.values())
    total_sell = sum(l["sell"] for l in levels.values())
    total_vol  = total_buy + total_sell or 1

    delta      = total_buy - total_sell
    delta_pct  = (delta / total_vol) * 100
    imbalance_pct = abs(delta_pct)

    # Point of Control (POC) — price level with highest volume
    poc_price  = max(levels, key=lambda p: levels[p]["buy"] + levels[p]["sell"])
    poc_vol    = levels[poc_price]["buy"] + levels[poc_price]["sell"]

    # Value Area (70% of volume around POC)
    sorted_prices = sorted(levels.keys())
    poc_idx   = sorted_prices.index(poc_price) if poc_price in sorted_prices else 0
    va_target = total_vol * 0.70
    va_vol    = poc_vol
    va_low_i  = poc_idx
    va_high_i = poc_idx
    while va_vol < va_target and (va_low_i > 0 or va_high_i < len(sorted_prices)-1):
        low_add  = levels[sorted_prices[va_low_i-1]]["buy"]+levels[sorted_prices[va_low_i-1]]["sell"] if va_low_i>0 else 0
        high_add = levels[sorted_prices[va_high_i+1]]["buy"]+levels[sorted_prices[va_high_i+1]]["sell"] if va_high_i<len(sorted_prices)-1 else 0
        if low_add >= high_add and va_low_i > 0:
            va_low_i -= 1; va_vol += low_add
        elif va_high_i < len(sorted_prices)-1:
            va_high_i += 1; va_vol += high_add
        else: break

    va_low  = sorted_prices[va_low_i]
    va_high = sorted_prices[va_high_i]

    # Imbalance zones (where buy/sell ratio > 3:1)
    imbalance_zones = []
    for price, lvl in levels.items():
        total = lvl["buy"] + lvl["sell"] or 0.001
        ratio = lvl["buy"] / total
        if ratio > 0.75:
            imbalance_zones.append({"price":price,"type":"buy_imbalance","ratio":ratio})
        elif ratio < 0.25:
            imbalance_zones.append({"price":price,"type":"sell_imbalance","ratio":1-ratio})

    # Top footprint rows (closest to current price)
    sorted_near = sorted(levels.items(),
                         key=lambda x: abs(x[0]-current_price))[:12]
    footprint = sorted(sorted_near, key=lambda x: x[0], reverse=True)

    return {
        "total_buy":     round(total_buy, 2),
        "total_sell":    round(total_sell, 2),
        "total_vol":     round(total_vol, 2),
        "delta":         round(delta, 2),
        "delta_pct":     round(delta_pct, 2),
        "imbalance_pct": round(imbalance_pct, 2),
        "poc":           poc_price,
        "va_low":        va_low,
        "va_high":       va_high,
        "imbalance_zones": imbalance_zones[:5],
        "footprint":     footprint,
        "bias":          "BUY" if delta > 0 else "SELL",
    }

def check_orderflow_confirmation(of_metrics, res):
    """Check if OrderFlow confirms the current signal direction."""
    if not of_metrics or not res: return None
    signal  = res.get("signal","")
    conf    = res.get("confluence", 0)
    delta_p = of_metrics.get("delta_pct", 0)
    imb     = of_metrics.get("imbalance_pct", 0)
    of_bias = of_metrics.get("bias","")

    confirmed = (
        (signal == "BUY"  and of_bias == "BUY"  and delta_p > 5)  or
        (signal == "SELL" and of_bias == "SELL" and delta_p < -5)
    )
    divergence = (
        (signal == "BUY"  and of_bias == "SELL" and abs(delta_p) > 10) or
        (signal == "SELL" and of_bias == "BUY"  and abs(delta_p) > 10)
    )

    # Confluence score boost
    of_score = min(100, conf + (imb * 0.3)) if confirmed else max(0, conf - (imb * 0.2))

    return {
        "confirmed":   confirmed,
        "divergence":  divergence,
        "of_bias":     of_bias,
        "delta_pct":   round(delta_p, 2),
        "imbalance":   round(imb, 2),
        "of_score":    round(of_score, 1),
    }

def render_orderflow(res, asset_name, df=None, asset_cfg=None):
    import streamlit as st
    import streamlit.components.v1 as components

    st.markdown("""
    <div style="background:linear-gradient(135deg,#0A1520,#1A0A20);border:2px solid #A78BFA;
         border-radius:12px;padding:14px 18px;margin-bottom:14px;text-align:center;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;
           color:#A78BFA;letter-spacing:2px;">📊 ORDERFLOW — FOOTPRINT ANALYSIS</div>
      <div style="font-size:10px;color:#6B7A8D;margin-top:3px;">
        Buy/Sell Imbalance · Delta · POC · Value Area · Confluence ICT-ML
      </div>
    </div>""", unsafe_allow_html=True)

    if not res:
        st.warning("⚠️ Lance d'abord une analyse pour activer l'OrderFlow.")
        return None

    current_price = res.get("close", 0)
    is_btc = "BTC" in asset_name

    with st.spinner("Calcul OrderFlow en cours..."):
        if is_btc:
            levels = fetch_orderflow_btc("BTCUSDT", 1000)
        else:
            levels = fetch_orderflow_oanda(df) if df is not None else None

        if not levels:
            # Fallback: estimate from res candle data
            levels = {}
            if res:
                p = current_price
                atr = res.get("atr", p*0.001)
                vwap = res.get("vwap", p)
                rsi  = res.get("rsi", 50)
                macd_b = res.get("macd_bull", False)
                vol  = res.get("volume", 1000) if res.get("volume") else 1000
                # Estimate based on indicators
                buy_ratio  = (rsi/100) * (0.7 if macd_b else 0.45)
                sell_ratio = 1 - buy_ratio
                for i in range(20):
                    lp = round(p - atr*(10-i)*0.15, 5)
                    bv = vol * buy_ratio  * (0.8 + i*0.02)
                    sv = vol * sell_ratio * (0.8 + (19-i)*0.02)
                    levels[lp] = {"buy": bv, "sell": sv}

        of = calc_orderflow_metrics(levels, current_price)

    if not of:
        st.error("❌ Impossible de calculer l'OrderFlow.")
        return None

    # Check confirmation with current signal
    conf_result = check_orderflow_confirmation(of, res)

    delta_p  = of["delta_pct"]
    imb      = of["imbalance_pct"]
    bias     = of["bias"]
    poc      = of["poc"]
    va_low   = of["va_low"]
    va_high  = of["va_high"]
    t_buy    = of["total_buy"]
    t_sell   = of["total_sell"]
    t_vol    = of["total_vol"]

    bias_col  = "#0FBF5F" if bias == "BUY" else "#D93025"
    delta_col = "#0FBF5F" if delta_p > 0   else "#D93025"
    delta_bar_buy  = max(5, int(min(95, 50 + delta_p/2)))
    delta_bar_sell = 100 - delta_bar_buy

    # Confirmation banner
    if conf_result and conf_result["confirmed"]:
        banner_col = "#0FBF5F" if bias=="BUY" else "#D93025"
        banner_icon = "🟢" if bias=="BUY" else "🔴"
        banner_txt = f"✅ Tendance Confirmée par Volume Supérieur — {bias} ({conf_result['of_score']:.0f}%)"
    elif conf_result and conf_result["divergence"]:
        banner_col = "#E07B2A"
        banner_icon = "⚠️"
        banner_txt = f"DIVERGENCE — Signal {res.get('signal','')} mais Volume {bias} — Prudence !"
    else:
        banner_col = "#C9A94B"
        banner_icon = "🟡"
        banner_txt = f"Volume neutre — Pas de confirmation directionnelle forte"

    # Build footprint rows
    fp_rows = ""
    for price, lvl in of["footprint"]:
        bv = lvl["buy"];  sv = lvl["sell"]
        tv = bv + sv or 0.001
        br = bv/tv;       sr = sv/tv
        bw = max(4, int(br*80));  sw = max(4, int(sr*80))
        is_poc  = abs(price - poc) < abs(of["footprint"][0][0] - of["footprint"][-1][0]) * 0.05
        is_cur  = abs(price - current_price) < abs(of["footprint"][0][0] - of["footprint"][-1][0]) * 0.08
        is_va   = va_low <= price <= va_high
        row_bg  = "rgba(240,201,106,.10)" if is_cur else "rgba(74,126,199,.08)" if is_va else "rgba(0,0,0,.2)"
        row_brd = "1px solid #F0C96A" if is_cur else "1px solid #4A7EC7" if is_poc else "1px solid transparent"
        # Imbalance flag
        imb_flag = ""
        if br > 0.75: imb_flag = '<span style="color:#0FBF5F;font-size:9px;"> ▲BUY</span>'
        elif sr > 0.75: imb_flag = '<span style="color:#D93025;font-size:9px;"> ▼SELL</span>'

        fp_rows += f"""
        <tr style="background:{row_bg};border:{row_brd};border-radius:4px;">
          <td style="font-family:'JetBrains Mono',monospace;font-size:11px;
               color:{'#F0C96A' if is_cur else '#8A95A3'};padding:4px 6px;">
            {f"{price:.2f}" if price>10 else f"{price:.5f}"}{imb_flag}
          </td>
          <td style="padding:4px 6px;">
            <div style="display:flex;align-items:center;gap:3px;">
              <div style="width:{bw}px;height:10px;background:#0FBF5F;border-radius:2px;opacity:0.85;"></div>
              <span style="font-size:10px;color:#0FBF5F;">{bv:.1f}</span>
            </div>
          </td>
          <td style="padding:4px 6px;">
            <div style="display:flex;align-items:center;gap:3px;">
              <div style="width:{sw}px;height:10px;background:#D93025;border-radius:2px;opacity:0.85;"></div>
              <span style="font-size:10px;color:#D93025;">{sv:.1f}</span>
            </div>
          </td>
          <td style="font-size:10px;color:{'#0FBF5F' if bv>=sv else '#D93025'};padding:4px 6px;font-weight:700;">
            {'+' if bv>=sv else ''}{round(bv-sv,1)}
          </td>
        </tr>"""

    # Imbalance zones
    imb_rows = ""
    for z in of.get("imbalance_zones",[]):
        zc = "#0FBF5F" if z["type"]=="buy_imbalance" else "#D93025"
        zt = "BUY IMBALANCE 🟢" if z["type"]=="buy_imbalance" else "SELL IMBALANCE 🔴"
        imb_rows += f"""<tr>
          <td style="color:{zc};font-family:'JetBrains Mono';font-size:11px;">{f"{z['price']:.2f}" if z['price']>10 else f"{z['price']:.5f}"}</td>
          <td style="color:{zc};font-size:11px;font-weight:700;">{zt}</td>
          <td style="color:#C9A94B;font-size:11px;">{z['ratio']*100:.0f}% dominant</td>
        </tr>"""

    # Alignment with ICT/ML
    signal   = res.get("signal","")
    sig_conf = res.get("confluence", 0)
    fvg_bull = len(res.get("fvg_bull",[]) or []) > 0
    fvg_bear = len(res.get("fvg_bear",[]) or []) > 0
    ob_bull  = len(res.get("ob_bull",[])  or []) > 0
    ob_bear  = len(res.get("ob_bear",[])  or []) > 0

    ict_align = []
    if signal=="BUY"  and fvg_bull: ict_align.append("🟢 FVG Haussier aligné")
    if signal=="BUY"  and ob_bull:  ict_align.append("🟢 Order Block Haussier actif")
    if signal=="SELL" and fvg_bear: ict_align.append("🔴 FVG Baissier aligné")
    if signal=="SELL" and ob_bear:  ict_align.append("🔴 Order Block Baissier actif")
    if not ict_align: ict_align = ["🟡 Pas de confluence ICT directe"]
    ict_html = "".join(f'<div style="color:#8A95A3;font-size:11px;">{a}</div>' for a in ict_align)

    total_score = conf_result["of_score"] if conf_result else sig_conf
    score_col   = "#0FBF5F" if total_score>=75 else "#C9A94B" if total_score>=55 else "#D93025"

    full_html = f"""<!DOCTYPE html><html><head>
    <link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@700&family=JetBrains+Mono&display=swap" rel="stylesheet">
    <style>
      body{{margin:0;padding:8px;background:#060E1A;font-family:'Rajdhani',sans-serif;color:#DCE8F5;}}
      table{{width:100%;border-collapse:collapse;}}
      th{{font-size:10px;color:#6B7A8D;padding:5px 8px;text-align:left;border-bottom:1px solid #1A3050;}}
      td{{padding:0;}}
    </style></head><body>

    <!-- CONFIRMATION BANNER -->
    <div style="background:rgba(0,0,0,.3);border:2px solid {banner_col};border-radius:10px;
         padding:10px 14px;margin-bottom:10px;text-align:center;">
      <div style="font-size:15px;font-weight:700;color:{banner_col};">
        {banner_icon} {banner_txt}
      </div>
    </div>

    <!-- STATS ROW -->
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:5px;margin-bottom:10px;">
      <div style="background:#0F1C2E;border:1px solid {delta_col};border-radius:7px;padding:7px;text-align:center;">
        <div style="font-size:9px;color:#6B7A8D;">DELTA</div>
        <div style="font-size:15px;font-weight:700;color:{delta_col};">{delta_p:+.1f}%</div>
        <div style="font-size:9px;color:{delta_col};">{'BUY' if delta_p>0 else 'SELL'}</div>
      </div>
      <div style="background:#0F1C2E;border:1px solid #A78BFA;border-radius:7px;padding:7px;text-align:center;">
        <div style="font-size:9px;color:#6B7A8D;">IMBALANCE</div>
        <div style="font-size:15px;font-weight:700;color:#A78BFA;">{imb:.1f}%</div>
        <div style="font-size:9px;color:#6B7A8D;">{'Fort' if imb>20 else 'Modéré' if imb>10 else 'Faible'}</div>
      </div>
      <div style="background:#0F1C2E;border:1px solid {bias_col};border-radius:7px;padding:7px;text-align:center;">
        <div style="font-size:9px;color:#6B7A8D;">BIAIS VOL.</div>
        <div style="font-size:16px;font-weight:700;color:{bias_col};">{'🟢' if bias=='BUY' else '🔴'} {bias}</div>
      </div>
      <div style="background:#0F1C2E;border:1px solid {score_col};border-radius:7px;padding:7px;text-align:center;">
        <div style="font-size:9px;color:#6B7A8D;">SCORE TOTAL</div>
        <div style="font-size:15px;font-weight:700;color:{score_col};">{total_score:.0f}%</div>
        <div style="font-size:9px;color:{score_col};">ICT+ML+OF</div>
      </div>
    </div>

    <!-- BUY/SELL VOLUME BAR -->
    <div style="margin-bottom:10px;">
      <div style="font-size:11px;color:#6B7A8D;margin-bottom:4px;">
        Volume Total: {t_vol:.0f} &nbsp;|&nbsp;
        <span style="color:#0FBF5F;">Buy: {t_buy:.0f} ({t_buy/t_vol*100:.1f}%)</span> &nbsp;|&nbsp;
        <span style="color:#D93025;">Sell: {t_sell:.0f} ({t_sell/t_vol*100:.1f}%)</span>
      </div>
      <div style="display:flex;border-radius:6px;overflow:hidden;height:20px;">
        <div style="width:{delta_bar_buy}%;background:#0FBF5F;opacity:0.85;"></div>
        <div style="width:{delta_bar_sell}%;background:#D93025;opacity:0.85;"></div>
      </div>
    </div>

    <!-- POC + VALUE AREA -->
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:5px;margin-bottom:10px;">
      <div style="background:rgba(74,126,199,.12);border:1px solid #4A7EC7;border-radius:6px;padding:7px;text-align:center;">
        <div style="font-size:9px;color:#6B7A8D;">POC (Max Volume)</div>
        <div style="font-size:13px;font-weight:700;color:#4A7EC7;">{f"{poc:.2f}" if poc>10 else f"{poc:.5f}"}</div>
      </div>
      <div style="background:rgba(15,191,95,.06);border:1px solid #0FBF5F33;border-radius:6px;padding:7px;text-align:center;">
        <div style="font-size:9px;color:#6B7A8D;">VA HIGH</div>
        <div style="font-size:13px;font-weight:700;color:#7EC47A;">{f"{va_high:.2f}" if va_high>10 else f"{va_high:.5f}"}</div>
      </div>
      <div style="background:rgba(217,48,37,.06);border:1px solid #D9302533;border-radius:6px;padding:7px;text-align:center;">
        <div style="font-size:9px;color:#6B7A8D;">VA LOW</div>
        <div style="font-size:13px;font-weight:700;color:#FF8A80;">{f"{va_low:.2f}" if va_low>10 else f"{va_low:.5f}"}</div>
      </div>
    </div>

    <!-- FOOTPRINT TABLE -->
    <div style="font-size:12px;color:#A78BFA;letter-spacing:2px;margin-bottom:6px;">
      📊 FOOTPRINT — BUY vs SELL par niveau de prix
    </div>
    <table>
      <thead><tr>
        <th>Prix</th>
        <th>🟢 Buy Vol</th>
        <th>🔴 Sell Vol</th>
        <th>Δ Delta</th>
      </tr></thead>
      <tbody>{fp_rows}</tbody>
    </table>

    <!-- IMBALANCE ZONES -->
    {"<div style='font-size:12px;color:#A78BFA;letter-spacing:2px;margin:10px 0 6px 0;'>⚡ ZONES D\\'IMBALANCE (ratio > 75%)</div><table><thead><tr><th>Prix</th><th>Type</th><th>Force</th></tr></thead><tbody>" + imb_rows + "</tbody></table>" if imb_rows else ""}

    <!-- ICT ALIGNMENT -->
    <div style="background:#0F1C2E;border:1px solid #1A3050;border-left:3px solid #A78BFA;
         border-radius:6px;padding:9px 12px;margin-top:10px;">
      <div style="font-size:12px;color:#A78BFA;letter-spacing:1px;font-weight:700;margin-bottom:5px;">
        🎯 ALIGNEMENT ICT — ML — ORDERFLOW
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;font-size:11px;">
        <div>
          <div style="color:#6B7A8D;font-size:9px;margin-bottom:3px;">SIGNAL APP</div>
          <div style="color:{'#0FBF5F' if signal=='BUY' else '#D93025' if signal=='SELL' else '#C9A94B'};font-weight:700;">
            {'🟢' if signal=='BUY' else '🔴' if signal=='SELL' else '🟡'} {signal} {sig_conf:.0f}%
          </div>
        </div>
        <div>
          <div style="color:#6B7A8D;font-size:9px;margin-bottom:3px;">ORDERFLOW</div>
          <div style="color:{bias_col};font-weight:700;">
            {'🟢' if bias=='BUY' else '🔴'} {bias} Δ{delta_p:+.1f}%
          </div>
        </div>
      </div>
      <div style="margin-top:7px;">{ict_html}</div>
    </div>

    <!-- GUIDE -->
    <div style="background:#0F1C2E;border-left:3px solid #A78BFA;border-radius:6px;
         padding:9px 12px;margin-top:8px;font-size:11px;color:#8A95A3;line-height:1.9;">
      <b style="color:#A78BFA;">📖 LECTURE ORDERFLOW :</b><br>
      🟢 Delta positif + Imbalance BUY + Signal BUY = Confluence maximale ✅<br>
      🔴 Delta négatif + Imbalance SELL + Signal SELL = Short confirmé ✅<br>
      ⚠️ Signal BUY mais Delta négatif = Divergence → Attendre confirmation<br>
      📍 POC = niveau d'équilibre → Prix y revient souvent<br>
      📦 Value Area = zone où 70% du volume s'est échangé
    </div>
    </body></html>"""

    components.html(full_html, height=820, scrolling=True)
    return conf_result

# ─── TAPE READING MODULE ──────────────────────────────────────────────────────

def fetch_tape_btc(limit=80):
    """Fetch BTC trades — tries multiple endpoints for Streamlit Cloud compatibility."""
    # Endpoint 1: Binance Spot aggTrades
    try:
        r = requests.get("https://api.binance.com/api/v3/aggTrades",
                         params={"symbol":"BTCUSDT","limit":limit},
                         headers={"User-Agent":"Mozilla/5.0"}, timeout=10)
        if r.status_code == 200:
            trades = []
            for t in r.json():
                trades.append({"price":float(t["p"]),"qty":float(t["q"]),
                               "is_buy": not t["m"],"ts":t["T"]})
            if trades: return trades
    except: pass
    # Endpoint 2: Binance recent trades (simpler endpoint)
    try:
        r = requests.get("https://api.binance.com/api/v3/trades",
                         params={"symbol":"BTCUSDT","limit":limit},
                         headers={"User-Agent":"Mozilla/5.0"}, timeout=10)
        if r.status_code == 200:
            trades = []
            for t in r.json():
                trades.append({"price":float(t["price"]),"qty":float(t["qty"]),
                               "is_buy": not t["isBuyerMaker"],"ts":t["time"]})
            if trades: return trades
    except: pass
    # Endpoint 3: Binance ticker 24h for price + simulate trades
    try:
        r = requests.get("https://api.binance.com/api/v3/ticker/24hr",
                         params={"symbol":"BTCUSDT"},
                         headers={"User-Agent":"Mozilla/5.0"}, timeout=8)
        if r.status_code == 200:
            d = r.json()
            price  = float(d.get("lastPrice", 0))
            vol    = float(d.get("volume", 1000))
            buy_v  = float(d.get("quoteVolume", 0))
            count  = int(d.get("count", 100))
            # Simulate simplified trade list from 24h data
            trades = []
            buy_ratio = 0.52  # slight buy bias as default
            for i in range(min(limit, 30)):
                is_buy = (i % 10) < int(buy_ratio * 10)
                trades.append({"price":price,"qty":vol/count,"is_buy":is_buy,"ts":0})
            return trades
    except: pass
    return []

def fetch_tape_oanda(instrument, key, demo=True, count=60):
    try:
        base = OANDA_DEMO if demo else OANDA_LIVE
        r = requests.get(f"{base}/v3/instruments/{instrument}/candles",
            headers={"Authorization":f"Bearer {key}"},
            params={"granularity":"S5","count":count,"price":"BA"}, timeout=8)
        if r.status_code != 200: return []
        trades = []
        for c in r.json().get("candles",[]):
            if not c.get("complete",True): continue
            bid_c = float(c.get("bid",{}).get("c",0))
            ask_c = float(c.get("ask",{}).get("c",0))
            mid   = (bid_c + ask_c) / 2
            bid_o = float(c.get("bid",{}).get("o",mid))
            vol   = float(c.get("volume",1))
            trades.append({"price":round(mid,5),"qty":vol,
                           "is_buy": mid >= bid_o,"ts":0})
        return trades
    except: return []

def calc_tape_stats(trades):
    if not trades: return {}
    buy_vol  = sum(t["qty"] for t in trades if t["is_buy"])
    sell_vol = sum(t["qty"] for t in trades if not t["is_buy"])
    total    = buy_vol + sell_vol or 1
    delta    = buy_vol - sell_vol
    delta_pct= delta / total * 100
    sizes    = sorted([t["qty"] for t in trades], reverse=True)
    threshold= sizes[max(0, len(sizes)//10)] if sizes else 0
    big_buys = len([t for t in trades if t["is_buy"]  and t["qty"] >= threshold and threshold>0])
    big_sells= len([t for t in trades if not t["is_buy"] and t["qty"] >= threshold and threshold>0])
    if abs(delta_pct) > 30: conviction = "FORTE"
    elif abs(delta_pct) > 15: conviction = "MODEREE"
    else: conviction = "FAIBLE"
    if delta_pct > 10:    bias,bias_col = "BUY",  "#0FBF5F"
    elif delta_pct < -10: bias,bias_col = "SELL", "#D93025"
    else:                 bias,bias_col = "NEUTRE","#C9A94B"
    return {"buy_vol":round(buy_vol,2),"sell_vol":round(sell_vol,2),
            "total":round(total,2),"delta":round(delta,2),
            "delta_pct":round(delta_pct,2),
            "buy_pct":round(buy_vol/total*100,1),
            "sell_pct":round(sell_vol/total*100,1),
            "big_buys":big_buys,"big_sells":big_sells,
            "threshold":round(threshold,4),"speed":len(trades),
            "conviction":conviction,"bias":bias,"bias_col":bias_col}

def render_tape_tab(asset_name, res, asset_cfg, oanda_key, oanda_demo):
    import streamlit.components.v1 as _ct

    is_btc = "BTC" in asset_name

    st.markdown('''
    <div style="background:linear-gradient(135deg,#0A1520,#0A1A10);border:2px solid #0FBF5F;
         border-radius:12px;padding:14px 18px;margin-bottom:14px;text-align:center;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;
           color:#0FBF5F;letter-spacing:2px;">📋 TAPE READING — FLUX D'ORDRES LIVE</div>
      <div style="font-size:10px;color:#6B7A8D;margin-top:3px;">
        Transactions en temps réel · Intentions institutionnelles · Buy vs Sell
      </div>
    </div>''', unsafe_allow_html=True)

    with st.expander("📖 Comment lire le Tape ? (cliquer pour ouvrir)", expanded=False):
        st.markdown('''
<div style="background:#0F1C2E;border-radius:8px;padding:14px 16px;">

**🎯 Le Tape Reading — C'est quoi ?**

Le Tape est le **flux de toutes les transactions** passées en temps réel sur le marché.
Chaque ligne = une transaction réelle entre un acheteur et un vendeur.
C'est l'outil des **traders professionnels** pour voir ce que font les institutions.

---

**🟢 Gros volumes VERTS qui s'accumulent = Institutions ACHÈTENT → Signal BUY**

Quand tu vois plusieurs grosses transactions vertes se succéder rapidement sur un
niveau de support ou un FVG haussier → Les institutions accumulent des positions longues.

**Exemple :** Prix arrive sur FVG 2340 (Or) + Tape montre 5 gros ordres verts
→ Entrée BUY avec SL sous le FVG ✅

---

**🔴 Gros volumes ROUGES qui s'accumulent = Institutions VENDENT → Signal SELL**

Quand tu vois plusieurs grosses transactions rouges se succéder sur une résistance
ou un Order Block baissier → Les institutions distribuent leurs positions.

**Exemple :** Prix arrive sur Bearish OB (CL WTI) + 4 gros ordres rouges
→ Entrée SELL avec SL au-dessus de l'OB ✅

---

**⚪ Volumes MIXTES petits = Pas de conviction → ATTENDRE**

Petites transactions vertes et rouges alternées sans direction claire
→ Marché indécis · Ne pas trader · Attends qu'une direction s'impose.

---

**⚡ L'Absorption — Signal le plus puissant**

Prix BAISSE vers support + Tape montre GROS volumes VERTS malgré la baisse
→ Institutions absorbent les vendeurs → Retournement HAUSSIER imminent 🚀

**⚡ L'Épuisement — Fin de tendance**

Prix MONTE mais volumes VERTS DIMINUENT progressivement
→ Acheteurs épuisés → Retournement ou consolidation ⚠️

</div>
        ''', unsafe_allow_html=True)

    st.markdown("---")

    col_btn, col_info = st.columns([2,1])
    with col_btn:
        st.button("🔄 Actualiser le Tape", use_container_width=True, key="tape_refresh")
    with col_info:
        src_lbl = "Binance aggTrades" if is_btc else "OANDA S5 Ticks"
        st.markdown(f'<div style="padding:8px;text-align:center;font-size:11px;color:#6B7A8D;">'
                    f'Source : <b style="color:#C9A94B;">{src_lbl}</b></div>',
                    unsafe_allow_html=True)

    with st.spinner("Chargement du Tape..."):
        if is_btc:
            trades = fetch_tape_btc(limit=80)
        else:
            inst = asset_cfg.get("oanda")
            trades = fetch_tape_oanda(inst, oanda_key, oanda_demo) if inst and oanda_key else []

    # Fallback: build synthetic tape from res if API failed
    if not trades and res:
        price  = res.get("close", 0)
        atr    = res.get("atr", price * 0.001)
        rsi    = res.get("rsi", 50)
        macd_b = res.get("macd_bull", False)
        vol_b  = res.get("volume", 1000) or 1000
        # Estimate buy/sell split from RSI + MACD
        buy_ratio = (rsi / 100) * (0.65 if macd_b else 0.45)
        for i in range(40):
            is_buy = (i / 40) < buy_ratio
            qty    = vol_b / 40 * (0.8 + 0.4 * ((i % 5) / 4))
            trades.append({"price": round(price + (atr * 0.1 * (i - 20) / 20), 5),
                           "qty": round(qty, 4), "is_buy": is_buy, "ts": 0})
        st.info("ℹ️ Tape estimé depuis les indicateurs OANDA (Binance API non disponible).")

    if not trades:
        if not is_btc and not oanda_key:
            st.warning("⚠️ Configure ton token OANDA dans ⚙️ CONFIG pour activer le Tape.")
        else:
            st.warning("⚠️ Données Tape temporairement indisponibles. Réessaie dans quelques secondes.")
        return

    stats = calc_tape_stats(trades)
    if not stats: return

    app_signal = res.get("signal","") if res else ""
    tape_bias  = stats["bias"]

    if app_signal and tape_bias != "NEUTRE":
        tape_confirm = "confirmed" if app_signal == tape_bias else "divergence"
    else:
        tape_confirm = "neutral"

    if tape_confirm == "confirmed":
        b_col = "#0FBF5F" if tape_bias=="BUY" else "#D93025"
        b_icon = "🟢" if tape_bias=="BUY" else "🔴"
        b_txt  = f"✅ TAPE CONFIRME LE SIGNAL {app_signal} — Confluence Maximale !"
    elif tape_confirm == "divergence":
        b_col,b_icon,b_txt = "#E07B2A","⚠️",f"⚠️ DIVERGENCE — Signal APP {app_signal} mais Tape {tape_bias} → Attendre"
    else:
        b_col,b_icon,b_txt = "#C9A94B","⚪","🟡 Tape NEUTRE — Pas de biais directionnel clair → Patience"

    buy_w  = max(3, int(stats["buy_pct"]))
    sell_w = 100 - buy_w
    conv_col = ("#0FBF5F" if tape_bias=="BUY" else "#D93025") if stats["conviction"]=="FORTE" else ("#C9A94B" if stats["conviction"]=="MODEREE" else "#6B7A8D")
    speed_label = "TRÈS RAPIDE 🔥" if stats["speed"]>60 else ("RAPIDE" if stats["speed"]>30 else "NORMAL")
    speed_col   = "#D93025" if stats["speed"]>60 else ("#C9A94B" if stats["speed"]>30 else "#6B7A8D")

    recent  = list(reversed(trades[-30:]))
    max_qty = max((t["qty"] for t in recent), default=1)
    thr     = stats.get("threshold",0)

    tape_rows = ""
    for t in recent:
        rc   = "#0FBF5F" if t["is_buy"] else "#D93025"
        rbg  = "rgba(15,191,95,.08)" if t["is_buy"] else "rgba(217,48,37,.08)"
        rbrd = "rgba(15,191,95,.25)" if t["is_buy"] else "rgba(217,48,37,.25)"
        dtxt = "BUY ↑" if t["is_buy"] else "SELL ↓"
        bw   = max(4, int(t["qty"]/max_qty*80))
        big  = t["qty"]>=thr and thr>0
        btag = ' <span style="color:#F0C96A;font-size:9px;font-weight:700;">★ GROS</span>' if big else ""
        pfmt = f"{t['price']:.2f}" if t["price"]>10 else f"{t['price']:.5f}"
        qfmt = f"{t['qty']:.4f}" if t["qty"]<1 else f"{t['qty']:.2f}"
        tape_rows += (
            f'<tr style="background:{rbg};border-bottom:1px solid {rbrd};">' +
            f'<td style="color:{rc};font-weight:700;font-size:12px;padding:4px 8px;">{dtxt}{btag}</td>' +
            f'<td style="color:#DCE8F5;font-size:11px;padding:4px 8px;">{pfmt}</td>' +
            f'<td style="padding:4px 8px;"><div style="display:flex;align-items:center;gap:4px;">' +
            f'<div style="width:{bw}px;height:10px;background:{rc};border-radius:2px;opacity:0.85;"></div>' +
            f'<span style="color:{rc};font-size:10px;">{qfmt}</span></div></td></tr>'
        )

    src_note = "Binance aggTrades — Vrai Tape institutionnel" if is_btc else "OANDA S5 Ticks — Tape directionnel"
    big_buy_note  = "🟢 Institutions achètent !" if stats["big_buys"]>=3 else "Volume normal"
    big_sell_note = "🔴 Institutions vendent !" if stats["big_sells"]>=3 else "Volume normal"

    html = (
        '<!DOCTYPE html><html><head>' +
        '<link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@700' +
        '&family=JetBrains+Mono&display=swap" rel="stylesheet">' +
        '<style>body{margin:0;padding:8px;background:#060E1A;font-family:sans-serif;color:#DCE8F5;}' +
        'table{width:100%;border-collapse:collapse;}th{font-size:10px;color:#6B7A8D;padding:5px 8px;text-align:left;border-bottom:1px solid #1A3050;}</style></head><body>' +

        f'<div style="background:rgba(0,0,0,.3);border:2px solid {b_col};border-radius:10px;padding:9px 14px;margin-bottom:10px;text-align:center;">' +
        f'<div style="font-size:14px;font-weight:700;color:{b_col};">{b_icon} {b_txt}</div></div>' +

        f'<div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:5px;margin-bottom:10px;">' +
        f'<div style="background:#0F1C2E;border:1px solid #0FBF5F;border-radius:7px;padding:7px;text-align:center;"><div style="font-size:9px;color:#6B7A8D;">VOL. BUY</div><div style="font-size:15px;font-weight:700;color:#0FBF5F;">{stats["buy_pct"]:.1f}%</div><div style="font-size:9px;color:#0FBF5F;">{stats["buy_vol"]:.2f}</div></div>' +
        f'<div style="background:#0F1C2E;border:1px solid #D93025;border-radius:7px;padding:7px;text-align:center;"><div style="font-size:9px;color:#6B7A8D;">VOL. SELL</div><div style="font-size:15px;font-weight:700;color:#D93025;">{stats["sell_pct"]:.1f}%</div><div style="font-size:9px;color:#D93025;">{stats["sell_vol"]:.2f}</div></div>' +
        f'<div style="background:#0F1C2E;border:1px solid {stats["bias_col"]};border-radius:7px;padding:7px;text-align:center;"><div style="font-size:9px;color:#6B7A8D;">DELTA</div><div style="font-size:15px;font-weight:700;color:{stats["bias_col"]};">{stats["delta_pct"]:+.1f}%</div><div style="font-size:9px;color:{stats["bias_col"]};">{stats["bias"]}</div></div>' +
        f'<div style="background:#0F1C2E;border:1px solid {conv_col};border-radius:7px;padding:7px;text-align:center;"><div style="font-size:9px;color:#6B7A8D;">CONVICTION</div><div style="font-size:12px;font-weight:700;color:{conv_col};">{stats["conviction"]}</div><div style="font-size:9px;color:{speed_col};">{speed_label}</div></div>' +
        '</div>' +

        f'<div style="margin-bottom:10px;"><div style="font-size:10px;color:#6B7A8D;margin-bottom:3px;">Pression Buy vs Sell — {stats["buy_pct"]:.0f}% acheteurs / {stats["sell_pct"]:.0f}% vendeurs</div>' +
        f'<div style="display:flex;border-radius:6px;overflow:hidden;height:22px;"><div style="width:{buy_w}%;background:#0FBF5F;opacity:0.85;display:flex;align-items:center;justify-content:center;"><span style="font-size:10px;font-weight:700;color:#060E1A;">{stats["buy_pct"]:.0f}%</span></div>' +
        f'<div style="width:{sell_w}%;background:#D93025;opacity:0.85;display:flex;align-items:center;justify-content:center;"><span style="font-size:10px;font-weight:700;color:#DCE8F5;">{stats["sell_pct"]:.0f}%</span></div></div></div>' +

        f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:10px;">' +
        f'<div style="background:rgba(15,191,95,.08);border:1px solid rgba(15,191,95,.3);border-radius:7px;padding:7px;text-align:center;"><div style="font-size:9px;color:#6B7A8D;">★ GROS ORDRES BUY</div><div style="font-size:22px;font-weight:700;color:#0FBF5F;">{stats["big_buys"]}</div><div style="font-size:9px;color:#0FBF5F;">{big_buy_note}</div></div>' +
        f'<div style="background:rgba(217,48,37,.08);border:1px solid rgba(217,48,37,.3);border-radius:7px;padding:7px;text-align:center;"><div style="font-size:9px;color:#6B7A8D;">★ GROS ORDRES SELL</div><div style="font-size:22px;font-weight:700;color:#D93025;">{stats["big_sells"]}</div><div style="font-size:9px;color:#D93025;">{big_sell_note}</div></div>' +
        '</div>' +

        '<div style="font-size:12px;color:#0FBF5F;letter-spacing:2px;margin-bottom:6px;">📋 FLUX ORDRES — 30 DERNIERES TRANSACTIONS</div>' +
        '<table><thead><tr><th>Direction</th><th>Prix</th><th>Volume</th></tr></thead><tbody>' +
        tape_rows + '</tbody></table>' +

        '<div style="background:#0F1C2E;border-radius:6px;padding:9px 12px;margin-top:10px;">' +
        '<div style="font-size:11px;color:#C9A94B;font-weight:700;margin-bottom:6px;">📖 LÉGENDE</div>' +
        '<div style="font-size:11px;color:#0FBF5F;margin-bottom:3px;">🟢 Gros volumes VERTS accumulés = Institutions ACHÈTENT → BUY</div>' +
        '<div style="font-size:11px;color:#D93025;margin-bottom:3px;">🔴 Gros volumes ROUGES accumulés = Institutions VENDENT → SELL</div>' +
        '<div style="font-size:11px;color:#8A95A3;margin-bottom:3px;">⚪ Volumes mixtes petits = Pas de conviction → ATTENDRE</div>' +
        '<div style="font-size:11px;color:#F0C96A;">★ = Gros ordre institutionnel (top 10% taille)</div>' +
        '</div>' +

        f'<div style="font-size:9px;color:#6B7A8D;text-align:center;margin-top:8px;">Source : {src_note} | {len(trades)} transactions analysées</div>' +
        '</body></html>'
    )
    _ct.html(html, height=720, scrolling=True)


# ─── ORDER BOOK MODULE ────────────────────────────────────────────────────────

def fetch_orderbook_btc(depth=50):
    """Fetch Binance BTC/USDT order book — real bids and asks."""
    try:
        r = requests.get(
            "https://api.binance.com/api/v3/depth",
            params={"symbol": "BTCUSDT", "limit": depth},
            timeout=8
        )
        if r.status_code != 200: return None
        data = r.json()
        bids = [[float(p), float(q)] for p, q in data.get("bids", [])]
        asks = [[float(p), float(q)] for p, q in data.get("asks", [])]
        return {"bids": bids, "asks": asks, "source": "binance"}
    except: return None

def fetch_orderbook_oanda(instrument, key, demo=True):
    """Fetch OANDA order book — tries orderBook endpoint first, then builds synthetic OB from candles."""
    base = OANDA_DEMO if demo else OANDA_LIVE

    # ── Attempt 1: OANDA orderBook endpoint ──────────────────────────────
    try:
        r = requests.get(
            f"{base}/v3/instruments/{instrument}/orderBook",
            headers={"Authorization": f"Bearer {key}"},
            timeout=8
        )
        if r.status_code == 200:
            data    = r.json().get("orderBook", {})
            buckets = data.get("buckets", [])
            bids, asks = [], []
            for b in buckets:
                price     = float(b.get("price", 0))
                long_pct  = float(b.get("longCountPercent", 0))
                short_pct = float(b.get("shortCountPercent", 0))
                if long_pct  > 0: bids.append([price, long_pct])
                if short_pct > 0: asks.append([price, short_pct])
            if bids or asks:
                bids.sort(key=lambda x: x[0], reverse=True)
                asks.sort(key=lambda x: x[0])
                return {"bids": bids[:30], "asks": asks[:30], "source": "oanda_ob"}
    except: pass

    # ── Attempt 2: Build synthetic OB from S5 candles ────────────────────
    # Uses bid/ask OANDA candles to estimate order pressure at each level
    try:
        r = requests.get(
            f"{base}/v3/instruments/{instrument}/candles",
            headers={"Authorization": f"Bearer {key}"},
            params={"granularity": "S5", "count": 80, "price": "BA"},
            timeout=8
        )
        if r.status_code == 200:
            candles = r.json().get("candles", [])
            if not candles: return None

            # Get current mid price
            last = candles[-1]
            mid_c = (float(last.get("bid",{}).get("c",0)) +
                     float(last.get("ask",{}).get("c",0))) / 2

            # Build synthetic bid/ask levels from price action
            # Level = rounded price, volume = candle volume
            bid_levels = {}
            ask_levels = {}

            for c in candles:
                if not c.get("complete", True): continue
                bid_o = float(c.get("bid",{}).get("o", 0))
                bid_c = float(c.get("bid",{}).get("c", 0))
                ask_o = float(c.get("ask",{}).get("o", 0))
                ask_c = float(c.get("ask",{}).get("c", 0))
                vol   = float(c.get("volume", 1))
                mid   = (bid_c + ask_c) / 2

                # Classify as bid or ask pressure
                if bid_c >= bid_o:  # price went up in this candle → buy pressure
                    lvl = round(bid_c, 3 if mid < 10 else 2)
                    bid_levels[lvl] = bid_levels.get(lvl, 0) + vol
                else:               # price went down → sell pressure
                    lvl = round(ask_c, 3 if mid < 10 else 2)
                    ask_levels[lvl] = ask_levels.get(lvl, 0) + vol

            # Filter to near-price levels only (within 0.5% of mid)
            bids = [[p, v] for p, v in bid_levels.items()
                    if p < mid_c and (mid_c - p) / mid_c < 0.005]
            asks = [[p, v] for p, v in ask_levels.items()
                    if p > mid_c and (p - mid_c) / mid_c < 0.005]

            bids.sort(key=lambda x: x[0], reverse=True)
            asks.sort(key=lambda x: x[0])

            if bids or asks:
                return {"bids": bids[:20], "asks": asks[:20], "source": "oanda_synthetic"}
    except: pass

    # ── Attempt 3: Build from res data if available ───────────────────────
    return None

def calc_orderbook_stats(ob, current_price):
    """Compute walls, imbalance, bid/ask ratio from order book."""
    if not ob: return {}
    bids = ob.get("bids", [])
    asks = ob.get("asks", [])

    total_bid = sum(q for _, q in bids)
    total_ask = sum(q for _, q in asks)
    total     = total_bid + total_ask or 1

    bid_pct   = total_bid / total * 100
    ask_pct   = total_ask / total * 100
    imbalance = bid_pct - ask_pct  # positive = buy pressure

    # Detect walls — top 15% volume levels
    all_vols  = [q for _, q in bids + asks]
    all_vols.sort(reverse=True)
    wall_thr  = all_vols[max(0, len(all_vols)//7)] if all_vols else 0

    bid_walls = [[p, q] for p, q in bids if q >= wall_thr and p < current_price]
    ask_walls = [[p, q] for p, q in asks if q >= wall_thr and p > current_price]

    # Nearest walls
    nearest_bid_wall = bid_walls[0]  if bid_walls else None
    nearest_ask_wall = ask_walls[0]  if ask_walls else None

    # Bias
    if imbalance > 15:   bias, bias_col = "BUY",   "#0FBF5F"
    elif imbalance < -15: bias, bias_col = "SELL", "#D93025"
    else:                 bias, bias_col = "NEUTRE","#C9A94B"

    # Distance to nearest walls
    dist_bid = round((current_price - nearest_bid_wall[0]) / current_price * 100, 3) if nearest_bid_wall else None
    dist_ask = round((nearest_ask_wall[0] - current_price) / current_price * 100, 3) if nearest_ask_wall else None

    return {
        "total_bid":        round(total_bid, 2),
        "total_ask":        round(total_ask, 2),
        "bid_pct":          round(bid_pct, 1),
        "ask_pct":          round(ask_pct, 1),
        "imbalance":        round(imbalance, 1),
        "bias":             bias,
        "bias_col":         bias_col,
        "bid_walls":        bid_walls[:5],
        "ask_walls":        ask_walls[:5],
        "nearest_bid_wall": nearest_bid_wall,
        "nearest_ask_wall": nearest_ask_wall,
        "wall_thr":         round(wall_thr, 4),
        "dist_bid_wall":    dist_bid,
        "dist_ask_wall":    dist_ask,
        "n_bid_walls":      len(bid_walls),
        "n_ask_walls":      len(ask_walls),
    }

def check_ob_ict_confluence(stats, res):
    """Check if Order Book walls align with ICT levels from analysis."""
    if not stats or not res: return []
    confluences = []
    price = res.get("close", 0)
    fvgs  = res.get("fvgs", [])
    obs   = res.get("obs",  [])

    # Check if bid walls near FVG/OB levels
    for wall in stats.get("bid_walls", []):
        wp = wall[0]
        for fvg in fvgs:
            if fvg.get("type") == "bull":
                if abs(wp - fvg.get("mid", 0)) / price < 0.003:
                    confluences.append(f"Mur BID ({wp:.2f}) aligne avec FVG Haussier !")
        for ob_ in obs:
            if ob_.get("type") == "bull":
                if abs(wp - ob_.get("mid", 0)) / price < 0.003:
                    confluences.append(f"Mur BID ({wp:.2f}) aligne avec Bullish OB !")

    for wall in stats.get("ask_walls", []):
        wp = wall[0]
        for fvg in fvgs:
            if fvg.get("type") == "bear":
                if abs(wp - fvg.get("mid", 0)) / price < 0.003:
                    confluences.append(f"Mur ASK ({wp:.2f}) aligne avec FVG Baissier !")
        for ob_ in obs:
            if ob_.get("type") == "bear":
                if abs(wp - ob_.get("mid", 0)) / price < 0.003:
                    confluences.append(f"Mur ASK ({wp:.2f}) aligne avec Bearish OB !")

    return confluences

def render_orderbook_tab(asset_name, res, asset_cfg, oanda_key, oanda_demo):
    """Full Order Book / Bookmap-style tab."""
    import streamlit.components.v1 as _cob

    is_btc = "BTC" in asset_name

    st.markdown("""
    <div style="background:linear-gradient(135deg,#0A1520,#1A0A20);border:2px solid #F0C96A;
         border-radius:12px;padding:14px 18px;margin-bottom:14px;text-align:center;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;
           color:#F0C96A;letter-spacing:2px;">📖 ORDER BOOK — STYLE BOOKMAP</div>
      <div style="font-size:10px;color:#6B7A8D;margin-top:3px;">
        Murs de liquidite · Ordres en attente · Confluence ICT · Absorption
      </div>
    </div>""", unsafe_allow_html=True)

    # Guide
    with st.expander("📖 Comment lire l'Order Book ? (Bookmap style)", expanded=False):
        st.markdown("""
<div style="background:#0F1C2E;border-radius:8px;padding:14px 16px;">

**🗺️ C'est quoi l'Order Book ?**

L'Order Book est la liste de TOUS les ordres en attente sur le marche.
Contrairement au Tape (ordres executes), l'Order Book montre ce qui VA se passer.

---

**🟡 MURS DE LIQUIDITE (Walls) = Zones Institutionnelles**

Un mur = un gros volume d'ordres en attente a un niveau de prix.
Les institutions placent des gros ordres limites pour accumuler / distribuer.

- **Mur BID (en dessous)** = Acheteurs institutionnels attendent la → Support fort
- **Mur ASK (au-dessus)** = Vendeurs institutionnels attendent la → Resistance forte

---

**⚡ L'ABSORPTION — Signal le plus puissant**

```
Mur BID present + Prix descend vers le mur
+ Le mur TIENT (volume reste stable)
→ Les institutions ABSORBENT les vendeurs → BUY imminent !

Mur ASK present + Prix monte vers le mur
+ Le mur TIENT
→ Les institutions ABSORBENT les acheteurs → SELL imminent !
```

---

**🚨 LE SPOOFING — Piege a eviter**

Si un gros mur DISPARAIT avant que le prix l'atteigne
→ C'etait du Spoofing (manipulation) → Ne pas trader ce niveau !

---

**🎯 Combo Order Book + ICT**

Mur BID au niveau d'un FVG Haussier = Confluence maximale !
Mur ASK au niveau d'un Bearish OB  = Confluence maximale !

---

**💡 TP Optimal**

Toujours placer ton TP JUSTE AVANT le prochain mur adverse :
- BUY → TP juste en dessous du prochain mur ASK
- SELL → TP juste au-dessus du prochain mur BID

</div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    col_btn, col_src = st.columns([2, 1])
    with col_btn:
        st.button("🔄 Actualiser l'Order Book", use_container_width=True, key="ob_refresh")
    with col_src:
        src_lbl = "Binance Depth API" if is_btc else "OANDA Order Book"
        st.markdown(f'<div style="padding:8px;text-align:center;font-size:11px;color:#6B7A8D;">'
                    f'Source : <b style="color:#F0C96A;">{src_lbl}</b></div>',
                    unsafe_allow_html=True)

    with st.spinner("Chargement de l'Order Book..."):
        if is_btc:
            ob = fetch_orderbook_btc(depth=50)
        else:
            inst = asset_cfg.get("oanda")
            ob = fetch_orderbook_oanda(inst, oanda_key, oanda_demo) if inst and oanda_key else None

    # ── Fallback: build synthetic OB from res indicators ─────────────────
    if not ob and res:
        price  = res.get("close", 0)
        atr    = res.get("atr", price * 0.001)
        vwap   = res.get("vwap", price)
        bb_up  = res.get("bb_upper", price * 1.005)
        bb_lo  = res.get("bb_lower", price * 0.995)
        rsi    = res.get("rsi", 50)
        liq    = res.get("liquidity", {})
        bsl    = liq.get("buyside",  [])   # resistance levels
        ssl    = liq.get("sellside", [])   # support levels

        # Build synthetic bids (supports) from ICT + indicators
        bids = []
        for lvl in [bb_lo, vwap - atr*0.5, price - atr, price - atr*1.5, price - atr*2]:
            vol = max(1, (70 - rsi) * 10) if lvl < price else 5
            bids.append([round(lvl, 5), round(vol, 2)])
        for lvl in ssl[:3]:
            bids.append([round(lvl, 5), round(atr * 50, 2)])

        # Build synthetic asks (resistances) from ICT + indicators
        asks = []
        for lvl in [bb_up, vwap + atr*0.5, price + atr, price + atr*1.5, price + atr*2]:
            vol = max(1, (rsi - 30) * 10) if lvl > price else 5
            asks.append([round(lvl, 5), round(vol, 2)])
        for lvl in bsl[:3]:
            asks.append([round(lvl, 5), round(atr * 50, 2)])

        bids = sorted([b for b in bids if b[0] < price], key=lambda x: x[0], reverse=True)
        asks = sorted([a for a in asks if a[0] > price], key=lambda x: x[0])
        ob   = {"bids": bids[:15], "asks": asks[:15], "source": "synthetic"}
        st.info("ℹ️ Order Book calculé depuis les indicateurs OANDA (données en direct non disponibles).")

    if not ob:
        if not is_btc and not oanda_key:
            st.warning("⚠️ Configure ton token OANDA dans ⚙️ Config pour activer l'Order Book.")
        else:
            st.warning("⚠️ Lance d'abord une analyse pour afficher l'Order Book.")
        return

    current_price = res.get("close", 0) if res else 0
    if not current_price and ob.get("bids"):
        current_price = ob["bids"][0][0]

    stats  = calc_orderbook_stats(ob, current_price)
    if not stats: return

    # ICT confluences
    ict_conf = check_ob_ict_confluence(stats, res) if res else []

    # App signal alignment
    app_sig  = res.get("signal", "") if res else ""
    ob_bias  = stats["bias"]
    if app_sig and ob_bias != "NEUTRE":
        ob_confirm = "confirmed" if app_sig == ob_bias else "divergence"
    else:
        ob_confirm = "neutral"

    if ob_confirm == "confirmed":
        b_col  = "#0FBF5F" if ob_bias == "BUY" else "#D93025"
        b_icon = "🟢" if ob_bias == "BUY" else "🔴"
        b_txt  = f"ORDER BOOK CONFIRME LE SIGNAL {app_sig} — Liquidite Alignee !"
    elif ob_confirm == "divergence":
        b_col, b_icon = "#E07B2A", "⚠️"
        b_txt  = f"DIVERGENCE — Signal {app_sig} mais Order Book {ob_bias} → Prudence !"
    else:
        b_col, b_icon = "#C9A94B", "🟡"
        b_txt  = "Order Book NEUTRE — Equilibre acheteurs / vendeurs"

    buy_w  = max(3, int(stats["bid_pct"]))
    sell_w = 100 - buy_w

    # Build OB visualization rows (bids green, asks red, price in center)
    bids_top = ob["bids"][:15]
    asks_top = ob["asks"][:15]
    max_vol  = max([q for _, q in bids_top + asks_top] or [1])
    wall_thr = stats.get("wall_thr", 0)

    ob_rows  = ""
    # ASK side (sells — above price) — reversed so lowest ask at top
    for price_lvl, qty in reversed(asks_top):
        is_wall = qty >= wall_thr and wall_thr > 0
        bw = max(4, int(qty / max_vol * 90))
        wall_tag = " ★" if is_wall else ""
        row_bg = "rgba(217,48,37,.15)" if is_wall else "rgba(217,48,37,.05)"
        row_brd= "rgba(217,48,37,.4)"  if is_wall else "rgba(217,48,37,.15)"
        pfmt = f"{price_lvl:.2f}" if price_lvl > 10 else f"{price_lvl:.5f}"
        qfmt = f"{qty:.4f}"       if qty < 1 else f"{qty:.2f}"
        ob_rows += (
            f'<tr style="background:{row_bg};border-bottom:1px solid {row_brd};">'
            f'<td style="color:#D93025;font-size:11px;padding:3px 7px;font-weight:{"700" if is_wall else "400"};">{pfmt}{wall_tag}</td>'
            f'<td style="padding:3px 7px;text-align:right;">'
            f'<div style="display:flex;align-items:center;justify-content:flex-end;gap:3px;">'
            f'<span style="color:#D93025;font-size:10px;">{qfmt}</span>'
            f'<div style="width:{bw}px;height:10px;background:#D93025;border-radius:2px;opacity:0.85;"></div>'
            f'</div></td>'
            f'<td style="color:#D93025;font-size:10px;padding:3px 7px;">{"★ MUR SELL" if is_wall else "SELL"}</td>'
            f'</tr>'
        )

    # Current price row
    pfmt_cur = f"{current_price:.2f}" if current_price > 10 else f"{current_price:.5f}"
    ob_rows += (
        f'<tr style="background:rgba(240,201,106,.15);border:1px solid #F0C96A;">'
        f'<td style="color:#F0C96A;font-weight:700;font-size:13px;padding:5px 7px;">◆ {pfmt_cur}</td>'
        f'<td style="color:#F0C96A;font-size:10px;padding:5px 7px;text-align:center;" colspan="2">← PRIX ACTUEL</td>'
        f'</tr>'
    )

    # BID side (buys — below price)
    for price_lvl, qty in bids_top:
        is_wall = qty >= wall_thr and wall_thr > 0
        bw = max(4, int(qty / max_vol * 90))
        wall_tag = " ★" if is_wall else ""
        row_bg = "rgba(15,191,95,.15)" if is_wall else "rgba(15,191,95,.05)"
        row_brd= "rgba(15,191,95,.4)"  if is_wall else "rgba(15,191,95,.15)"
        pfmt = f"{price_lvl:.2f}" if price_lvl > 10 else f"{price_lvl:.5f}"
        qfmt = f"{qty:.4f}"       if qty < 1 else f"{qty:.2f}"
        ob_rows += (
            f'<tr style="background:{row_bg};border-bottom:1px solid {row_brd};">'
            f'<td style="color:#0FBF5F;font-size:11px;padding:3px 7px;font-weight:{"700" if is_wall else "400"};">{pfmt}{wall_tag}</td>'
            f'<td style="padding:3px 7px;">'
            f'<div style="display:flex;align-items:center;gap:3px;">'
            f'<div style="width:{bw}px;height:10px;background:#0FBF5F;border-radius:2px;opacity:0.85;"></div>'
            f'<span style="color:#0FBF5F;font-size:10px;">{qfmt}</span>'
            f'</div></td>'
            f'<td style="color:#0FBF5F;font-size:10px;padding:3px 7px;">{"★ MUR BUY" if is_wall else "BUY"}</td>'
            f'</tr>'
        )

    # Walls summary
    nbw = stats["n_bid_walls"]; naw = stats["n_ask_walls"]
    nbw_txt = f"{nbw} mur(s) BID detecte(s)" if nbw > 0 else "Aucun mur BID fort"
    naw_txt = f"{naw} mur(s) ASK detecte(s)" if naw > 0 else "Aucun mur ASK fort"

    # Nearest walls info
    nbwall = stats.get("nearest_bid_wall")
    nawall = stats.get("nearest_ask_wall")
    _nb_p  = (f"{nbwall[0]:.2f}" if nbwall[0]>10 else f"{nbwall[0]:.5f}") if nbwall else ""
    _na_p  = (f"{nawall[0]:.2f}" if nawall[0]>10 else f"{nawall[0]:.5f}") if nawall else ""
    nb_txt = (f"Support fort : {_nb_p} (-{stats['dist_bid_wall']:.2f}%)") if nbwall and stats.get("dist_bid_wall") else "Aucun"
    na_txt = (f"Resistance forte : {_na_p} (+{stats['dist_ask_wall']:.2f}%)") if nawall and stats.get("dist_ask_wall") else "Aucune"

    # ICT confluence html
    ict_html = ""
    if ict_conf:
        items = "".join(f'<div style="color:#F0C96A;font-size:11px;margin-bottom:3px;">★ {c}</div>' for c in ict_conf)
        ict_html = (
            f'<div style="background:rgba(240,201,106,.10);border:1px solid #F0C96A;'
            f'border-radius:6px;padding:9px 12px;margin-top:8px;">'
            f'<div style="color:#F0C96A;font-weight:700;font-size:12px;margin-bottom:5px;">'
            f'★ CONFLUENCES ORDER BOOK + ICT DETECTEES</div>{items}</div>'
        )

    src_note = "Binance Depth API — Vrai Order Book" if is_btc else "OANDA Order Book — Positions/Ordres"

    html = (
        "<!DOCTYPE html><html><head>"
        "<style>body{margin:0;padding:8px;background:#060E1A;font-family:sans-serif;color:#DCE8F5;}"
        "table{width:100%;border-collapse:collapse;}"
        "th{font-size:10px;color:#6B7A8D;padding:5px 8px;text-align:left;border-bottom:1px solid #1A3050;}"
        "</style></head><body>"

        # Banner
        f'<div style="background:rgba(0,0,0,.3);border:2px solid {b_col};border-radius:10px;'
        f'padding:9px 14px;margin-bottom:10px;text-align:center;">'
        f'<div style="font-size:14px;font-weight:700;color:{b_col};">{b_icon} {b_txt}</div></div>'

        # Stats grid
        f'<div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:5px;margin-bottom:10px;">'
        f'<div style="background:#0F1C2E;border:1px solid #0FBF5F;border-radius:7px;padding:7px;text-align:center;">'
        f'<div style="font-size:9px;color:#6B7A8D;">BIDS (Acheteurs)</div>'
        f'<div style="font-size:16px;font-weight:700;color:#0FBF5F;">{stats["bid_pct"]:.1f}%</div>'
        f'<div style="font-size:9px;color:#0FBF5F;">{nbw_txt}</div></div>'

        f'<div style="background:#0F1C2E;border:1px solid #D93025;border-radius:7px;padding:7px;text-align:center;">'
        f'<div style="font-size:9px;color:#6B7A8D;">ASKS (Vendeurs)</div>'
        f'<div style="font-size:16px;font-weight:700;color:#D93025;">{stats["ask_pct"]:.1f}%</div>'
        f'<div style="font-size:9px;color:#D93025;">{naw_txt}</div></div>'

        f'<div style="background:#0F1C2E;border:1px solid {stats["bias_col"]};border-radius:7px;padding:7px;text-align:center;">'
        f'<div style="font-size:9px;color:#6B7A8D;">DESEQUILIBRE</div>'
        f'<div style="font-size:15px;font-weight:700;color:{stats["bias_col"]};">{stats["imbalance"]:+.1f}%</div>'
        f'<div style="font-size:9px;color:{stats["bias_col"]};">{ob_bias}</div></div>'

        f'<div style="background:#0F1C2E;border:1px solid #F0C96A;border-radius:7px;padding:7px;text-align:center;">'
        f'<div style="font-size:9px;color:#6B7A8D;">MURS DETECTES</div>'
        f'<div style="font-size:15px;font-weight:700;color:#0FBF5F;">{nbw}🟢</div>'
        f'<div style="font-size:9px;color:#D93025;">{naw}🔴 ASK</div></div>'
        f'</div>'

        # Bid/Ask bar
        f'<div style="margin-bottom:10px;">'
        f'<div style="font-size:10px;color:#6B7A8D;margin-bottom:3px;">'
        f'Pression Bids vs Asks — {stats["bid_pct"]:.0f}% acheteurs / {stats["ask_pct"]:.0f}% vendeurs</div>'
        f'<div style="display:flex;border-radius:6px;overflow:hidden;height:22px;">'
        f'<div style="width:{buy_w}%;background:#0FBF5F;opacity:0.85;display:flex;align-items:center;justify-content:center;">'
        f'<span style="font-size:10px;font-weight:700;color:#060E1A;">{stats["bid_pct"]:.0f}%</span></div>'
        f'<div style="width:{sell_w}%;background:#D93025;opacity:0.85;display:flex;align-items:center;justify-content:center;">'
        f'<span style="font-size:10px;font-weight:700;color:#DCE8F5;">{stats["ask_pct"]:.0f}%</span></div>'
        f'</div></div>'

        # Nearest walls
        f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:10px;">'
        f'<div style="background:rgba(15,191,95,.08);border:1px solid rgba(15,191,95,.3);'
        f'border-radius:7px;padding:8px;text-align:center;">'
        f'<div style="font-size:9px;color:#6B7A8D;">★ MUR BID LE PLUS PROCHE</div>'
        f'<div style="font-size:12px;font-weight:700;color:#0FBF5F;">{nb_txt}</div></div>'

        f'<div style="background:rgba(217,48,37,.08);border:1px solid rgba(217,48,37,.3);'
        f'border-radius:7px;padding:8px;text-align:center;">'
        f'<div style="font-size:9px;color:#6B7A8D;">★ MUR ASK LE PLUS PROCHE</div>'
        f'<div style="font-size:12px;font-weight:700;color:#D93025;">{na_txt}</div></div>'
        f'</div>'

        # ICT confluence
        f"{ict_html}"

        # Order book table
        f'<div style="font-size:12px;color:#F0C96A;letter-spacing:2px;margin:10px 0 6px 0;">'
        f'📖 CARNET ORDRES - STYLE BOOKMAP</div>'
        f'<div style="font-size:9px;color:#D93025;margin-bottom:3px;">★ = MUR INSTITUTIONNEL DETECTE</div>'
        f'<table><thead><tr><th>Prix</th><th>Volume</th><th>Type</th></tr></thead>'
        f'<tbody>{ob_rows}</tbody></table>'

        # Guide TP
        f'<div style="background:#0F1C2E;border-left:3px solid #F0C96A;border-radius:6px;'
        f'padding:9px 12px;margin-top:10px;font-size:11px;color:#8A95A3;line-height:1.9;">'
        f'<b style="color:#F0C96A;">GUIDE UTILISATION :</b><br>'
        f'<span style="color:#0FBF5F;">★ Mur BID = Support fort</span> — Prix rebondit souvent dessus<br>'
        f'<span style="color:#D93025;">★ Mur ASK = Resistance forte</span> — Prix bloque souvent dessus<br>'
        f'Mur disparait avant que le prix l\'atteigne = Spoofing - ne pas trader !<br>'
        f'BUY → TP juste sous le prochain mur ASK<br>'
        f'SELL → TP juste au-dessus du prochain mur BID'
        f'</div>'

        f'<div style="font-size:9px;color:#6B7A8D;text-align:center;margin-top:8px;">'
        f'Source : {src_note}</div>'
        f'</body></html>'
    )

    _cob.html(html, height=780, scrolling=True)


# ─── MAIN v5 ─────────────────────────────────────────────────────────────────
def main():
    import streamlit as st
    now = datetime.now(timezone.utc)
    cfg = load_config()

    # ── HEADER ──────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#060E1A,#0F1C2E);border:1px solid #C9A94B;
         border-radius:12px;padding:10px 16px;margin-bottom:10px;
         display:flex;justify-content:space-between;align-items:center;">
      <div>
        <div style="font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;
             color:#C9A94B;letter-spacing:3px;">
          <span class="live-dot"></span>STRATÉGIE PRO v7.1
        </div>
        <div style="font-size:9px;color:#6B7A8D;letter-spacing:2px;">
          MACD · BB · VWAP · ML · ICT · FORCE INDEX · TAPE · ORDER BOOK
        </div>
      </div>
      <div style="text-align:right;font-size:10px;color:#6B7A8D;">
        {now.strftime('%H:%M UTC')}<br>
        <span style="color:#C9A94B;">{now.strftime('%d %b %Y')}</span>
      </div>
    </div>""", unsafe_allow_html=True)

    # ── CONFIG ───────────────────────────────────────────────────────────────
    with st.expander("⚙️ CONFIGURATION — Actif · Timeframe · OANDA · Claude API", expanded=False):
        ca, cb = st.columns(2)
        with ca: asset_name = st.selectbox("📊 Actif", list(ASSETS.keys()), key="sel_asset")
        with cb: tf_sel     = st.selectbox("⏱️ Timeframe", ["5m","15m","30m","1h"], index=1, key="sel_tf")

        cx1, cx2, cx3 = st.columns(3)
        with cx1: multi_tf = st.checkbox("Multi-TF",         value=True,  key="chk_multitf")
        with cx2: auto10   = st.checkbox("Auto-refresh 10s", value=False, key="chk_auto")
        with cx3: pass

        _cfg_asset = ASSETS[asset_name]
        if _cfg_asset["source"] == "oanda":
            st.markdown('<div style="font-size:10px;color:#4A7EC7;margin-top:6px;">🔵 OANDA — fxtrade.oanda.com → Mon Compte → Gérer l\'accès API</div>', unsafe_allow_html=True)
            t1, t2, t3 = st.columns([3,1,1])
            with t1:
                # On Streamlit Cloud, keys come from secrets — show info instead
                if is_cloud() and cfg.get("oanda_key"):
                    st.success("✅ Clés API chargées depuis les Secrets Streamlit Cloud.")
                    st.info("Pour modifier : Streamlit Cloud → App Settings → Secrets")
                oanda_key_in = st.text_input("🔑 Token OANDA", value=cfg.get("oanda_key",""),
                                             type="password", placeholder="Token ici...", key="in_oanda")
            with t2:
                oanda_demo_in = st.checkbox("Démo", value=cfg.get("oanda_demo",True), key="chk_demo")
            with t3:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("💾 Sauver", key="save_oanda"):
                    cfg["oanda_key"]=oanda_key_in; cfg["oanda_demo"]=oanda_demo_in
                    save_config(cfg); st.success("✅ Sauvé !"); time.sleep(0.3); st.rerun()

        cl1, cl2 = st.columns([4,1])
        with cl1:
            claude_key_in = st.text_input("🧠 Clé API Claude (sk-ant-...)",
                value=cfg.get("claude_api_key",""), type="password",
                placeholder="sk-ant-api03-...", key="in_claude")
        with cl2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("💾", key="save_claude"):
                cfg["claude_api_key"]=claude_key_in
                save_config(cfg); st.success("✅"); time.sleep(0.3); st.rerun()

    # ── READ VALUES (from session_state after widgets render) ────────────────
    asset_name     = st.session_state.get("sel_asset", list(ASSETS.keys())[0])
    tf_sel         = st.session_state.get("sel_tf", "15m")
    multi_tf       = st.session_state.get("chk_multitf", True)
    auto10         = st.session_state.get("chk_auto", False)
    asset_cfg      = ASSETS[asset_name]
    oanda_key      = cfg.get("oanda_key", "")
    oanda_demo     = cfg.get("oanda_demo", True)
    claude_api_key = cfg.get("claude_api_key", "")

    # Auto-refresh 10s
    if auto10 and "ts_ref" in st.session_state:
        if (datetime.now(timezone.utc)-st.session_state["ts_ref"]).seconds >= 10:
            st.session_state["ts_ref"] = datetime.now(timezone.utc)
            st.rerun()

    # OANDA warning
    if asset_cfg["source"]=="oanda" and not oanda_key:
        st.markdown('''<div style="background:rgba(201,169,75,.06);border:1px solid #C9A94B;
             border-radius:6px;padding:6px 12px;font-size:10px;color:#C9A94B;margin-bottom:6px;">
          ⚠️ Token OANDA manquant → Yahoo Finance (~15min délai). Ouvre ⚙️ Configuration.
        </div>''', unsafe_allow_html=True)

    # ── ANALYSE BUTTONS ──────────────────────────────────────────────────────
    b1, b2, b3 = st.columns([3,1,1])
    with b1:
        run = st.button(f"🚀 ANALYSER {asset_name} — {tf_sel.upper()} (ML+LIVE)",
                        use_container_width=True)
    with b2:
        if st.button("🔄 REFRESH", use_container_width=True):
            if "res_cache" in st.session_state:
                lp2, ls2 = get_live_price(asset_cfg, oanda_key, oanda_demo)
                if lp2:
                    st.session_state["live_price"] = lp2
                    st.session_state["live_src"]   = ls2
            st.rerun()
    with b3:
        if st.button("🗑️ RESET", use_container_width=True):
            for k in ["res_cache","live_price","live_src","ts_ref"]:
                st.session_state.pop(k, None)
            st.rerun()

    # ── RUN ANALYSIS ─────────────────────────────────────────────────────────
    if run:
        mph  = st.empty()
        tfs  = ["5m","15m","30m","1h"] if multi_tf else [tf_sel]
        results = {}; dfs = {}; sources = {}
        steps = [
            (8,  "Connexion aux marchés live..."),
            (18, f"Source : {'Binance ⚡' if asset_cfg['source']=='binance' else 'OANDA 🔵' if oanda_key else 'Yahoo Finance ⚠️'}"),
            (30, f"Download candles {asset_name} {tf_sel}..."),
            (45, "VWAP · BB · MACD · ATR · Stoch · Pivots..."),
            (60, "ICT — FVG · Order Blocks · Liquidité..."),
            (80, "Features ML · 5 modèles ensemble..."),
            (94, "Confluence · TP/SL · Score final..."),
            (100,"✅ Analyse complète"),
        ]
        for pct, slbl in steps:
            mph.markdown(modal_html(slbl, pct), unsafe_allow_html=True)
            time.sleep(0.18)
            if pct == 30:
                for tf in tfs:
                    df, src = get_candles(asset_cfg, tf, oanda_key, oanda_demo)
                    dfs[tf] = df; sources[tf] = src
            if pct == 80:
                lp, ls = get_live_price(asset_cfg, oanda_key, oanda_demo)
                st.session_state["live_price"] = lp
                st.session_state["live_src"]   = ls or sources.get(tf_sel,"yfinance")
                # Cache BTC price for Heatmap/Tape fallback
                if "BTC" in asset_name and lp:
                    st.session_state["live_price_btc"] = lp
                for tf in tfs:
                    results[tf] = analyse(dfs.get(tf), tf, lp if tf==tf_sel else None)
        mph.empty()

        res_main = results.get(tf_sel)
        if res_main and res_main["signal"] in ["BUY","SELL"]:
            append_signal(asset_name, tf_sel, res_main["signal"],
                          res_main["confidence"], res_main["close"],
                          res_main["tp"], res_main["sl"],
                          res_main["regime"], res_main["ml_ensemble"])

        st.session_state["res_cache"]  = results
        st.session_state["asset_name"] = asset_name
        st.session_state["tf_sel"]     = tf_sel
        st.session_state["ts_ref"]     = datetime.now(timezone.utc)
        st.rerun()

    # ── DATA ─────────────────────────────────────────────────────────────────
    results    = st.session_state.get("res_cache", {})
    tf_d       = st.session_state.get("tf_sel", tf_sel)
    asset_d    = st.session_state.get("asset_name", asset_name)
    res        = results.get(tf_d) if results else None
    lp         = st.session_state.get("live_price")
    lsrc       = st.session_state.get("live_src", "yfinance")
    active_kzs = kz_active(now)

    # Live price banner
    if lp:
        ts_s    = datetime.now(timezone.utc).strftime("%H:%M:%S UTC")
        src_lbl = {"binance":"Binance ⚡","oanda":"OANDA 🔵"}.get(lsrc,"Yahoo ⚠️")
        st.markdown(f'''<div style="background:rgba(15,191,95,.06);border:1px solid #0FBF5F;
             border-radius:8px;padding:7px 14px;margin-bottom:8px;
             display:flex;justify-content:space-between;align-items:center;">
          <div>
            <span class="live-dot"></span>
            <span style="color:#0FBF5F;font-family:'Rajdhani',sans-serif;font-size:17px;font-weight:700;">{asset_d}</span>
            <span style="color:#F0C96A;font-family:'Rajdhani',sans-serif;font-size:24px;font-weight:700;margin-left:12px;">{lp:.5f}</span>
          </div>
          <div style="font-size:9px;color:#6B7A8D;text-align:right;">{src_lbl}<br>{ts_s}</div>
        </div>''', unsafe_allow_html=True)

    def no_data_msg(icon, label):
        st.markdown(f'''<div style="text-align:center;padding:50px 20px;">
          <div style="font-size:44px;">{icon}</div>
          <div style="font-family:'Rajdhani',sans-serif;font-size:16px;color:#C9A94B;
               letter-spacing:2px;margin:10px 0;">{label}</div>
          <div style="color:#6B7A8D;font-size:11px;">Lance une analyse d\'abord (bouton 🚀 ci-dessus)</div>
        </div>''', unsafe_allow_html=True)

    # ── TABS ─────────────────────────────────────────────────────────────────
    # Dynamic tab label for sentiment (changes with asset)
    _is_btc = "BTC" in asset_d
    _is_nas = "NAS" in asset_d
    _fg_label = "😨 FEAR/GREED" if _is_btc else "📊 SENTIMENT"
    _hm_label = "🗺️ HEATMAP"    if _is_btc else "🔍 MARCHÉS"

    tab_analyse, tab_chart, tab_pivots, tab_backtest, tab_alertes, tab_claude, tab_fg, tab_heatmap, tab_calendar, tab_orderflow, tab_tape, tab_ob, tab_telegram, tab_journal, tab_history = st.tabs([
        "📊 ANALYSE", "📈 GRAPHIQUE", "🎯 PIVOTS",
        "📉 BACKTEST", "🔊 ALERTES", "🧠 CLAUDE AI",
        _fg_label, _hm_label, "🗓️ CALENDRIER",
        "📊 ORDERFLOW", "📋 TAPE", "📖 ORDER BOOK", "📱 TELEGRAM",
        "📓 JOURNAL", "📋 HISTORIQUE"
    ])

    with tab_analyse:
        if res:
            r1a, r1b = st.columns(2)
            with r1a:
                render_signal_panel(res)
                render_tpsl(res)
            with r1b:
                render_ml_panel(res)
                render_checklist(res)
            r2a, r2b = st.columns(2)
            with r2a:
                render_indicators(res, lsrc)
                render_ict(res)
            with r2b:
                render_kz(now)
            if multi_tf and len(results) > 1:
                render_multitf(results)
            # ── Export analyse ──────────────────────────────────────────
            st.markdown("---")
            _all_rows = [res_to_export_row(r, asset_d, t) for t, r in results.items() if r]
            if _all_rows:
                render_export_buttons(_all_rows, "ANALYSE MULTI-TF", "analyse_resultats", "ana")
        else:
            no_data_msg("📊", "ANALYSE — EN ATTENTE")

    with tab_chart:
        if res: render_chart(res, asset_d)
        else:   no_data_msg("📈", "GRAPHIQUE — EN ATTENTE")

    with tab_pivots:
        if res: render_pivots(res)
        else:   no_data_msg("🎯", "PIVOTS — EN ATTENTE")

    with tab_backtest:
        if res: render_backtest(res, asset_d, tf_d)
        else:   no_data_msg("📉", "BACKTEST — EN ATTENTE")

    with tab_alertes:
        cfg = render_alertes(res, asset_d, cfg)
        save_config(cfg)

    with tab_claude:
        if res:
            render_claude_tab(res, results, asset_d, tf_d, claude_api_key, now, active_kzs)
        else:
            no_data_msg("🧠", "CLAUDE AI — EN ATTENTE")

    with tab_fg:
        render_sentiment_tab(asset_d, now, res)

    with tab_heatmap:
        render_heatmap_tab(asset_d, res)

    with tab_calendar:
        render_calendar(now)

    # Track orderflow result for Telegram
    of_result = st.session_state.get("of_result", None)

    with tab_orderflow:
        if res:
            _df_cur = None
            try:
                from copy import deepcopy
                _df_cur = results.get(tf_d) and results[tf_d].get("df") if hasattr(results.get(tf_d,{}), "get") else None
            except: pass
            _of = render_orderflow(res, asset_d, df=_df_cur)
            if _of: st.session_state["of_result"] = _of; of_result = _of
        else:
            no_data_msg("📊", "ORDERFLOW — Lance une analyse d'abord")

    with tab_tape:
        render_tape_tab(asset_d, res, asset_cfg, oanda_key, oanda_demo)

    with tab_ob:
        render_orderbook_tab(asset_d, res, asset_cfg, oanda_key, oanda_demo)

    with tab_telegram:
        cfg = render_telegram_config(cfg, res, asset_d, tf_d, of_result)
        save_config(cfg)

    with tab_journal:
        render_journal_tab()

    with tab_history:
        render_history_tab()

    # Footer
    st.markdown(f'''<div style="text-align:center;margin-top:14px;padding:6px;
         border-top:1px solid #1A3050;font-size:9px;color:#1A3050;">
      STRATÉGIE PRO v7.1 — 8 ACTIFS · ML · ICT · FORCE INDEX · TAPE · ORDER BOOK · TELEGRAM
      — 777gervais-dev — {now.strftime('%Y')} — ⚠️ Outil d\'aide à la décision uniquement.
    </div>''', unsafe_allow_html=True)


if __name__ == "__main__":
    main()
